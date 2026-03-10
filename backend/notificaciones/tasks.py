from celery import shared_task
from django.utils import timezone
from datetime import timedelta
from django.conf import settings
from capacitaciones.models import progresoCapacitaciones, Capacitaciones
from capacitaciones.utils import enviar_correo_batch
import time
import logging

logger = logging.getLogger(__name__)



def enviar_correos_por_lotes(destinatarios_bcc, subject, text_message, html_message, delay=None):
    """
    Función auxiliar para enviar correos masivos respetando el límite de GoDaddy.
    Usa la función mejorada de capacitaciones.utils.enviar_correo_batch con
    rate-limiting automático (450 emails/hora).

    Args:
        destinatarios_bcc (list): Lista de correos a enviar
        subject (str): Asunto del correo
        text_message (str): Contenido de texto
        html_message (str): Contenido HTML
        delay (int|None): Segundos entre lotes (None = calculado automático por rate-limit)

    Returns:
        dict: Estadísticas de envío {enviados, fallidos, total, tasa_exito}
    """
    if not destinatarios_bcc:
        logger.warning(f"enviar_correos_por_lotes: Lista de correos vacía")
        return {"enviados": 0, "fallidos": 0, "total": 0, "tasa_exito": 0}

    try:
        logger.info(f"enviar_correos_por_lotes: Enviando a {len(destinatarios_bcc)} colaboradores")
        # Usa rate-limiting automático para respetar 500 emails/hora de GoDaddy
        enviados, fallidos, errores = enviar_correo_batch(
            correos=destinatarios_bcc,
            subject=subject,
            text_message=text_message,
            html_message=html_message,
            delay_entre_lotes=delay,
            fail_silently=False
        )
        
        total = len(destinatarios_bcc)
        tasa_exito = (enviados / total * 100) if total > 0 else 0
        
        resultado = {
            "enviados": enviados,
            "fallidos": fallidos,
            "total": total,
            "tasa_exito": tasa_exito,
            "errores": errores
        }
        
        logger.info(f"enviar_correos_por_lotes: Resultado - {resultado}")
        return resultado
        
    except Exception as e:
        logger.error(f"enviar_correos_por_lotes: Error - {str(e)}", exc_info=True)
        return {
            "enviados": 0,
            "fallidos": len(destinatarios_bcc),
            "total": len(destinatarios_bcc),
            "tasa_exito": 0,
            "errores": [str(e)]
        }


@shared_task
def enviar_correo_capacitaciones_activas_y_activar():
    """
    Activa capacitaciones que inician hoy Y envía correos a colaboradores.
    Combina dos tareas: activación + notificación con batching automático.
    Se ejecuta cada día a las 12:00.
    
    IMPORTANTE: Filtra colaboradores DESACTIVADOS (estadocolaborador != 1)
    """
    hoy = timezone.now().date()
    
    capacitaciones_a_activar = Capacitaciones.objects.filter(
        fecha_inicio__date=hoy,
        estado=0
    )

    for cap in capacitaciones_a_activar:
        if getattr(cap, 'estado', None) != 0:
            continue
        cap.estado = 1
        cap.save()
        
        # FILTRO: Solo colaboradores ACTIVOS (estadocolaborador=1)
        correos = list(
            progresoCapacitaciones.objects.filter(
                capacitacion=cap,
                colaborador__estadocolaborador=1  # FILTRO: Solo activos
            )
            .values_list("colaborador__correocolaborador", flat=True)
            .exclude(colaborador__correocolaborador__isnull=True)
            .exclude(colaborador__correocolaborador__exact="")
            .distinct()
        )

        if not correos:
            continue

        subject = f"🎓 Nueva Capacitación Activada: {cap.titulo}"

        text_message = (
            f"Estimado colaborador,\n\n"
            f"Reciba un cordial saludo.\n"
            f"Nos complace informarle que ha sido matriculado en la formación '{cap.titulo}'.\n\n"
            f"Fecha de inicio: {cap.fecha_inicio.date()}\n"
            f"Fecha de finalización: {cap.fecha_fin.date()}\n\n"
            f"Podrá acceder a la plataforma en: https://formacion.cloudregencyapps.com/login\n\n"
            f"Agradecemos su disposición e interés en fortalecer sus competencias.\n"
            f"Atentamente,\n\n"
            f"Área de Formación Empresarial"
        )

        html_message = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <p>Estimado colaborador,</p>
            <p>Reciba un cordial saludo.</p>
            <p>
                Nos complace informarle que ha sido matriculado en la formación
                <strong>{cap.titulo}</strong>. A continuación, encontrará los detalles:
            </p>
            <ul>
                <li><strong>Fecha de inicio:</strong> {cap.fecha_inicio.date()}</li>
                <li><strong>Fecha de finalización:</strong> {cap.fecha_fin.date()}</li>
            </ul>
            <p>
                Podrá acceder a la plataforma de formación a través del siguiente enlace:<br>
                <a href="https://formacion.cloudregencyapps.com/login" target="_blank">Acceder a la plataforma</a>
            </p>
            <p>
                Si olvidó su contraseña, puede restablecerla desde la plataforma.
            </p>
            <p>
                Agradecemos su disposición e interés en fortalecer sus competencias.<br>
                Le deseamos una experiencia de aprendizaje provechosa.
            </p>
            <p><strong>Atentamente,</strong><br>
            Área de Formación Empresarial</p>
        </body>
        </html>
        """

        # Batching con rate-limiting automático (respeta 450 emails/hora de GoDaddy)
        enviar_correos_por_lotes(
            destinatarios_bcc=correos,
            subject=subject,
            text_message=text_message,
            html_message=html_message
        )

        # Esperar 60s entre capacitaciones como buffer adicional
        time.sleep(60)


@shared_task
def notificar_capacitacion_por_vencer_7_dias():
    """
    Notifica sobre capacitaciones que vencen en 7 días.
    Usa batching automático con rate-limiting (450 emails/hora GoDaddy).
    Se ejecuta cada día a las 10:00.
    
    IMPORTANTE: Filtra colaboradores DESACTIVADOS (estadocolaborador != 1)
    """
    hoy = timezone.now().date()
    fecha_objetivo = hoy + timedelta(days=7)

    capacitaciones = Capacitaciones.objects.filter(
        fecha_fin__date=fecha_objetivo,
        estado=1  # FILTRO: Solo capacitaciones activas
    )

    for cap in capacitaciones:

        # FILTRO: Solo colaboradores ACTIVOS (estadocolaborador=1)
        pendientes = progresoCapacitaciones.objects.filter(
            capacitacion=cap,
            completada=False,
            colaborador__estadocolaborador=1  # FILTRO: Solo activos
        ).select_related("colaborador")

        correos = [
            p.colaborador.correocolaborador
            for p in pendientes
            if p.colaborador.correocolaborador
        ]

        if not correos:
            continue

        subject = f"⚠️ Capacitación próxima a finalizar: {cap.titulo}"

        text_message = (
            f"Estimado colaborador,\n\n"
            f"Le recordamos que la capacitación \"{cap.titulo}\" finalizará en 7 días.\n\n"
            f"Fecha de finalización: {cap.fecha_fin.date()}\n\n"
            f"Según nuestros registros, aún no ha completado esta formación.\n\n"
            f"Acceda a la plataforma: https://formacion.cloudregencyapps.com/login\n\n"
            f"Lo invitamos a ingresar a la plataforma y finalizarla lo antes posible.\n\n"
            f"Atentamente,\n"
            f"Área de Formación Empresarial"
        )

        html_message = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <p>Estimado colaborador,</p>
            <p>
                La capacitación <strong>{cap.titulo}</strong> se encuentra próxima a finalizar.
            </p>
            <ul>
                <li><strong>Fecha de finalización:</strong> {cap.fecha_fin.date()}</li>
                <li><strong>Estado:</strong> No completada</li>
            </ul>
            <p>
                Le recomendamos ingresar a la plataforma para completar sus actividades:
                <br>
                <a href="https://formacion.cloudregencyapps.com/login" target="_blank">Ir a la plataforma</a>
            </p>
            <p>
                <strong>Quedan solo 7 días.</strong>
            </p>
            <p>
                Atentamente,<br>
                Área de Formación Empresarial
            </p>
        </body>
        </html>
        """

        # Batching con rate-limiting automático (respeta 450 emails/hora de GoDaddy)
        enviar_correos_por_lotes(
            destinatarios_bcc=correos,
            subject=subject,
            text_message=text_message,
            html_message=html_message
        )

        # Esperar 60s entre capacitaciones como buffer adicional
        time.sleep(60)

@shared_task
def desactivar_capacitaciones():
    """
    Desactiva capacitaciones que vencen hoy.
    Se ejecuta cada día a las 23:59.
    """
    hoy = timezone.now().date()

    capacitaciones_a_desactivar = Capacitaciones.objects.filter(
        fecha_fin__date=hoy,
        estado=1
    )

    for cap in capacitaciones_a_desactivar:
        cap.estado = 0
        cap.save()

@shared_task
def notificar_capacitacion_por_vencer_1_dia():
    """
    Último aviso para capacitaciones que vencen mañana. 
    Usa batching automático para soportar 1500+ colaboradores.
    Se ejecuta cada día a las 07:30.
    
    IMPORTANTE: Filtra colaboradores DESACTIVADOS (estadocolaborador != 1)
    """
    hoy = timezone.now().date()
    fecha_objetivo = hoy + timedelta(days=1)

    capacitaciones = Capacitaciones.objects.filter(
        fecha_fin__date=fecha_objetivo,
        estado=1  # FILTRO: Solo capacitaciones activas
    )

    for cap in capacitaciones:

        # FILTRO: Solo colaboradores ACTIVOS (estadocolaborador=1)
        pendientes = progresoCapacitaciones.objects.filter(
            capacitacion=cap,
            completada=False,
            colaborador__estadocolaborador=1  # FILTRO: Solo activos
        ).select_related("colaborador")

        correos = [
            p.colaborador.correocolaborador
            for p in pendientes
            if p.colaborador.correocolaborador
        ]

        if not correos:
            continue

        subject = f"🚨 Último aviso: {cap.titulo} vence mañana"

        text_message = (
            f"Estimado colaborador,\n\n"
            f"Le informamos que mañana finaliza la capacitación \"{cap.titulo}\".\n\n"
            f"Aún aparece como NO completada en el sistema.\n\n"
            f"Fecha de finalización: {cap.fecha_fin.date()}\n\n"
            f"Acceda a la plataforma: https://formacion.cloudregencyapps.com/login\n\n"
            f"Le recomendamos completarla hoy mismo para evitar quedar como pendiente.\n\n"
            f"Atentamente,\n"
            f"Área de Formación Empresarial"
        )

        html_message = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <p>Estimado colaborador,</p>
            <p>
                La capacitación <strong>{cap.titulo}</strong> finaliza <strong>mañana</strong>.
            </p>
            <ul>
                <li><strong>Fecha de finalización:</strong> {cap.fecha_fin.date()}</li>
                <li><strong>Estado:</strong> Aún no completada</li>
            </ul>
            <p>
                Este es el <strong>último recordatorio</strong>.
            </p>
            <p>
                Acceda aquí y finalice su capacitación:
                <br>
                <a href="https://formacion.cloudregencyapps.com/login" target="_blank">Ir a la plataforma</a>
            </p>
            <p>
                Área de Formación Empresarial
            </p>
        </body>
        </html>
        """

        # Batching con rate-limiting automático (respeta 450 emails/hora de GoDaddy)
        enviar_correos_por_lotes(
            destinatarios_bcc=correos,
            subject=subject,
            text_message=text_message,
            html_message=html_message
        )

        # Esperar 60s entre capacitaciones como buffer adicional
        time.sleep(60)


@shared_task
def notificar_jefes_por_colaboradores_sin_progreso():
    """
    Notifica a los jefes de proyecto sobre el estado de capacitaciones de sus colaboradores.
    Envía un archivo Excel adjunto personalizado por proyecto con las columnas:
    Nombre Completo, Correo, Capacitación, Estado (Completado / No Completado).
    
    Se ejecuta cada lunes a las 18:00.
    
    LÓGICA DE FILTRADO:
    1. Solo capacitaciones activas (estado = 1)
    2. Solo proyectos activos (estadoproyecto = 1) con jefe asignado
    3. Solo centros operativos activos (estadocentrop = 1) del proyecto
    4. Solo colaboradores activos (estadocolaborador = 1) de esos centros
    5. Solo colaboradores ASIGNADOS a cada capacitación (tienen registro en progresoCapacitaciones)
    6. Del registro en progresoCapacitaciones se obtiene si está completada o no
    7. Solo notifica si hay al menos un colaborador con capacitación pendiente
    
    SALIDA:
    - Cada jefe recibe un Excel con solo SUS colaboradores del proyecto
    - Cada fila muestra: Nombre, Correo, Capacitación, Estado (Completado/No Completado)
    - No incluye colaboradores inactivos, ni capacitaciones desactivadas/eliminadas
    - No incluye colaboradores no asignados a las capacitaciones
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from django.core.mail import EmailMultiAlternatives
    from analitica.models import Proyecto, Centroop
    from usuarios.models import Colaboradores

    # Obtener capacitaciones activas
    capacitaciones_activas = list(Capacitaciones.objects.filter(estado=1))

    if not capacitaciones_activas:
        logger.info("notificar_jefes: No hay capacitaciones activas")
        return

    # Obtener proyectos activos con jefe asignado
    proyectos = Proyecto.objects.filter(
        estadoproyecto=1,
        idcolaborador__isnull=False
    ).select_related('idcolaborador')

    for proyecto in proyectos:
        jefe = proyecto.idcolaborador
        if not jefe or not jefe.correocolaborador:
            continue

        # Obtener centros operativos activos del proyecto
        centros = Centroop.objects.filter(
            id_proyecto=proyecto,
            estadocentrop=1
        )

        if not centros.exists():
            continue

        # Obtener colaboradores activos de esos centros
        colaboradores = Colaboradores.objects.filter(
            centroop__in=centros,
            estadocolaborador=1
        ).order_by('nombrecolaborador', 'apellidocolaborador')

        if not colaboradores.exists():
            continue

        # Construir filas del Excel: solo colaboradores ASIGNADOS a cada capacitación
        filas_excel = []
        tiene_pendientes = False

        for colaborador in colaboradores:
            for cap in capacitaciones_activas:
                # Solo incluir si el colaborador está ASIGNADO a esta capacitación
                progreso = progresoCapacitaciones.objects.filter(
                    capacitacion=cap,
                    colaborador=colaborador
                ).first()

                # Si no tiene registro en progresoCapacitaciones, no está asignado → saltar
                if not progreso:
                    continue

                if progreso.completada:
                    estado = "Completado"
                else:
                    estado = "No Completado"
                    tiene_pendientes = True

                nombre_completo = f"{colaborador.nombrecolaborador} {colaborador.apellidocolaborador}"
                correo = colaborador.correocolaborador or "Sin correo"

                filas_excel.append({
                    "nombre": nombre_completo,
                    "correo": correo,
                    "capacitacion": cap.titulo,
                    "estado": estado,
                })

        # Solo notificar si hay al menos un colaborador con capacitación pendiente
        if not tiene_pendientes or not filas_excel:
            continue

        # ── Generar archivo Excel ──
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Capacitaciones"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        completado_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        completado_font = Font(color="006100")
        no_completado_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        no_completado_font = Font(color="9C0006")

        # Encabezados
        headers = ["Nombre Completo", "Correo", "Capacitación", "Estado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Datos
        for row_idx, fila in enumerate(filas_excel, 2):
            ws.cell(row=row_idx, column=1, value=fila["nombre"]).border = thin_border
            ws.cell(row=row_idx, column=2, value=fila["correo"]).border = thin_border
            ws.cell(row=row_idx, column=3, value=fila["capacitacion"]).border = thin_border

            estado_cell = ws.cell(row=row_idx, column=4, value=fila["estado"])
            estado_cell.border = thin_border
            estado_cell.alignment = Alignment(horizontal="center")

            if fila["estado"] == "Completado":
                estado_cell.fill = completado_fill
                estado_cell.font = completado_font
            else:
                estado_cell.fill = no_completado_fill
                estado_cell.font = no_completado_font

        # Ancho de columnas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 18

        # Guardar en memoria
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # ── Estadísticas para el cuerpo del correo ──
        nombres_unicos = set(f["nombre"] for f in filas_excel)
        total_colab = len(nombres_unicos)
        total_pendientes = sum(1 for f in filas_excel if f["estado"] == "No Completado")
        total_completados = sum(1 for f in filas_excel if f["estado"] == "Completado")

        # ── Enviar correo con Excel adjunto ──
        subject = f"📊 Reporte Semanal de Capacitaciones - Proyecto: {proyecto.nombreproyecto}"

        text_message = (
            f"Estimado/a {jefe.nombrecolaborador},\n\n"
            f"Adjunto encontrará el reporte de capacitaciones del proyecto "
            f"'{proyecto.nombreproyecto}'.\n\n"
            f"Resumen:\n"
            f"- Total colaboradores: {total_colab}\n"
            f"- Capacitaciones completadas: {total_completados}\n"
            f"- Capacitaciones pendientes: {total_pendientes}\n\n"
            f"Le recomendamos realizar el seguimiento correspondiente para "
            f"garantizar el cumplimiento del proceso de formación.\n\n"
            f"Atentamente,\n"
            f"Área de Formación Empresarial"
        )

        html_message = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <p>Estimado/a <strong>{jefe.nombrecolaborador}</strong>,</p>
            <p>
                Adjunto encontrará el reporte de capacitaciones del proyecto
                <strong>{proyecto.nombreproyecto}</strong>.
            </p>
            <table style="border-collapse: collapse; margin: 15px 0;">
                <tr>
                    <td style="padding: 8px 15px; border: 1px solid #ddd; background: #f5f5f5;">
                        <strong>Total colaboradores</strong>
                    </td>
                    <td style="padding: 8px 15px; border: 1px solid #ddd; text-align: center;">
                        {total_colab}
                    </td>
                </tr>
                <tr>
                    <td style="padding: 8px 15px; border: 1px solid #ddd; background: #f5f5f5;">
                        <strong>Capacitaciones completadas</strong>
                    </td>
                    <td style="padding: 8px 15px; border: 1px solid #ddd; text-align: center; color: #006100;">
                        ✅ {total_completados}
                    </td>
                </tr>
                <tr>
                    <td style="padding: 8px 15px; border: 1px solid #ddd; background: #f5f5f5;">
                        <strong>Capacitaciones pendientes</strong>
                    </td>
                    <td style="padding: 8px 15px; border: 1px solid #ddd; text-align: center; color: #9C0006;">
                        ❌ {total_pendientes}
                    </td>
                </tr>
            </table>
            <p>
                Le recomendamos realizar el seguimiento correspondiente para
                garantizar el cumplimiento del proceso de formación.
            </p>
            <p>
                <em>El archivo Excel adjunto contiene el detalle por colaborador.</em>
            </p>
            <p>
                <strong>Atentamente,</strong><br>
                Área de Formación Empresarial
            </p>
        </body>
        </html>
        """

        # Crear email individual para cada jefe con Excel adjunto
        nombre_archivo = f"Reporte_Capacitaciones_{proyecto.nombreproyecto.replace(' ', '_')}.xlsx"

        try:
            email = EmailMultiAlternatives(
                subject=subject,
                body=text_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[jefe.correocolaborador],
            )
            email.attach_alternative(html_message, "text/html")
            email.attach(
                nombre_archivo,
                excel_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            email.send(fail_silently=False)

            logger.info(
                f"notificar_jefes: ✅ Reporte enviado a {jefe.correocolaborador} "
                f"(Proyecto: {proyecto.nombreproyecto}, "
                f"Colaboradores: {total_colab}, Pendientes: {total_pendientes})"
            )
        except Exception as e:
            logger.error(
                f"notificar_jefes: ❌ Error enviando a {jefe.correocolaborador} "
                f"(Proyecto: {proyecto.nombreproyecto}): {str(e)}",
                exc_info=True,
            )

        # Esperar entre envíos para evitar rate-limiting
        time.sleep(2)
