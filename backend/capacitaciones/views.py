# ==================== STANDARD LIBRARY ====================
import csv
import hashlib
import io
from io import BytesIO
import os
import shutil
import tempfile
import zipfile
import traceback

from django.conf import settings
import uuid

import unicodedata
import re
from pathlib import Path
# ==================== DJANGO ====================
from django.conf import settings
from django.core.cache import cache
from django.db import transaction
from django.db.models import Count, Max, Prefetch, OuterRef, Subquery
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone

# ==================== THIRD PARTY ====================
import cloudinary
import cloudinary.uploader
from docx import Document
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.exceptions import NotAcceptable

import logging
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Opcional: pypandoc para conversión DOCX a PDF
import pypandoc
import pypandoc as _pypandoc

# Logger para registrar eventos
logger = logging.getLogger(__name__)

# ==================== LOCAL ====================
from .models import (
    Capacitaciones,
    CertificadoGenerado,
    Lecciones,
    Modulos,
    PreguntasLecciones,
    Respuestas,
    RespuestasColaboradores,
    progresoCapacitaciones,
    progresolecciones,
)
from capacitaciones.serializers import (
    CapacitacionDetalleSerializer,
    CapacitacionProgresoSerializer,
    capacitacionSerializer,
    CrearCapacitacionSerializer,
    MisCapacitacionesSerializer,
)
from capacitaciones.utils import actualizar_progreso_leccion
from usuarios.models import Colaboradores
from usuarios.permissions import IsAdminUser, IsSuperAdmin, IsUsuarioEspecial, IsGestionEmpresarial


# ==================== HELPERS DE CACHE ====================
def get_cache_key(prefix, *args):
    """Genera una clave de cache única basada en prefijo y argumentos"""
    key_data = f"{prefix}:" + ":".join(str(arg) for arg in args)
    return hashlib.md5(key_data.encode()).hexdigest()


def invalidate_capacitacion_cache(capacitacion_id=None, colaborador_id=None):
    """Invalida caches relacionadas con capacitaciones"""
    if capacitacion_id:
        cache.delete(get_cache_key('cap_detail', capacitacion_id))
    if colaborador_id:
        cache.delete(get_cache_key('mis_caps', colaborador_id))
    # Invalidar lista general de capacitaciones
    cache.delete('capacitaciones_list_admin')


class CrearCapacitacionView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]
    """Crear o editar una capacitación (Solo Admin y SuperAdmin)
       GET with `capacitacion_id`: retorna datos para edición
       POST: crea nueva capacitación
       PATCH: actualiza/sincroniza una capacitación existente
    """

    def get(self, request, capacitacion_id, *args, **kwargs):
        """Obtener datos de la capacitación para edición (Admin only)"""
        try:
            capacitacion = Capacitaciones.objects.get(pk=capacitacion_id)
        except Capacitaciones.DoesNotExist:
            return Response({'error': 'Capacitación no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        serializer = CapacitacionDetalleSerializer(capacitacion)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @transaction.atomic
    def patch(self, request, capacitacion_id, *args, **kwargs):
        """Editar campos de la capacitación (Admin only) - Sin eliminar datos existentes, solo actualizar o agregar"""
        if not getattr(request.user, 'is_staff', False) and not getattr(request.user, 'is_superuser', False):
            return Response({'error': 'No tienes permiso para editar esta capacitación.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            capacitacion = Capacitaciones.objects.get(pk=capacitacion_id)
        except Capacitaciones.DoesNotExist:
            return Response({'error': 'Capacitación no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        # Obtener colaboradores actuales para invalidar cache
        current_collaborators = set(progresoCapacitaciones.objects.filter(capacitacion=capacitacion).values_list('colaborador_id', flat=True))

        archivos_creados = []
        
        try:
            from capacitaciones.utils import procesar_archivo_multimedia, limpiar_archivo
            import json
            
            # Extraer datos del request - soporta tanto multipart como JSON
            # IMPORTANTE: request.data en DRF mezcla archivos + texto,
            # request.POST solo tiene campos de texto (seguro para serializer)
            data = {}
            try:
                if hasattr(request, 'POST') and request.POST:
                    data = dict(request.POST.items())
                elif hasattr(request, 'data'):
                    data = dict(request.data)
                else:
                    data = {}
            except Exception:
                data = {}

            # Filtrar objetos archivo que puedan haber llegado desde request.data
            # (InMemoryUploadedFile, etc.) - estos se procesan desde request.FILES
            for k in list(data.keys()):
                val = data[k]
                if hasattr(val, 'read') or hasattr(val, 'chunks'):
                    del data[k]

            # Eliminar cadenas vacías (no relevantes para PATCH parcial)
            for k in list(data.keys()):
                if isinstance(data.get(k), str) and data[k] == '':
                    data.pop(k)

            # NO procesar colaboradores desde el request - se ignoran completamente
            if 'colaboradores' in data:
                data.pop('colaboradores')
            
            # Parsear modulos si vienen como JSON string
            if 'modulos' in data:
                try:
                    if isinstance(data['modulos'], str):
                        data['modulos'] = json.loads(data['modulos'])
                except (json.JSONDecodeError, ValueError) as e:
                    logger.error(f"Error al parsear módulos en PATCH: {str(e)}")
                    return Response(
                        {'error': 'Estructura de módulos inválida'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Procesar imagen principal si viene en request.FILES
            if 'imagen' in request.FILES:
                imagen_file = request.FILES['imagen']
                url_imagen, error = procesar_archivo_multimedia(
                    imagen_file, 'imagen'
                )
                if error:
                    return Response({'error': error}, status=status.HTTP_400_BAD_REQUEST)
                data['imagen'] = url_imagen
                archivos_creados.append(url_imagen)
            
            # Procesar archivos de lecciones y preguntas si vienen en multipart
            if request.FILES:
                modulos_data = data.get('modulos', [])
                if modulos_data and isinstance(modulos_data, list):
                    for modulo_idx, modulo_data in enumerate(modulos_data):
                        lecciones = modulo_data.get('lecciones', [])
                        
                        for leccion_idx, leccion_data in enumerate(lecciones):
                            # Procesar archivo de lección (video/PDF)
                            leccion_tipo = leccion_data.get('tipo_leccion', '').lower()
                            if leccion_tipo in ['video', 'pdf', 'imagen']:
                                file_key = f'leccion_{modulo_idx}_{leccion_idx}'
                                
                                if file_key in request.FILES:
                                    tipo_archivo = 'video' if leccion_tipo == 'video' else leccion_tipo
                                    url_leccion, error = procesar_archivo_multimedia(
                                        request.FILES[file_key], tipo_archivo
                                    )
                                    if error:
                                        for url in archivos_creados:
                                            limpiar_archivo(url)
                                        return Response({'error': error}, status=status.HTTP_400_BAD_REQUEST)
                                    
                                    leccion_data['url'] = url_leccion
                                    archivos_creados.append(url_leccion)
                            
                            # Procesar preguntas
                            preguntas = leccion_data.get('preguntas', [])
                            for pregunta_idx, pregunta_data in enumerate(preguntas):
                                # Procesar multimedia de pregunta
                                file_key_pregunta = f'pregunta_{modulo_idx}_{leccion_idx}_{pregunta_idx}'
                                if file_key_pregunta in request.FILES:
                                    url_pregunta, error = procesar_archivo_multimedia(
                                        request.FILES[file_key_pregunta], 'imagen'
                                    )
                                    if error:
                                        for url in archivos_creados:
                                            limpiar_archivo(url)
                                        return Response({'error': error}, status=status.HTTP_400_BAD_REQUEST)
                                    
                                    pregunta_data['url_multimedia'] = url_pregunta
                                    archivos_creados.append(url_pregunta)
                                
                                # Procesar respuestas
                                respuestas = pregunta_data.get('respuestas', [])
                                for respuesta_idx, respuesta_data in enumerate(respuestas):
                                    file_key_respuesta = f'respuesta_{modulo_idx}_{leccion_idx}_{pregunta_idx}_{respuesta_idx}'
                                    if file_key_respuesta in request.FILES:
                                        url_respuesta, error = procesar_archivo_multimedia(
                                            request.FILES[file_key_respuesta], 'imagen'
                                        )
                                        if error:
                                            for url in archivos_creados:
                                                limpiar_archivo(url)
                                            return Response({'error': error}, status=status.HTTP_400_BAD_REQUEST)
                                        
                                        respuesta_data['url_imagen'] = url_respuesta
                                        archivos_creados.append(url_respuesta)

            serializer = CrearCapacitacionSerializer(capacitacion, data=data, partial=True)
            if serializer.is_valid():
                serializer.save()

                # Invalidar cache para la capacitación y colaboradores afectados
                invalidate_capacitacion_cache(capacitacion_id=capacitacion.id)
                new_collaborators = set(progresoCapacitaciones.objects.filter(capacitacion=capacitacion).values_list('colaborador_id', flat=True))
                affected = current_collaborators | new_collaborators
                for cid in affected:
                    invalidate_capacitacion_cache(colaborador_id=cid)

                return Response(serializer.data, status=status.HTTP_200_OK)
            
            # Limpiar archivos si hay error en validación
            for url in archivos_creados:
                limpiar_archivo(url)
            
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            # Limpiar archivos en caso de error
            for url in archivos_creados:
                limpiar_archivo(url)
            logger.error(f"Error en PATCH capacitación: {str(e)}")
            return Response(
                {'error': f'Error al actualizar capacitación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @transaction.atomic
    def post(self, request, capacitacion_id=None, *args, **kwargs):
        """
        Crear nueva capacitación con soporte para archivos multimedia en el mismo POST.
        
        Soporta multipart/form-data con archivos incrustados.
        Los archivos se procesan durante la creación y se rollback en caso de error.
        
        Estructura del payload (multipart/form-data):
        - titulo: string
        - descripcion: string
        - imagen: file (opcional)
        - tipo: string
        - fecha_inicio: datetime
        - fecha_fin: datetime
        - modulos: JSON string (con estructura de módulos, lecciones y archivos)
        - colaboradores: JSON string (array de IDs)
        - archivo_leccion_<modulo_idx>_<leccion_idx>: file (opcional)
        - archivo_pregunta_<modulo_idx>_<leccion_idx>_<pregunta_idx>: file (opcional)
        - archivo_respuesta_<modulo_idx>_<leccion_idx>_<pregunta_idx>_<respuesta_idx>: file (opcional)
        """
        archivos_creados = []  # Rastrear archivos creados para limpiarlos en caso de error
        
        try:
            from capacitaciones.utils import procesar_archivo_multimedia, limpiar_archivo, ALLOWED_IMAGE_EXTENSIONS, ALLOWED_PDF_EXTENSION
            import json
            
            # Extraer datos del request
            data = {}
            try:
                # Usar POST data para campos simples
                data = dict(request.POST.items())
            except Exception:
                return Response(
                    {'error': 'Error al procesar datos del formulario'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Filtrar objetos archivo que puedan colarse en data
            for k in list(data.keys()):
                val = data[k]
                if hasattr(val, 'read') or hasattr(val, 'chunks'):
                    del data[k]
            
            # Extraer colaboradores
            colaboradores_ids = []
            if 'colaboradores' in data:
                try:
                    if isinstance(data['colaboradores'], str):
                        colaboradores_ids = json.loads(data.pop('colaboradores'))
                    else:
                        colaboradores_ids = data.pop('colaboradores')
                except (json.JSONDecodeError, ValueError):
                    colaboradores_ids = []
            
            # Parsear modulos
            modulos_data = []
            if 'modulos' in data:
                try:
                    modulos_data = json.loads(data.pop('modulos'))
                except (json.JSONDecodeError, ValueError) as e:
                    logger.error(f"Error al parsear módulos: {str(e)}")
                    return Response(
                        {'error': 'Estructura de módulos inválida'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Procesar imagen principal
            if 'imagen' in request.FILES:
                imagen_file = request.FILES['imagen']
                url_imagen, error = procesar_archivo_multimedia(
                    imagen_file, 'imagen'
                )
                if error:
                    return Response({'error': f'Error al procesar imagen: {error}'}, status=status.HTTP_400_BAD_REQUEST)
                data['imagen'] = url_imagen
                archivos_creados.append(url_imagen)
                logger.info(f"Imagen procesada exitosamente: {url_imagen}")
            elif not data.get('imagen'):
                # Solo vaciar si no vino una URL existente en el cuerpo del formulario
                data['imagen'] = ''
            
            # Procesar archivos de lecciones y preguntas
            for modulo_idx, modulo_data in enumerate(modulos_data):
                lecciones = modulo_data.get('lecciones', [])
                
                for leccion_idx, leccion_data in enumerate(lecciones):
                    # Procesar archivo de lección (video/PDF)
                    leccion_tipo = leccion_data.get('tipo_leccion', '').lower()
                    if leccion_tipo in ['video', 'pdf', 'imagen']:
                        file_key = f'leccion_{modulo_idx}_{leccion_idx}'
                        
                        if file_key in request.FILES:
                            tipo_archivo = 'video' if leccion_tipo == 'video' else leccion_tipo
                            url_leccion, error = procesar_archivo_multimedia(
                                request.FILES[file_key], tipo_archivo
                            )
                            if error:
                                # Limpiar archivos creados antes del error
                                for url in archivos_creados:
                                    limpiar_archivo(url)
                                return Response({'error': error}, status=status.HTTP_400_BAD_REQUEST)
                            
                            leccion_data['url'] = url_leccion
                            archivos_creados.append(url_leccion)
                    
                    # Procesar preguntas
                    preguntas = leccion_data.get('preguntas', [])
                    for pregunta_idx, pregunta_data in enumerate(preguntas):
                        # Procesar multimedia de pregunta
                        file_key_pregunta = f'pregunta_{modulo_idx}_{leccion_idx}_{pregunta_idx}'
                        if file_key_pregunta in request.FILES:
                            url_pregunta, error = procesar_archivo_multimedia(
                                request.FILES[file_key_pregunta], 'imagen'
                            )
                            if error:
                                for url in archivos_creados:
                                    limpiar_archivo(url)
                                return Response({'error': error}, status=status.HTTP_400_BAD_REQUEST)
                            
                            pregunta_data['url_multimedia'] = url_pregunta
                            archivos_creados.append(url_pregunta)
                        
                        # Procesar respuestas
                        respuestas = pregunta_data.get('respuestas', [])
                        for respuesta_idx, respuesta_data in enumerate(respuestas):
                            file_key_respuesta = f'respuesta_{modulo_idx}_{leccion_idx}_{pregunta_idx}_{respuesta_idx}'
                            if file_key_respuesta in request.FILES:
                                url_respuesta, error = procesar_archivo_multimedia(
                                    request.FILES[file_key_respuesta], 'imagen'
                                )
                                if error:
                                    for url in archivos_creados:
                                        limpiar_archivo(url)
                                    return Response({'error': error}, status=status.HTTP_400_BAD_REQUEST)
                                
                                respuesta_data['url_imagen'] = url_respuesta
                                archivos_creados.append(url_respuesta)
            
            # Agregar módulos nuevamente a los datos para el serializer
            data['modulos'] = modulos_data
            data['colaboradores'] = colaboradores_ids
            
            # Asegurar que imagen tiene un valor (requerido por BD)
            if 'imagen' not in data or not data.get('imagen'):
                data['imagen'] = ''
            
            # Crear capacitación con transacción
            serializer = CrearCapacitacionSerializer(data=data)
            if not serializer.is_valid():
                # Limpiar archivos si hay error en validación
                for url in archivos_creados:
                    limpiar_archivo(url)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            # Guardar capacitación (la transacción atómica se maneja en el serializer)
            capacitacion = serializer.save()
            
            # Procesar colaboradores si se proporcionaron
            if colaboradores_ids:
                colaboradores = Colaboradores.objects.filter(
                    idcolaborador__in=colaboradores_ids,
                    estadocolaborador=1
                )
                
                # Obtener IDs ya existentes para esta capacitación (evitar duplicados)
                existentes = set(
                    progresoCapacitaciones.objects.filter(
                        capacitacion=capacitacion
                    ).values_list('colaborador_id', flat=True)
                )
                
                progreso_records = [
                    progresoCapacitaciones(
                        capacitacion=capacitacion,
                        colaborador=colaborador,
                        completada=0,
                        progreso=0
                    )
                    for colaborador in colaboradores
                    if colaborador.idcolaborador not in existentes
                ]
                
                if progreso_records:
                    progresoCapacitaciones.objects.bulk_create(progreso_records)
                
                # Enviar correos
                try:
                    from capacitaciones.utils import enviar_correo_capacitacion_creada_batch
                    transaction.on_commit(lambda: enviar_correo_capacitacion_creada_batch(capacitacion))
                    logger.info(f"Correos programados para capacitación {capacitacion.id}")
                except Exception as e:
                    logger.error(f"Error al programar correos: {str(e)}", exc_info=True)
            
            # Invalidar cache
            cache.delete('capacitaciones_list_admin')
            
            return Response(
                {
                    'id': capacitacion.id,
                    'titulo': capacitacion.titulo,
                    'imagen': capacitacion.imagen,
                    'colaboradores_asignados': len(colaboradores_ids),
                    'archivos_procesados': len(archivos_creados)
                },
                status=status.HTTP_201_CREATED
            )
            
        except Exception as e:
            # En caso de excepción no controlada, limpiar archivos
            logger.error(f"Error al crear capacitación: {str(e)}", exc_info=True)
            for url in archivos_creados:
                try:
                    limpiar_archivo(url)
                except Exception as cleanup_error:
                    logger.error(f"Error al limpiar archivo {url}: {str(cleanup_error)}")
            
            return Response(
                {'error': f'Error al crear capacitación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CapacitacionesView(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdmin | IsAdminUser]
    """Lista todas las capacitaciones activas (Solo Admin y SuperUsuario)
    
    Optimización: 
    - Cache de 5 minutos para lista de capacitaciones
    - Select_related para evitar N+1 queries
    - Only() para cargar solo campos necesarios
    """
    
    def get(self, request, *args, **kwargs):
        try:
            # Intentar obtener de cache
            cache_key = 'capacitaciones_list_admin'
            cached_data = cache.get(cache_key)
            
            if cached_data is not None:
                return Response(cached_data, status=status.HTTP_200_OK)
            
            # Query optimizada: solo campos necesarios, ordenado por fecha
            capacitaciones = Capacitaciones.objects.exclude(
                estado=3
            ).only(
                'id', 'titulo', 'descripcion', 'estado',
                'fecha_creacion', 'fecha_inicio', 'fecha_fin', 'tipo'
            ).order_by('-fecha_creacion')
            
            serializer = capacitacionSerializer(capacitaciones, many=True)
            
            # Guardar en cache (5 minutos)
            cache_ttl = getattr(settings, 'CACHE_TTL_CAPACITACIONES_LIST', 300)
            cache.set(cache_key, serializer.data, cache_ttl)
            
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CapacitacionDetailView(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdmin | IsAdminUser]
    """
    Detalle de una capacitación específica con prefetch optimizado.
    
    Solo el usuario autenticado puede ver las capacitaciones en las que está inscrito.
    
    Optimización: 
    - Prefetch_related profundo para cargar toda la estructura (226 → 4 queries)
    - Cache por capacitación (10 minutos) - estructura no cambia frecuentemente
    """
    
    def get(self, request, capacitacion_id, *args, **kwargs):
        try:
            # Verificar que el usuario tiene colaborador asociado
            colaborador = getattr(request.user, 'idcolaboradoru', None)
            if not colaborador:
                return Response(
                    {'error': 'El usuario no tiene un colaborador asociado'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Verificar inscripción (query rápida con exists())
            inscrito = progresoCapacitaciones.objects.filter(
                colaborador=colaborador,
                capacitacion_id=capacitacion_id
            ).only('id').exists()
            
            if not inscrito:
                return Response(
                    {'error': 'No tienes acceso a esta capacitación. No estás inscrito.'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Intentar obtener de cache (la estructura de la capacitación no cambia)
            cache_key = get_cache_key('cap_detail', capacitacion_id)
            cached_data = cache.get(cache_key)
            
            if cached_data is not None:
                return Response(cached_data, status=status.HTTP_200_OK)
            
            # Prefetch profundo de toda la estructura (optimizado)
            capacitacion = Capacitaciones.objects.prefetch_related(
                Prefetch(
                    'modulos_set',
                    queryset=Modulos.objects.prefetch_related(
                        Prefetch(
                            'lecciones_set',
                            queryset=Lecciones.objects.prefetch_related(
                                Prefetch(
                                    'preguntaslecciones_set',
                                    queryset=PreguntasLecciones.objects.prefetch_related(
                                        'respuestas_set'
                                    )
                                )
                            ).order_by('id')
                        )
                    ).order_by('id')
                )
            ).get(pk=capacitacion_id)
            
            serializer = CapacitacionDetalleSerializer(capacitacion)
            
            # Guardar en cache (10 minutos)
            cache_ttl = getattr(settings, 'CACHE_TTL_CAPACITACION_DETAIL', 600)
            cache.set(cache_key, serializer.data, cache_ttl)
            
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Capacitaciones.DoesNotExist:
            return Response(
                {'error': 'Capacitación no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RegistrarProgresoView(APIView):
    permission_classes = [IsAuthenticated]
    """Registrar progreso en una lección y actualizar progreso de módulo y capacitación
    
    Optimización:
    - Select_related para obtener módulo y capacitación en una sola query
    - Invalidación de cache del colaborador al actualizar progreso
    """
    
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            # Obtener colaborador desde el usuario autenticado
            if not request.user.idcolaboradoru:
                return Response(
                    {'error': 'El usuario no tiene un colaborador asociado'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            colaborador = request.user.idcolaboradoru
            leccion_id = request.data.get('leccion_id')
            progreso = request.data.get('progreso', 0)
            
            if not leccion_id:
                return Response(
                    {'error': 'leccion_id es requerido'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar que el progreso sea un número entre 0 y 100
            try:
                progreso = max(0, min(100, float(progreso)))
            except (ValueError, TypeError):
                progreso = 0
            
            # Determinar completada basado en el progreso real (no confiar en el frontend)
            completada = progreso >= 100
            
            # Select_related para obtener módulo y capacitación en una query
            leccion = Lecciones.objects.select_related(
                'idmodulo__idcapacitacion'
            ).get(id=leccion_id)
            
            # Usar la función de utils que actualiza toda la cadena de progreso
            progreso_data = actualizar_progreso_leccion(
                colaborador_id=colaborador.idcolaborador,
                leccion=leccion,
                progreso=progreso,
                completada=completada
            )
            
            # Invalidar cache del colaborador
            invalidate_capacitacion_cache(colaborador_id=colaborador.idcolaborador)
            
            return Response(
                {
                    'mensaje': 'Progreso actualizado exitosamente',
                    'leccion_id': leccion_id,
                    'progreso_leccion': progreso,
                    'completada': completada,
                    'progreso_modulo': progreso_data.get('progreso_modulo', 0),
                    'progreso_capacitacion': progreso_data.get('progreso_capacitacion', 0)
                },
                status=status.HTTP_200_OK
            )
        except Lecciones.DoesNotExist:
            return Response(
                {'error': 'Lección no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CompletarLeccionView(APIView):
    permission_classes = [IsAuthenticated]
    """Marcar una lección como completada y actualizar progreso de módulo y capacitación
    
    Optimización:
    - Select_related para obtener módulo y capacitación en una sola query
    - Verificación de inscripción optimizada con only()
    - Invalidación de cache al completar
    """
    
    @transaction.atomic
    def post(self, request, leccion_id, *args, **kwargs):
        try:
            # Obtener colaborador desde el usuario autenticado
            if not request.user.idcolaboradoru:
                return Response(
                    {'error': 'El usuario no tiene un colaborador asociado'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            colaborador = request.user.idcolaboradoru
            
            # Select_related para obtener módulo y capacitación en una query
            leccion = Lecciones.objects.select_related(
                'idmodulo__idcapacitacion'
            ).get(id=leccion_id)
            
            # Verificar inscripción (query optimizada)
            modulo = leccion.idmodulo
            capacitacion = modulo.idcapacitacion
            
            inscrito = progresoCapacitaciones.objects.filter(
                colaborador=colaborador,
                capacitacion=capacitacion
            ).only('id').exists()
            
            if not inscrito:
                return Response(
                    {'error': 'No tienes acceso a esta lección. No estás inscrito en la capacitación correspondiente.'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Usar la función de utils que actualiza toda la cadena de progreso
            progreso_data = actualizar_progreso_leccion(
                colaborador_id=colaborador.idcolaborador,
                leccion=leccion,
                progreso=100,
                completada=True
            )
            
            # Invalidar cache del colaborador
            invalidate_capacitacion_cache(colaborador_id=colaborador.idcolaborador)
            
            return Response(
                {
                    'mensaje': 'Lección completada exitosamente',
                    'leccion_id': leccion_id,
                    'colaborador_id': colaborador.idcolaborador,
                    'progreso_modulo': progreso_data.get('progreso_modulo', 0),
                    'progreso_capacitacion': progreso_data.get('progreso_capacitacion', 0)
                },
                status=status.HTTP_200_OK
            )
        except Lecciones.DoesNotExist:
            return Response(
                {'error': 'Lección no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Error al completar lección: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ResponderCuestionarioView(APIView):
    permission_classes = [IsAuthenticated]
    """Responder cuestionario de una lección tipo formulario
    
    Optimización:
    - Select_related para obtener lección, módulo y capacitación en una query
    - Bulk operations para guardar respuestas
    - Prefetch de preguntas con respuestas correctas
    - Transacción atómica para consistencia
    """
    
    @transaction.atomic
    def post(self, request, leccion_id, *args, **kwargs):
        try:
            # Obtener colaborador desde el usuario autenticado
            if not request.user.idcolaboradoru:
                return Response(
                    {'error': 'El usuario no tiene un colaborador asociado'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            colaborador = request.user.idcolaboradoru
            respuestas_ids = request.data.get('respuestas', [])
            # Dict { str(pregunta_id): texto } para preguntas abiertas
            textos_abiertos = request.data.get('respuestas_abiertas', {})

            if not respuestas_ids and not textos_abiertos:
                return Response(
                    {'error': 'Se requiere al menos una respuesta'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Obtener lección con relaciones en una sola query
            leccion = Lecciones.objects.select_related(
                'idmodulo__idcapacitacion'
            ).get(id=leccion_id)
            
            modulo = leccion.idmodulo
            capacitacion = modulo.idcapacitacion
            tipo_capacitacion = capacitacion.tipo
            
            # Verificar inscripción (query optimizada)
            inscrito = progresoCapacitaciones.objects.filter(
                colaborador=colaborador,
                capacitacion=capacitacion
            ).only('id').exists()
            
            if not inscrito:
                return Response(
                    {'error': 'No tienes acceso a esta lección. No estás inscrito en la capacitación correspondiente.'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Obtener preguntas con respuestas correctas en una query optimizada
            preguntas = PreguntasLecciones.objects.filter(
                id_leccion=leccion
            ).prefetch_related(
                Prefetch(
                    'respuestas_set',
                    queryset=Respuestas.objects.filter(escorrecto=1),
                    to_attr='respuestas_correctas'
                )
            )
            
            total_preguntas = preguntas.count()
            
            if total_preguntas == 0:
                return Response(
                    {'error': 'Esta lección no tiene preguntas asociadas'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Obtener IDs de respuestas correctas
            respuestas_correctas_ids = set()
            for pregunta in preguntas:
                for resp in pregunta.respuestas_correctas:
                    respuestas_correctas_ids.add(resp.id)
            
            # Eliminar respuestas anteriores en bulk
            preguntas_ids = [p.id for p in preguntas]
            RespuestasColaboradores.objects.filter(
                idcolaborador=colaborador,
                idpregunta_id__in=preguntas_ids
            ).delete()
            
            # Obtener respuestas válidas (preguntas cerradas) y crear en bulk
            respuestas_validas = Respuestas.objects.filter(
                id__in=respuestas_ids
            ).select_related('idpregunta')

            nuevas_respuestas = [
                RespuestasColaboradores(
                    idcolaborador=colaborador,
                    idpregunta=respuesta.idpregunta,
                    idrespuesta=respuesta,
                )
                for respuesta in respuestas_validas
            ]

            # Procesar preguntas abiertas: { pregunta_id: texto }
            for pregunta_id_str, texto in textos_abiertos.items():
                try:
                    pregunta_id = int(pregunta_id_str)
                    resp_abierta = Respuestas.objects.filter(
                        idpregunta_id=pregunta_id, escorrecto=1
                    ).first()
                    if resp_abierta:
                        nuevas_respuestas.append(RespuestasColaboradores(
                            idcolaborador=colaborador,
                            idpregunta_id=pregunta_id,
                            idrespuesta=resp_abierta,
                            texto_respuesta=texto,
                        ))
                except (ValueError, TypeError):
                    pass

            RespuestasColaboradores.objects.bulk_create(nuevas_respuestas)
            
            # Calcular respuestas correctas del usuario
            # Las preguntas abiertas con texto siempre cuentan como correctas (escorrecto=1 por defecto)
            respuestas_correctas_usuario = set(respuestas_ids) & respuestas_correctas_ids
            total_abiertas_respondidas = len(textos_abiertos)
            total_correctas = len(respuestas_correctas_usuario) + total_abiertas_respondidas

            # Calcular porcentaje de acierto
            porcentaje_acierto = (total_correctas / total_preguntas) * 100
            
            # Determinar si pasó la lección (>60%)
            aprobada = porcentaje_acierto >= 60


            if tipo_capacitacion == 'ENCUESTA':
                # Para encuestas, el porcentage es 100
                porcentaje_acierto = 100
                aprobada = True
                total_correctas = total_preguntas  # Contar todas como correctas para encuestas
            
            # Actualizar progreso usando la función de utils
            progreso = 100 if aprobada else 0
            progreso_data = actualizar_progreso_leccion(
                colaborador_id=colaborador.idcolaborador,
                leccion=leccion,
                progreso=progreso,
                completada=aprobada
            )
            
            # Invalidar cache del colaborador
            invalidate_capacitacion_cache(colaborador_id=colaborador.idcolaborador)
            
            return Response(
                {
                    'mensaje': 'Cuestionario respondido exitosamente',
                    'leccion_id': leccion_id,
                    'total_preguntas': total_preguntas,
                    'respuestas_correctas': total_correctas,
                    'porcentaje_acierto': round(porcentaje_acierto, 2),
                    'aprobada': aprobada,
                    'progreso_leccion': progreso,
                    'progreso_modulo': progreso_data.get('progreso_modulo', 0),
                    'progreso_capacitacion': progreso_data.get('progreso_capacitacion', 0)
                },
                status=status.HTTP_200_OK
            )
        except Lecciones.DoesNotExist:
            return Response(
                {'error': 'Lección no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Error al procesar cuestionario: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PrevisualizarColaboradoresView(APIView):
    permission_classes = [IsAuthenticated]
    """Cargar archivo CSV con cédulas y previsualizar colaboradores
    
    Optimización:
    - Búsqueda en bulk en lugar de consulta individual por cédula
    - Solo se cargan campos necesarios con only()
    
    Retorna errores detallados para mejor UX en el frontend
    """    
    def post(self, request, *args, **kwargs):
        try:
            archivo = request.FILES.get('archivo')
            if not archivo:
                return Response(
                    {
                        'error': 'No se proporcionó archivo',
                        'detalles': ['Por favor selecciona un archivo CSV']
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar que sea un archivo CSV
            if not archivo.name.endswith('.csv'):
                return Response(
                    {
                        'error': f'El archivo debe ser un CSV. Recibido: {archivo.content_type}',
                        'detalles': [
                            'Extiende correctamente el archivo con .csv',
                            f'Archivo recibido: {archivo.name}'
                        ]
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar tamaño del archivo
            max_size = 5 * 1024 * 1024  # 5MB
            if archivo.size > max_size:
                size_mb = round(archivo.size / 1024 / 1024, 2)
                return Response(
                    {
                        'error': f'El archivo supera el tamaño máximo permitido (5MB). Tamaño: {size_mb}MB',
                        'detalles': ['Divide el archivo en partes más pequeñas']
                    },
                    status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE
                )
            
            # Leer el archivo CSV
            try:
                decoded_file = archivo.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                
                # Detectar el delimitador
                first_line = decoded_file.split('\n')[0]
                if ';;' in first_line:
                    delimiter = ';'
                elif ';' in first_line:
                    delimiter = ';'
                else:
                    delimiter = ','
                
                csv_reader = csv.DictReader(io_string, delimiter=delimiter)

                fieldnames = csv_reader.fieldnames or []
                normalized_fieldnames = [fn.strip().lower() for fn in fieldnames if fn]

                # Validar que tenga la columna 'cedula'
                if 'cedula' not in normalized_fieldnames:
                    return Response(
                        {
                            'error': 'El archivo CSV debe contener una columna "cedula"',
                            'detalles': [
                                f'Columnas encontradas: {", ".join(fieldnames)}',
                                'Asegúrate de que una columna se llame "cedula"'
                            ]
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Mapear nombre real del campo para obtener los valores
                cedula_field = fieldnames[normalized_fieldnames.index('cedula')]
                
                # Recolectar todas las cédulas primero
                cedulas = []
                filas_sin_cedula = 0
                for idx, row in enumerate(csv_reader, start=2):  # start=2 porque fila 1 es header
                    cedula = row.get(cedula_field, '').strip()
                    if cedula:
                        cedulas.append(cedula)
                    else:
                        filas_sin_cedula += 1
                
                # Si no hay cédulas válidas
                if not cedulas:
                    return Response(
                        {
                            'error': 'No se encontraron cédulas válidas en el archivo',
                            'detalles': [
                                f'Filas sin cédula: {filas_sin_cedula}',
                                'Verifica que la columna "cedula" contenga valores'
                            ]
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Búsqueda en bulk - una sola query para todas las cédulas
                colaboradores_db = Colaboradores.objects.filter(
                    cccolaborador__in=cedulas
                ).only(
                    'idcolaborador', 'cccolaborador', 
                    'nombrecolaborador', 'apellidocolaborador'
                )
                
                # Crear diccionario para búsqueda rápida O(1)
                colaboradores_map = {
                    c.cccolaborador: c for c in colaboradores_db
                }
                
                colaboradores_encontrados = []
                colaboradores_no_encontrados = []
                
                for cedula in cedulas:
                    colaborador = colaboradores_map.get(cedula)
                    if colaborador:
                        colaboradores_encontrados.append({
                            'id': colaborador.idcolaborador,
                            'cc_colaborador': colaborador.cccolaborador,
                            'nombre': colaborador.nombrecolaborador,
                            'apellido': colaborador.apellidocolaborador
                        })
                    else:
                        # Solo almacenar las cédulas no encontradas (lista simple)
                        colaboradores_no_encontrados.append(cedula)
                
                # Determinar tipo de respuesta
                respuesta = {
                    'mensaje': 'Archivo procesado correctamente',
                    'total_procesados': len(cedulas),
                    'encontrados': len(colaboradores_encontrados),
                    'no_encontrados': len(colaboradores_no_encontrados),
                    'colaboradores': colaboradores_encontrados,
                    'colaboradores_no_encontrados': colaboradores_no_encontrados
                }
                
                # Si no se encontraron colaboradores
                if not colaboradores_encontrados:
                    respuesta['detalles'] = [
                        'Ninguna de las cédulas fue encontrada en el sistema',
                        'Verifica la ortografía de las cédulas',
                        'Asegúrate de que los colaboradores estén registrados'
                    ]
                    return Response(respuesta, status=status.HTTP_200_OK)
                
                return Response(respuesta, status=status.HTTP_200_OK)
                
            except UnicodeDecodeError as e:
                return Response(
                    {
                        'error': 'Error al decodificar el archivo',
                        'detalles': [
                            'El archivo no está en formato UTF-8',
                            'Codificación no soportada detectada',
                            'Solución: Guarda el archivo en formato UTF-8 desde Excel o un editor de texto'
                        ]
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            except csv.Error as e:
                return Response(
                    {
                        'error': f'Error al procesar el archivo CSV: Fila con formato incorrecto',
                        'detalles': [
                            f'Detalle: {str(e)}',
                            'Verifica que todas las filas tengan el mismo número de columnas',
                            'Comprueba que no haya saltos de línea dentro de los campos'
                        ]
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            except Exception as e:
                logger.error(f"Error inesperado al procesar CSV: {str(e)}", exc_info=True)
                return Response(
                    {
                        'error': 'Error inesperado al procesar el archivo',
                        'detalles': [
                            'Por favor intenta de nuevo',
                            'Si el problema persiste, contacta al administrador'
                        ]
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
                
        except Exception as e:
            logger.error(f"Error en PrevisualizarColaboradoresView: {str(e)}", exc_info=True)
            return Response(
                {
                    'error': f'Error al procesar archivo: {str(e)}',
                    'detalles': ['Por favor intenta de nuevo más tarde']
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CargarArchivoView(APIView):
    permission_classes = [IsAuthenticated]
    """Cargar archivos (imágenes y PDFs) usando procesar_archivo_multimedia"""
    
    def post(self, request, *args, **kwargs):
        try:
            from capacitaciones.utils import procesar_archivo_multimedia
            
            archivo = request.FILES.get('archivo')
            if not archivo:
                return Response(
                    {'error': 'No se proporcionó archivo'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Determinar tipo de archivo
            nombre_archivo = archivo.name
            extension = nombre_archivo.rsplit('.', 1)[-1].lower() if '.' in nombre_archivo else ''
            
            # Mapear extensión a tipo
            if extension in ['jpg', 'jpeg', 'png', 'gif', 'webp', 'bmp']:
                tipo_archivo = 'imagen'
            elif extension in ['pdf']:
                tipo_archivo = 'pdf'
            else:
                return Response(
                    {'error': f'Tipo de archivo no permitido. Extensiones válidas: jpg, jpeg, png, gif, webp, bmp, pdf'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Procesar archivo con procesar_archivo_multimedia (genera nombre con UUID_original)
            url, error = procesar_archivo_multimedia(archivo, tipo_archivo)
            
            if error:
                return Response(
                    {'error': error},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return Response(
                {
                    'url': url,
                    'filename': archivo.name,
                    'original_filename': nombre_archivo,
                    'extension': extension,
                    'size': archivo.size,
                    'local': True
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {'error': f'Error al subir archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DescargarCertificadoView(APIView):
    permission_classes = [IsAuthenticated]
    """Generar y descargar certificado de capacitación completada"""
    
    def _convertir_docx_a_pdf(self, docx_path, pdf_path):
        """
        Convertir DOCX a PDF
        Nota: Requiere una herramienta externa (pypandoc con backend disponible)
        Por ahora retorna False para indicar que se devuelve el DOCX
        """
        logger = logging.getLogger("certificado_debug")
        try:
            # 1) Intentar con docx2pdf (usa Word en Windows o LibreOffice en algunas plataformas)
            try:
                from docx2pdf import convert as _docx2pdf_convert
                logger.warning("Intentando conversión con docx2pdf...")
                try:
                    _docx2pdf_convert(docx_path, pdf_path)
                    if os.path.exists(pdf_path):
                        logger.warning("Conversión con docx2pdf exitosa")
                        return True
                except Exception as e:
                    logger.warning(f"docx2pdf falló: {e}")
            except Exception:
                logger.warning("docx2pdf no está disponible")

            # 2) Intentar con pypandoc (requiere pandoc y típicamente LaTeX para PDF)
            try:
                logger.warning("Intentando conversión con pypandoc...")
                try:
                    _pypandoc.convert_file(docx_path, 'pdf', outputfile=pdf_path)
                    if os.path.exists(pdf_path):
                        logger.warning("Conversión con pypandoc exitosa")
                        return True
                except Exception as e:
                    logger.warning(f"pypandoc falló: {e}")
            except Exception:
                logger.warning("pypandoc no está disponible")

            # 3) Intentar con LibreOffice (soffice) en modo headless
            try:
                import subprocess, shutil
                logger.warning("Intentando conversión con LibreOffice (soffice)...")
                out_dir = os.path.dirname(pdf_path) or os.getcwd()
                cmd = ['soffice', '--headless', '--convert-to', 'pdf', '--outdir', out_dir, docx_path]
                proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=120)
                logger.warning(f"soffice exit {proc.returncode}")
                # LibreOffice normalmente crea un PDF con mismo nombre base in out_dir
                expected = os.path.join(out_dir, os.path.splitext(os.path.basename(docx_path))[0] + '.pdf')
                if os.path.exists(expected):
                    try:
                        shutil.move(expected, pdf_path)
                    except Exception:
                        pass
                    if os.path.exists(pdf_path):
                        logger.warning("Conversión con LibreOffice exitosa")
                        return True
            except Exception as e:
                logger.warning(f"LibreOffice/soffice falló: {e}")

            logger.warning("No se pudo convertir DOCX a PDF con los backends disponibles")
            return False

        except Exception as e:
            logger.error(f"Error inesperado en conversión: {e}")
            return False

    def _generar_certificado_desde_svg(self, svg_path, datos, output_dir, filename_base):
        """
        Genera un certificado SVG insertando texto dinámico con el fondo gráfico.
        
        - Inserta elementos <text> en posiciones específicas del SVG.
        - Devuelve el SVG directamente (preserva todos los gráficos de fondo).
        
        - `svg_path`: ruta al archivo SVG base.
        - `datos`: dict con llaves como '{{NOMBRE}}', '{{CEDULA}}', '{{CURSO}}', '{{FECHA}}'.
        - `output_dir`: directorio donde escribir archivos temporales.
        - `filename_base`: nombre base para el archivo generado (sin extensión).
        
        Retorna: (ruta_archivo_generado, extension, content_type)
        
        Guía de posiciones en el certificado (viewBox 0 0 255.72 182.57):
        - "Felicitamos al señor(a)" → NOMBRE va debajo (y=~110)
        - "Con documento de identidad CC" → CEDULA va a la derecha (x=~180, y=~118)
        - "Por su aprobación de" → CURSO va debajo (y=~130)
        - "En constancia de lo anterior se firma el" → FECHA va a la derecha (x=~180, y=~145)
        """
        import xml.etree.ElementTree as ET
        logger = logging.getLogger("certificado_debug")
        
        # Registrar namespaces para SVG
        ET.register_namespace('', 'http://www.w3.org/2000/svg')
        ET.register_namespace('xlink', 'http://www.w3.org/1999/xlink')
        
        # Leer SVG
        tree = ET.parse(svg_path)
        root = tree.getroot()
        
        # Diccionario de posiciones para cada elemento de texto
        # Ajusta estas coordenadas según tu plantilla SVG específica
        posiciones_texto = {
            '{{NOMBRE}}': {
                'x': '150.86',
                'y': '84',
                'font-size': '8',
                'text-anchor': 'middle',
                'fill': '#000000',
                'font-family': 'Montserrat-Light, Montserrat',
                'font-weight': '400',
                'letter-spacing': '0em',
                'max-width': None
            },
            '{{CEDULA}}': {
                'x': '174',
                'y': '94.3',
                'font-size': '3.41',
                'text-anchor': 'start',
                'fill': '#000000',
                'font-family': 'Montserrat-Light, Montserrat',
                'font-weight': '300',
                'letter-spacing': '0em',
                'max-width': None
            },
            '{{CURSO}}': {
                'x': '156',
                'y': '120.5',
                'font-size': '6',
                'text-anchor': 'middle',
                'fill': '#000000',
                'font-family': 'Montserrat-Light, Montserrat',
                'font-weight': '400',
                'letter-spacing': '0em',
                'max-width': 35
            },
            '{{FECHA}}': {
                'x': '170',
                'y': '132.2',
                'font-size': '3.41',
                'text-anchor': 'start',
                'fill': '#000000',
                'font-family': 'Montserrat-Light, Montserrat',
                'font-weight': '300',
                'letter-spacing': '0em',
                'max-width': None
            }
        }
        # Ajustar posición Y de '{{CURSO}}' según la longitud del texto proporcionado
        try:
            curso_pos = posiciones_texto.get('{{CURSO}}', {})
            # Si no se define max-width en la plantilla, usamos 35 como fallback
            max_width = curso_pos.get('max-width') if curso_pos.get('max-width') is not None else 35
            curso_text = datos.get('{{CURSO}}', '') if isinstance(datos, dict) else ''
            if curso_text and len(str(curso_text)) > int(max_width):
                # Texto largo: utilizar la Y para curso largo
                curso_pos['y'] = '115.5'
            else:
                # Texto corto o vacío: usar la Y por defecto para curso corto
                curso_pos['y'] = '118.5'
            posiciones_texto['{{CURSO}}'] = curso_pos
        except Exception as e:
            logger.warning(f"Error al ajustar posición de CURSO: {e}")

        # Crear grupo para textos
        text_group = ET.Element('{http://www.w3.org/2000/svg}g')
        text_group.set('id', 'dynamic_text_overlay')
        text_group.set('class', 'dynamic-text')  # Usar la clase de texto del SVG
        
        # Función auxiliar para manejar textos largos con saltos de línea
        def create_text_elements(key, val, pos, parent_group):
            """Crea elementos de texto con saltos de línea si es necesario"""
            max_width = pos.get('max-width', None)
            text_val = str(val) if val else ""
            
            # Si el texto es muy largo y tiene max-width, hacer saltos de línea
            if max_width and len(text_val) > max_width:
                # Dividir el texto en palabras
                palabras = text_val.split()
                lineas = []
                linea_actual = ""
                
                for palabra in palabras:
                    if len(linea_actual) + len(palabra) + 1 <= max_width:
                        linea_actual += (" " if linea_actual else "") + palabra
                    else:
                        if linea_actual:
                            lineas.append(linea_actual)
                        linea_actual = palabra
                
                if linea_actual:
                    lineas.append(linea_actual)
                
                # Crear elemento de texto para cada línea
                y_offset = float(pos.get('y', '0'))
                line_height = float(pos.get('font-size', '4')) * 1.5  # 1.5x el tamaño de fuente
                
                for idx, linea in enumerate(lineas):
                    # Crear un solo elemento de texto
                    text_elem = ET.Element('{http://www.w3.org/2000/svg}text')
                    text_elem.set('x', pos.get('x', '0'))
                    text_elem.set('y', str(y_offset + (idx * line_height)))
                    text_elem.set('font-size', pos.get('font-size', '4'))
                    text_elem.set('text-anchor', pos.get('text-anchor', 'middle'))
                    text_elem.set('fill', pos.get('fill', '#000000'))
                    text_elem.set('font-family', pos.get('font-family', 'Montserrat-Light, Montserrat'))

                    if 'font-weight' in pos:
                        text_elem.set('font-weight', pos.get('font-weight'))

                    text_elem.text = linea
                    parent_group.append(text_elem)
                    logger.warning(f"Línea {idx+1} de '{key}': '{linea}'")
            else:
                # Crear un solo elemento de texto
                text_elem = ET.Element('{http://www.w3.org/2000/svg}text')
                text_elem.set('x', pos.get('x', '0'))
                text_elem.set('y', pos.get('y', '0'))
                text_elem.set('font-size', pos.get('font-size', '4'))
                text_elem.set('text-anchor', pos.get('text-anchor', 'middle'))
                text_elem.set('fill', pos.get('fill', '#000000'))
                text_elem.set('font-family', pos.get('font-family', 'montserrat light, Montserrat'))
                text_elem.text = text_val
                parent_group.append(text_elem)
                logger.warning(f"Añadido elemento de texto '{key}': '{text_val}'")
        
        # Añadir elementos <text> para cada dato
        for key, val in datos.items():
            if key in posiciones_texto:
                pos = posiciones_texto[key]
                create_text_elements(key, val, pos, text_group)
        
        # Insertar el grupo de textos al final del SVG (para que esté encima)
        root.append(text_group)

                
        # Agregar texto fijo "Con documento de identidad CC" antes de la fecha
        fixed_text_group = ET.Element('{http://www.w3.org/2000/svg}g')
        fixed_text_group.set('id', 'fixed_text_constancia')
        
        constancia_text = ET.Element('{http://www.w3.org/2000/svg}text')
        constancia_text.set('x', '145')
        constancia_text.set('y', '94.3')
        constancia_text.set('font-size', '3.41')
        constancia_text.set('text-anchor', 'middle')
        constancia_text.set('fill', '#000000')
        constancia_text.set('font-family', 'Montserrat-Light, Montserrat')
        constancia_text.set('font-weight', '400')
        constancia_text.text = 'Con documento de identidad CC'
        fixed_text_group.append(constancia_text)
        root.append(fixed_text_group)

        
        # Agregar texto fijo "En constancia de lo anterior se firma el" antes de la fecha
        fixed_text_group = ET.Element('{http://www.w3.org/2000/svg}g')
        fixed_text_group.set('id', 'fixed_text_constancia')
        
        constancia_text = ET.Element('{http://www.w3.org/2000/svg}text')
        constancia_text.set('x', '135')
        constancia_text.set('y', '132.2')
        constancia_text.set('font-size', '3.41')
        constancia_text.set('text-anchor', 'middle')
        constancia_text.set('fill', '#000000')
        constancia_text.set('font-family', 'Montserrat-Light, Montserrat')
        constancia_text.set('font-weight', '400')
        constancia_text.text = 'En constancia de lo anterior se firma el'
        fixed_text_group.append(constancia_text)
        root.append(fixed_text_group)
        
        # Guardar SVG temporal con encoding UTF-8 explícito
        temp_svg = os.path.join(output_dir, f"{filename_base}.svg")
        
        # Escribir el SVG asegurando UTF-8
        tree.write(temp_svg, encoding='utf-8', xml_declaration=True)
        
        # Leer y reescribir para garantizar UTF-8 correcto
        with open(temp_svg, 'r', encoding='utf-8') as f:
            svg_content = f.read()
        
        # Asegurar que la declaración XML especifique UTF-8
        if "<?xml" not in svg_content:
            svg_content = "<?xml version='1.0' encoding='utf-8'?>\n" + svg_content
        elif "encoding=" not in svg_content:
            svg_content = svg_content.replace("<?xml version='1.0'?>", "<?xml version='1.0' encoding='utf-8'?>")
        
        # Reescribir con UTF-8 garantizado
        with open(temp_svg, 'w', encoding='utf-8') as f:
            f.write(svg_content)
        
        logger.warning(f"SVG con texto dinámico guardado en: {temp_svg}")
        logger.warning(f"SVG content preview (first 500 chars): {svg_content[:500]}")
        
        logger.warning(f"SVG con texto dinámico guardado en: {temp_svg}")
        
        # Convertir SVG a PDF usando weasyprint (preserva gráficos y tamaño)
        try:
            from weasyprint import HTML, CSS
            from io import BytesIO
            
            pdf_output = os.path.join(output_dir, f"{filename_base}.pdf")
            
            # Crear CSS para asegurar que el PDF tenga el mismo tamaño que el SVG
            # SVG viewBox: 0 0 255.72 182.57 ≈ 9.05" x 6.46" en 28.35 DPI (estándar SVG)
            css_string = '''
            @page {
                size: 255.72mm 182.57mm;
                margin: 0;
            }
            body {
                margin: 0;
                padding: 0;
            }
            svg {
                width: 100%;
                height: 100%;
            }
            '''
            
            # Leer el SVG como string con encoding UTF-8 explícito
            with open(temp_svg, 'r', encoding='utf-8') as f:
                svg_content = f.read()
            
            logger.warning(f"SVG a convertir contiene: {svg_content[200:400]}")
            
            # Usar HTML.from_string con encoding UTF-8 explícito
            html_obj = HTML(string=svg_content, encoding='utf-8')
            html_obj.write_pdf(pdf_output, stylesheets=[CSS(string=css_string)])
            
            logger.warning(f"✓ SVG convertido a PDF con tamaño correcto: {pdf_output}")
            return (pdf_output, 'pdf', 'application/pdf')
        except Exception as e:
            logger.warning(f"Weasyprint falló ({e}), intentando con cairosvg...")
            try:
                import cairosvg
                pdf_output = os.path.join(output_dir, f"{filename_base}.pdf")
                cairosvg.svg2pdf(url=temp_svg, write_to=pdf_output)
                logger.warning(f"✓ SVG convertido a PDF con cairosvg: {pdf_output}")
                return (pdf_output, 'pdf', 'application/pdf')
            except Exception as e2:
                logger.error(f"❌ CRÍTICO: Ambas librerías de conversión PDF fallaron!")
                logger.error(f"   Weasyprint error: {e}")
                logger.error(f"   Cairosvg error: {e2}")
                # NO retornar SVG, lanzar excepción para que se maneje en el endpoint
                raise Exception(f"No se pudo convertir SVG a PDF. Weasyprint: {e}, Cairosvg: {e2}")


    
    def get(self, request, id_capacitacion, id_colaborador=None, *args, **kwargs):
        import logging
        logger = logging.getLogger("certificado_debug")
        logger.warning(f"=== DESCARGA DE CERTIFICADO ===")
        logger.warning(f"request.user: {request.user}")
        logger.warning(f"request.user.id: {getattr(request.user, 'id', None)}")
        logger.warning(f"request.user.idcolaboradoru_id: {getattr(request.user, 'idcolaboradoru_id', None)}")
        logger.warning(f"id_capacitacion (URL): {id_capacitacion}")
        logger.warning(f"id_colaborador (URL): {id_colaborador}")
        
        try:
            # VALIDACIÓN 1: Obtener el colaborador del usuario autenticado
            user_colaborador_id = getattr(request.user, 'idcolaboradoru_id', None)
            
            if not user_colaborador_id:
                logger.error(f"SEGURIDAD: Usuario {request.user.id} no tiene colaborador asociado")
                return Response(
                    {'error': 'El usuario no tiene un colaborador asociado'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            logger.warning(f"Colaborador del usuario autenticado: {user_colaborador_id}")
            
            # VALIDACIÓN 2: Si se proporciona id_colaborador en la URL, verificar que coincida
            if id_colaborador:
                logger.warning(f"Validando que id_colaborador ({id_colaborador}) coincida con usuario ({user_colaborador_id})")
                if int(id_colaborador) != user_colaborador_id:
                    logger.error(f"INTENTO DE ACCESO NO AUTORIZADO: Usuario {request.user.id} (col {user_colaborador_id}) intentó descargar certificado de col {id_colaborador}")
                    return Response(
                        {'error': 'No tienes permiso para descargar este certificado'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            else:
                # Si no se proporciona, usar el del usuario autenticado
                id_colaborador = user_colaborador_id
                logger.warning(f"id_colaborador asignado del usuario autenticado: {id_colaborador}")
            
            logger.warning(f"VALIDACIÓN EXITOSA: Usuario {request.user.id} descargando certificado de colaborador {id_colaborador}")
            
            # Obtener colaborador con centro y cargo (necesarios para derivar empresa)
            colaborador = Colaboradores.objects.select_related(
                'centroop__id_proyecto__id_unidad__id_empresa',
                'cargocolaborador'
            ).only(
                'idcolaborador', 'cccolaborador', 'nombrecolaborador',
                'apellidocolaborador', 'correocolaborador', 'cargocolaborador', 'centroop'
            ).get(idcolaborador=id_colaborador)
            
            # ======================================================================
            # VALIDACIÓN: Verificar que la empresa del usuario está autorizada
            # ======================================================================
            empresa = None
            try:
                centro = getattr(colaborador, 'centroop', None)
                if centro and getattr(centro, 'id_proyecto', None) and getattr(centro.id_proyecto, 'id_unidad', None):
                    empresa = getattr(centro.id_proyecto.id_unidad, 'id_empresa', None)
            except Exception as e:
                logger.error(f"Error obtener empresa: {e}")
                empresa = None

            if not empresa:
                return Response(
                    {'error': 'El colaborador no tiene empresa asociada. No se puede descargar el certificado.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Lista de empresas autorizadas para descargar certificados
            # Ajusta esta lista según tus necesidades
            EMPRESAS_AUTORIZADAS = [
                'CONSORCIO',
                'PROTINCO',
                'REGENCY HEALTH',
                'REGENCY TECH',
                'REGENCY',
            ]

            empresa_nombre = (empresa.nombre_empresa or '').upper().strip()

            # Verificar si la empresa del usuario está en la lista autorizada
            empresa_autorizada = False
            for empresa_auth in EMPRESAS_AUTORIZADAS:
                if empresa_auth in empresa_nombre:
                    empresa_autorizada = True
                    break

            # Si no está autorizada, rechazar (a menos que sea admin)
            if not empresa_autorizada:
                user_role = getattr(request.user, 'tipousuario', None)
                # Solo SuperAdmin (4) y Admin (1) pueden descargar de empresas no autorizadas
                if user_role not in [1, 4]:
                    return Response(
                        {
                            'error': f'La empresa "{empresa_nombre}" no está autorizada para descargar certificados. Por favor contacta con administración.',
                            'empresa': empresa_nombre
                        },
                        status=status.HTTP_403_FORBIDDEN
                    )
                logger.warning(f"Admin ({user_role}) descargando certificado de empresa no autorizada: {empresa_nombre}")

            capacitacion = get_object_or_404(Capacitaciones, id=id_capacitacion)
            
            # Verificar que el colaborador completó la capacitación
            progreso = progresoCapacitaciones.objects.filter(
                colaborador=colaborador,
                capacitacion=capacitacion,
                completada=1
            ).first()
            
            if not progreso:
                return Response(
                    {'error': 'No has completado esta capacitación. Debes completarla al 100% para obtener el certificado.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Siempre generar nuevo certificado - Sin cache
            logger.warning(f"Generando certificado nuevo para colaborador {id_colaborador}, capacitación {id_capacitacion}")
            
            # Buscar archivo SVG en el directorio de certificados
            empresa_raw = getattr(empresa, 'nombre_empresa', None)
            logger.warning(f"Empresa nombre (raw BD): {empresa_raw}")
            
            # MAPEO EXPLÍCITO: Nombre de empresa en BD → Archivo SVG esperado
            # Esto evita búsquedas difusas y usa la correspondencia directa
            EMPRESA_SVG_MAPPING = {
                'CONSORCIO PEAJES 2526': 'CONSORCIO.svg',
                'PROTECCION DE INFRAESTRUCTURA COLOMBIA - PROTINCO LTDA': 'PROTINCO.svg',
                'REGENCY HEALTH SERVICES S.A.S': 'REGENCY HEALTH.svg',
                'REGENCY SERVICES DE COLOMBIA S.A.S': 'REGENCY.svg',
            }
            
            # Normalizar nombre de empresa para búsqueda en mapeo
            empresa_normalizada_map = (empresa_raw or '').upper().strip() if empresa_raw else ''
            svg_nombre_esperado = EMPRESA_SVG_MAPPING.get(empresa_normalizada_map)
            
            plantilla_path = None
            plantilla_nombre = None
            svg_files = []
            
            # Múltiples rutas posibles según el entorno (Docker, Local, etc)
            posibles_dirs = [
                os.path.join(settings.BASE_DIR, 'capacitaciones', 'templates', 'certificados'),
                os.path.join(settings.BASE_DIR, '..', 'capacitaciones', 'templates', 'certificados'),
                os.path.join(settings.BASE_DIR, 'plantillas'),
                os.path.join(settings.MEDIA_ROOT, 'certificados'),
                '/app/capacitaciones/templates/certificados',
                'capacitaciones/templates/certificados',
            ]
            
            # Buscar en todas las rutas posibles
            for certificados_dir in posibles_dirs:
                certificados_dir = os.path.normpath(certificados_dir)
                logger.warning(f"Intentando directorio: {certificados_dir}")
                
                if not os.path.isdir(certificados_dir):
                    logger.warning(f"   ✗ No existe o no es directorio")
                    continue
                
                try:
                    dir_files = os.listdir(certificados_dir)
                    svg_files_in_dir = [f for f in dir_files if f.endswith('.svg')]
                    
                    if svg_files_in_dir:
                        logger.warning(f"   ✓ Directorio contiene {len(svg_files_in_dir)} archivos SVG: {svg_files_in_dir}")
                        svg_files.extend([(f, certificados_dir) for f in svg_files_in_dir])
                    else:
                        logger.warning(f"   ⚠️ Directorio existe pero no contiene SVG")
                        logger.warning(f"   Archivos en dir: {dir_files[:10]}...")
                        
                except Exception as e:
                    logger.warning(f"   ✗ Error al listar: {e}")
                    continue
            
            logger.warning(f"Total de SVG encontrados en todas las rutas: {len(svg_files)}")
            
            # ESTRATEGIA 1: Usar mapeo explícito si existe
            if svg_nombre_esperado:
                logger.warning(f"Buscando SVG mapeado para '{empresa_normalizada_map}': {svg_nombre_esperado}")
                for svg_file, svg_dir in svg_files:
                    if svg_file.lower() == svg_nombre_esperado.lower():
                        plantilla_nombre = svg_file
                        plantilla_path = os.path.join(svg_dir, svg_file)
                        logger.warning(f"✓✓✓ SVG encontrado (MAPEO EXPLÍCITO): {plantilla_nombre} en {svg_dir}")
                        break
            
            # ESTRATEGIA 2: Si no se encuentra por mapeo explícito, usar búsquedas progresivas como fallback
            if not plantilla_path and svg_files:
                logger.warning(f"SVG no encontrado por mapeo. Usando búsquedas progresivas como fallback...")
                
                # Búsqueda 1: Coincidencia EXACTA (case-insensitive)
                for svg_file, svg_dir in svg_files:
                    svg_name = svg_file.lower().replace('.svg', '')
                    empresa_lower = empresa_normalizada_map.lower()
                    if svg_name == empresa_lower:
                        plantilla_nombre = svg_file
                        plantilla_path = os.path.join(svg_dir, svg_file)
                        logger.warning(f"✓✓ SVG encontrado (coincidencia EXACTA): {plantilla_nombre}")
                        break
                
                # Búsqueda 2: Coincidencia de PALABRAS CLAVE
                if not plantilla_path:
                    empresa_keywords = empresa_normalizada_map.split()
                    logger.warning(f"Palabras clave de empresa: {empresa_keywords}")
                    for svg_file, svg_dir in svg_files:
                        svg_normalized = svg_file.upper().replace('.SVG', '')
                        matches = sum(1 for kw in empresa_keywords if kw in svg_normalized)
                        if matches >= 2:
                            plantilla_nombre = svg_file
                            plantilla_path = os.path.join(svg_dir, svg_file)
                            logger.warning(f"✓ SVG encontrado (coincidencia de {matches} keywords): {plantilla_nombre}")
                            break
                
                # Búsqueda 3: BÚSQUEDA PARCIAL
                if not plantilla_path:
                    for svg_file, svg_dir in svg_files:
                        if empresa_normalizada_map in svg_file.upper():
                            plantilla_nombre = svg_file
                            plantilla_path = os.path.join(svg_dir, svg_file)
                            logger.warning(f"✓ SVG encontrado (búsqueda parcial): {plantilla_nombre}")
                            break
                
                # Búsqueda 4: PRIMERA PALABRA coincide
                if not plantilla_path:
                    primera_palabra = empresa_normalizada_map.split()[0] if empresa_normalizada_map.split() else ''
                    if primera_palabra:
                        for svg_file, svg_dir in svg_files:
                            if primera_palabra in svg_file.upper():
                                plantilla_nombre = svg_file
                                plantilla_path = os.path.join(svg_dir, svg_file)
                                logger.warning(f"✓ SVG encontrado (primera palabra): {plantilla_nombre}")
                                break

            # Si no se encuentra SVG, retornar error detallado
            if not plantilla_path:
                plantillas_disp = [f"{fname} (en {dirname})" for fname, dirname in svg_files[:10]]
                logger.warning(f"✗✗✗ No se encontró SVG para empresa: {empresa_nombre}")
                logger.warning(f"Empresa normalizada: {empresa_normalizada}")
                logger.warning(f"Plantillas disponibles: {plantillas_disp}")
                
                return Response(
                    {
                        'error': f'No se encontró plantilla SVG para la empresa "{empresa_nombre}"',
                        'empresa_raw': empresa_raw,
                        'empresa_normalizada': empresa_normalizada,
                        'plantillas_disponibles': [f for f, _ in svg_files[:10]],
                        'directorios_buscados': [os.path.normpath(d) for d in posibles_dirs],
                        'total_plantillas_encontradas': len(svg_files)
                    },
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Preparar datos para reemplazar en la plantilla
            fecha_completada = progreso.fecha_completada or timezone.now()
            
            # Traducir mes a español
            meses = {
                1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
                5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
                9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
            }
            mes_nombre = meses[fecha_completada.month]
            fecha_formateada = f"{fecha_completada.day} de {mes_nombre} de {fecha_completada.year}"
            
            # Intentar cargar cargo y centro de forma segura
            cargo_nombre = ''
            centro_nombre = ''
            try:
                if colaborador.cargocolaborador:
                    cargo_nombre = str(colaborador.cargocolaborador.nombrecargo) if colaborador.cargocolaborador.nombrecargo else ''
            except Exception:
                cargo_nombre = ''
            
            try:
                if colaborador.centroop:
                    centro_nombre = str(colaborador.centroop.nombre_centrop) if colaborador.centroop.nombre_centrop else ''
            except Exception:
                centro_nombre = ''
            
            # Construir diccionario de datos con encoding UTF-8 explícito
            datos = {
                '{{NOMBRE}}': str(f"{colaborador.nombrecolaborador} {colaborador.apellidocolaborador}").encode('utf-8').decode('utf-8'),
                '{{CEDULA}}': str(colaborador.cccolaborador).encode('utf-8').decode('utf-8'),
                '{{CURSO}}': str(capacitacion.titulo).encode('utf-8').decode('utf-8'),
                '{{FECHA}}': str(fecha_formateada).encode('utf-8').decode('utf-8'),
                '{{nombre_completo}}': str(f"{colaborador.nombrecolaborador} {colaborador.apellidocolaborador}").encode('utf-8').decode('utf-8'),
                '{{nombre}}': str(colaborador.nombrecolaborador).encode('utf-8').decode('utf-8'),
                '{{apellido}}': str(colaborador.apellidocolaborador).encode('utf-8').decode('utf-8'),
                '{{cedula}}': str(colaborador.cccolaborador).encode('utf-8').decode('utf-8'),
                '{{capacitacion}}': str(capacitacion.titulo).encode('utf-8').decode('utf-8'),
                '{{fecha}}': str(fecha_formateada).encode('utf-8').decode('utf-8'),
                '{{fecha_corta}}': str(fecha_completada.strftime('%d/%m/%Y')).encode('utf-8').decode('utf-8'),
                '{{empresa}}': str(empresa.nombre_empresa).encode('utf-8').decode('utf-8'),
                '{{cargo}}': str(cargo_nombre).encode('utf-8').decode('utf-8'),
                '{{centro}}': str(centro_nombre).encode('utf-8').decode('utf-8'),
            }
            
            # Crear directorio temporal
            temp_dir = tempfile.mkdtemp()
            try:
                # Procesar el SVG (no DOCX) y generar el PDF
                resultado = self._generar_certificado_desde_svg(
                    svg_path=plantilla_path,
                    datos=datos,
                    output_dir=temp_dir,
                    filename_base='certificado'
                )
                
                if resultado is None:
                    return Response(
                        {'error': 'Error al generar certificado desde SVG'},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
                
                pdf_path, extension, content_type = resultado
                
                # Leer el archivo generado
                with open(pdf_path, 'rb') as f:
                    file_content = f.read()
                
                # Crear respuesta con el archivo
                response = FileResponse(
                    BytesIO(file_content),
                    content_type=content_type
                )
                
                # Determinar nombre del archivo de descarga basado en capacitación
                # Formato: certificado_[nombre_capacitacion].{extension}
                nombre_capacitacion = str(capacitacion.titulo).lower().replace(' ', '_')
                nombre_capacitacion = re.sub(r'[^a-z0-9_]', '', nombre_capacitacion)
                nombre_archivo = f"certificado_{nombre_capacitacion}.{extension}"
                response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                
                # Deshabilitar cache - Siempre generar nuevo certificado
                response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
                response['Pragma'] = 'no-cache'
                response['Expires'] = '0'
                
                logger.warning(f"✓ Certificado generado exitosamente: {nombre_archivo}")
                return response
                
            finally:
                # Limpiar directorio temporal
                try:
                    shutil.rmtree(temp_dir)
                except Exception as e:
                    logger.warning(f"Error al limpiar directorio temporal: {e}")
                
        except Exception as e:
            return Response(
                {'error': f'Error al generar certificado: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class MisCapacitacionesView(APIView):
    permission_classes = [IsAuthenticated]
    """Ver mis capacitaciones asignadas (una capacitación específica)"""
    
    def get(self, request, capacitacion_id, *args, **kwargs):
        try:
            # Obtener colaborador desde el token
            colaborador = getattr(request.user, 'idcolaboradoru', None)
            if not colaborador:
                return Response(
                    {'error': 'El usuario no tiene un colaborador asociado'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Prefetch optimizado con filtro del colaborador para progreso y estructura
            capacitacion = Capacitaciones.objects.prefetch_related(
                Prefetch(
                    'progresocapacitaciones_set',
                    queryset=progresoCapacitaciones.objects.filter(colaborador=colaborador),
                    to_attr='progreso_colaborador'
                ),
                'modulos_set__lecciones_set'
            ).get(pk=capacitacion_id)

            serializer = CapacitacionProgresoSerializer(
                capacitacion,
                context={'colaborador': colaborador}
            )
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Capacitaciones.DoesNotExist:
            return Response(
                {'error': 'Capacitación no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AdminCapacitacionProgresoView(APIView):
    permission_classes = [IsAuthenticated]
    """
    Ver progreso de un colaborador en una capacitación específica (admin).
    
    Optimización: Usa prefetch_related con to_attr para acceso directo
    al progreso del colaborador específico.
    """
    
    def get(self, request, capacitacion_id, colaborador_id, *args, **kwargs):
        try:
            colaborador = get_object_or_404(Colaboradores, id_colaborador=colaborador_id)
            
            # Prefetch optimizado con filtro del colaborador
            capacitacion = Capacitaciones.objects.prefetch_related(
                Prefetch(
                    'progresocapacitaciones_set',
                    queryset=progresoCapacitaciones.objects.filter(colaborador=colaborador),
                    to_attr='progreso_colaborador'
                ),
                'modulos_set__lecciones_set'
            ).get(pk=capacitacion_id)
            
            serializer = CapacitacionProgresoSerializer(
                capacitacion,
                context={'colaborador': colaborador}
            )
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Capacitaciones.DoesNotExist:
            return Response(
                {'error': 'Capacitación no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Colaboradores.DoesNotExist:
            return Response(
                {'error': 'Colaborador no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MisCapacitacionesListView(APIView):
    permission_classes = [IsAuthenticated]
    """
    Lista de capacitaciones asignadas al usuario autenticado.
    
    Solo el propio usuario puede ver sus capacitaciones.
    No se permite ver capacitaciones de otros usuarios.
    
    Optimización: 
    - Cache de 2 minutos por colaborador
    - Annotate + prefetch_related + to_attr para reducir queries (87% optimización)
    """
    
    def get(self, request, *args, **kwargs):
        try:
            # Obtener colaborador SOLO desde el token del usuario autenticado
            colaborador = getattr(request.user, 'idcolaboradoru', None)
            
            if not colaborador:
                return Response(
                    {'error': 'El usuario no tiene un colaborador asociado'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Intentar obtener de cache
            cache_key = get_cache_key('mis_caps', colaborador.idcolaborador)
            cached_data = cache.get(cache_key)
            
            if cached_data is not None:
                return Response(cached_data, status=status.HTTP_200_OK)
            
            # Optimización: Annotate + Prefetch con to_attr
            # Mostrar solo capacitaciones activas (estado=1) donde el progreso
            # del colaborador indicado NO esté completado (progresocapacitaciones.completada=0)
            capacitaciones = Capacitaciones.objects.filter(
                progresocapacitaciones__colaborador=colaborador,
                progresocapacitaciones__completada=0,
                estado=1
            ).annotate(
                total_lecciones_count=Count('modulos__lecciones', distinct=True),
                fecha_registro_progreso=Subquery(
                    progresoCapacitaciones.objects.filter(
                        capacitacion=OuterRef('pk'),
                        colaborador=colaborador
                    ).values('fecha_registro')[:1]
                )
            ).prefetch_related(
                Prefetch(
                    'progresocapacitaciones_set',
                    queryset=progresoCapacitaciones.objects.filter(colaborador=colaborador),
                    to_attr='progreso_colaborador'
                ),
                Prefetch(
                    'modulos_set',
                    queryset=Modulos.objects.prefetch_related(
                        Prefetch(
                            'lecciones_set',
                            queryset=Lecciones.objects.prefetch_related(
                                Prefetch(
                                    'progresolecciones_set',
                                    queryset=progresolecciones.objects.filter(idcolaborador=colaborador),
                                    to_attr='progreso_leccion_colaborador'
                                )
                            )
                        )
                    )
                )
            ).distinct().order_by('-fecha_registro_progreso')
            
            serializer = MisCapacitacionesSerializer(capacitaciones, many=True)
            
            # Guardar en cache (2 minutos - cambia más frecuentemente)
            cache_ttl = getattr(settings, 'CACHE_TTL_MIS_CAPACITACIONES', 120)
            cache.set(cache_key, serializer.data, cache_ttl)
            
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DesactivarCapacitacionesView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser, IsSuperAdmin]

    """Desactivar capacitaciones"""

    def put(self, request, *args, **kwargs):

        try:
            capacitacion_id = request.data.get('capacitacion_id')
            if not capacitacion_id:
                return Response({'error': 'capacitacion_id es requerido'}, status=status.HTTP_400_BAD_REQUEST)
            
            capacitacion = get_object_or_404(Capacitaciones, pk=capacitacion_id)
            # Toggle between estado 0 (desactivada) and 1 (activa)
            if capacitacion.estado == 0:
                capacitacion.estado = 1
                mensaje = 'Capacitación activada correctamente'
            elif capacitacion.estado == 1:
                capacitacion.estado = 0
                mensaje = 'Capacitación desactivada correctamente'
            else:
                return Response({'error': f'No se puede alternar estado desde el estado {capacitacion.estado}'}, status=status.HTTP_400_BAD_REQUEST)
            capacitacion.save()
            # If the capacitacion was activated, trigger the notification task after commit
            if capacitacion.estado == 1:
                try:
                    from capacitaciones.utils import enviar_correo_cap_activada_batch
                    # Ejecutar de forma síncrona después de que la transacción se confirme
                    transaction.on_commit(lambda: enviar_correo_cap_activada_batch(capacitacion))
                except Exception:
                    pass
            # Invalidate caches
            invalidate_capacitacion_cache(capacitacion_id=capacitacion.id)
            colaboradores_ids = progresoCapacitaciones.objects.filter(capacitacion=capacitacion).values_list('colaborador_id', flat=True)
            for col_id in colaboradores_ids:
                invalidate_capacitacion_cache(colaborador_id=col_id)
            cache.delete('capacitaciones_list_admin')
            return Response({'mensaje': mensaje, 'estado': capacitacion.estado}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    def delete(self, request, *args, **kwargs):
        try:
            # Support DELETE with body list or DELETE /desactivar-capacitaciones/<id>/
            cap_id = kwargs.get('capacitacion_id')
            if cap_id:
                ids_capacitaciones = [int(cap_id)]
            else:
                ids_capacitaciones = request.data.get('ids_capacitaciones') or request.data.get('capacitacion_ids') or []

            if isinstance(ids_capacitaciones, int):
                ids_capacitaciones = [ids_capacitaciones]

            if not isinstance(ids_capacitaciones, list) or not all(isinstance(i, int) for i in ids_capacitaciones):
                return Response(
                    {'error': 'ids_capacitaciones debe ser una lista de enteros'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            capacitaciones = Capacitaciones.objects.filter(id__in=ids_capacitaciones)
            actualizadas = capacitaciones.update(estado=3)  # 3 = Eliminada

            # Invalidate caches for affected capacitaciones and colaboradores
            for cap in capacitaciones:
                invalidate_capacitacion_cache(capacitacion_id=cap.id)
            colaboradores_ids = progresoCapacitaciones.objects.filter(capacitacion__id__in=ids_capacitaciones).values_list('colaborador_id', flat=True)
            for col_id in colaboradores_ids:
                invalidate_capacitacion_cache(colaborador_id=col_id)
            cache.delete('capacitaciones_list_admin')

            return Response(
                {'mensaje': f'Se eliminaron {actualizadas} capacitaciones.'},
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)     
        
        
class EditarColaboradorCapacitacionView(APIView):
    permission_classes = [IsAuthenticated, IsGestionEmpresarial]
    """Editar colaboradores asignados a una capacitación"""

    def get(self, request, capacitacion_id, *args, **kwargs):
        """
        Obtener lista de colaboradores asignados a la capacitación con id `capacitacion_id`.
        Retorna lista de IDs, nombre, apellido y cédula de colaboradores asignados.
        """
        try:
            capacitacion = Capacitaciones.objects.get(pk=capacitacion_id)
            colaboradores_ids = list(progresoCapacitaciones.objects.filter(capacitacion=capacitacion).values_list('colaborador_id', flat=True))
            # Obtener datos completos de los colaboradores
            colaboradores = list(
                Colaboradores.objects.filter(idcolaborador__in=colaboradores_ids)
                .values('idcolaborador', 'nombrecolaborador', 'apellidocolaborador', 'cccolaborador')
            )
            # Formatear para frontend (compatibilidad)
            colaboradores_data = [
                {
                    'id': c['idcolaborador'],
                    'nombre': c['nombrecolaborador'],
                    'apellido': c['apellidocolaborador'],
                    'cc': c['cccolaborador'],
                }
                for c in colaboradores
            ]
            return Response({'colaboradores': colaboradores_data}, status=status.HTTP_200_OK)
        except Capacitaciones.DoesNotExist:
            return Response({'error': 'Capacitación no encontrada'}, status=status.HTTP_404_NOT_FOUND)
    
    def put(self, request, capacitacion_id, *args, **kwargs):
        """
        Editar colaboradores asignados a la capacitación con id `capacitacion_id`.
        Espera un payload JSON con listas de IDs de colaboradores a agregar o remover:
        {
            "add": [1, 2, 3],
            "remove": [4, 5]
        }
        Agrega o remueve los colaboradores especificados de la capacitación.
        Retorna el estado actualizado de la capacitación.
        
        Si no se proporciona `capacitacion_id`: crear nueva capacitación (comportamiento previo).
        """
        if capacitacion_id:
            try:
                # permisos: sólo admin (IsAdminUser ya aplicado)
                capacitacion = Capacitaciones.objects.get(pk=capacitacion_id)
            except Capacitaciones.DoesNotExist:
                return Response({'error': 'Capacitación no encontrada'}, status=status.HTTP_404_NOT_FOUND)

            data = request.data or {}
            add_ids = set(data.get('add', []) or [])
            remove_ids = set(data.get('remove', []) or [])

            # validar que no haya intersección problemática
            if add_ids & remove_ids:
                return Response({'error': 'IDs en add y remove al mismo tiempo'}, status=status.HTTP_400_BAD_REQUEST)

            # validar existencia de colaboradores a agregar
            existing_add = set(Colaboradores.objects.filter(idcolaborador__in=add_ids).values_list('idcolaborador', flat=True))
            missing = add_ids - existing_add
            if missing:
                return Response({'error': f'Colaboradores no encontrados: {list(missing)}'}, status=status.HTTP_400_BAD_REQUEST)

            with transaction.atomic():
                added = []
                removed = []

                # Agregar nuevos (verificar que no existan ya)
                existentes = set(progresoCapacitaciones.objects.filter(
                    capacitacion=capacitacion
                ).values_list('colaborador_id', flat=True))
                to_add = add_ids - existentes
                bulk = []
                for cid in to_add:
                    # Verificación extra: solo agregar si realmente no existe
                    if not progresoCapacitaciones.objects.filter(
                        capacitacion=capacitacion, colaborador_id=cid
                    ).exists():
                        bulk.append(progresoCapacitaciones(
                            capacitacion=capacitacion,
                            colaborador_id=cid,
                            fecha_registro=timezone.now(),
                            completada=False,
                            progreso=0
                        ))
                        added.append(cid)
                if bulk:
                    progresoCapacitaciones.objects.bulk_create(bulk)

                # Enviar notificación solo a los agregados una vez la transacción se confirme
                # Solo enviar si la capacitación está activa (estado = 1)
                if added and capacitacion.estado == 1:
                    try:
                        from capacitaciones.utils import enviar_correo_nuevos_colaboradores
                        ids_added = list(added)
                        # Ejecutar de forma síncrona después de que la transacción se confirme
                        transaction.on_commit(lambda: enviar_correo_nuevos_colaboradores(capacitacion.id, ids_added))
                    except Exception:
                        pass

                # Eliminar solicitados
                to_remove = remove_ids & set(progresoCapacitaciones.objects.filter(capacitacion=capacitacion).values_list('colaborador_id', flat=True))
                if to_remove:
                    for cid in to_remove:
                        progresolecciones.objects.filter(idcolaborador_id=cid, idleccion__idmodulo__idcapacitacion=capacitacion).delete()
                        from .models import progresoModulo as _progresoModulo
                        _progresoModulo.objects.filter(colaborador_id=cid, modulo__idcapacitacion=capacitacion).delete()
                        progresoCapacitaciones.objects.filter(capacitacion=capacitacion, colaborador_id=cid).delete()
                        removed.append(cid)

                # invalidar caches para colaboradores afectados y para la capacitación
                invalidate_capacitacion_cache(capacitacion_id=capacitacion.id)
                for cid in set(added + removed):
                    invalidate_capacitacion_cache(colaborador_id=cid)
                # Limpiar cache de lista general de capacitaciones
                cache.delete('capacitaciones_list_admin')

                return Response({'added': added, 'removed': removed}, status=status.HTTP_200_OK)
            
            
class ObtenerColaboradoresCapacitacionView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser | IsSuperAdmin]
    """
    Listado de colaboradores de una capacitación con su avance.
    
    Comportamiento:
    - GET /capacitaciones/<capacitacion_id>/colaboradores/ → Devuelve TODOS los colaboradores
    - GET /capacitaciones/<capacitacion_id>/colaboradores/?limit=20&offset=0 → Paginado (20 por página)
    - GET /capacitaciones/<capacitacion_id>/colaboradores/?limit=50&offset=50 → Paginado (50 por página)
    
    Parámetros opcionales:
    - limit: Número de registros por página (máx 500). Si no se especifica, devuelve todos.
    - offset: Número de registros a saltar (default: 0)
    """
    def get(self, request, capacitacion_id, *args, **kwargs):
        try:
            # Detectar si se solicita paginación
            limit_param = request.GET.get('limit', None)
            offset_param = request.GET.get('offset', None)
            
            qs = progresoCapacitaciones.objects.filter(capacitacion_id=capacitacion_id)
            total = qs.count()
            
            # FIX: Usar Max('id') + GROUP BY colaborador para obtener el registro más reciente
            # Compatible con MySQL (no necesita DISTINCT ON ni LIMIT en subquery)
            latest_ids = list(
                qs.values('colaborador')
                .annotate(max_id=Max('id'))
                .values_list('max_id', flat=True)
            )
            
            # Filtrar solo los registros más recientes por colaborador
            qs_distinct = qs.filter(id__in=latest_ids).select_related('colaborador')
            
            # Si no se especifica limit, devolver TODOS
            if limit_param is None:
                limit = qs_distinct.count()
                offset = 0
                progreso_qs = qs_distinct
            else:
                # Si se especifica limit, aplicar paginación
                limit = int(limit_param)
                offset = int(offset_param) if offset_param else 0
                
                # Validaciones
                if limit > 500:
                    limit = 500
                if limit < 1:
                    limit = 1
                if offset < 0:
                    offset = 0
                
                progreso_qs = qs_distinct[offset:offset+limit]

            data = list(
                progreso_qs.values(
                    'colaborador__idcolaborador',
                    'colaborador__nombrecolaborador',
                    'colaborador__apellidocolaborador',
                    'colaborador__cccolaborador',
                    'progreso',
                    'completada',
                    'fecha_registro',
                    'fecha_completada',
                )
            )
            results = [
                {
                    'id': d['colaborador__idcolaborador'],
                    'nombre': d['colaborador__nombrecolaborador'],  # FIX: Nombre correcto
                    'apellido': d['colaborador__apellidocolaborador'],  # FIX: Apellido correcto
                    'cedula': d['colaborador__cccolaborador'],
                    'progreso': float(d['progreso']),
                    'completada': bool(d['completada']),
                    'fecha_registro': d['fecha_registro'],
                    'fecha_completada': d['fecha_completada'],
                }
                for d in data
            ]

            return Response({
                'count': total,
                'limit': limit,
                'offset': offset,
                'results': results
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ReporteCapacitacionesView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser | IsSuperAdmin]
    """
    Generar reporte de capacitaciones en formato exel.
    
    Casos de uso:
    1. Con ID de capacitación:
       GET /capacitaciones/reporte-capacitaciones/?capacitacion_id=1
       - Genera reporte del avance de colaboradores en esa capacitación
       - Incluye: nombre, apellido, cédula, % completación, fecha registro, fecha completación
       - Información de la capacitación: título, tipo, fecha de creación
    
    2. Sin ID (con rango de fechas):
       GET /capacitaciones/reporte-capacitaciones/?fecha_inicio=2025-01-01&fecha_fin=2025-12-31
       - Genera reporte de TODAS las capacitaciones en el rango de fechas
       - Excluye capacitaciones con estado 3 (eliminadas)
       - No cuenta colaboradores desactivados (estadocolaborador != 1)
    """
    
    def get(self, request, *args, **kwargs):
        try:
            capacitacion_id = request.GET.get('capacitacion_id')
            fecha_inicio = request.GET.get('fecha_inicio')
            fecha_fin = request.GET.get('fecha_fin')
            include_questions = request.GET.get('include_questions', 'false').lower() == 'true'
            
            if capacitacion_id:
                # CASO 1: Reporte de una capacitación específica
                return self._generar_reporte_capacitacion(capacitacion_id, include_questions)
            elif fecha_inicio and fecha_fin:
                # CASO 2: Reporte de rango de fechas
                return self._generar_reporte_rango_fechas(fecha_inicio, fecha_fin)
            else:
                return Response(
                    {'error': 'Debe proporcionar capacitacion_id o (fecha_inicio y fecha_fin)'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {'error': f'Error al generar reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _generar_reporte_capacitacion(self, capacitacion_id, include_questions=False):
        """
        Genera reporte de una capacitación específica con todos sus colaboradores.
        Si include_questions es True, incluye columnas de preguntas y respuestas por cada colaborador.
        """
        try:
            # Obtener capacitación
            capacitacion = get_object_or_404(Capacitaciones, id=capacitacion_id)
            
            # Obtener colaboradores inscritos (solo activos)
            progreso_qs = progresoCapacitaciones.objects.filter(
                capacitacion=capacitacion
            ).select_related(
                'colaborador',
                'colaborador__centroop',
                'colaborador__centroop__id_proyecto',
                'colaborador__centroop__id_proyecto__id_unidad',
                'colaborador__centroop__id_proyecto__id_unidad__id_empresa'
            ).order_by('colaborador__nombrecolaborador')
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Capacitación"
            
            # Estilos
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Información de la capacitación (encabezado)
            ws['A1'] = "REPORTE DE CAPACITACIÓN"
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['A2'] = f"Título: {capacitacion.titulo}"
            ws['A3'] = f"Tipo: {capacitacion.tipo}"
            ws['A4'] = f"Fecha de Creación: {capacitacion.fecha_creacion.strftime('%d/%m/%Y %H:%M') if capacitacion.fecha_creacion else 'N/A'}"
            
            # Headers de tabla
            headers = [
                "Empresa",
                "Unidad",
                "Descripción Unidad",
                "Proyecto",
                "Centro Op",
                "Cédula",
                "Cargo",
                "Nombre",
                "Apellido",
                "Correo",
                "% Completación",
                "Fecha Registro",
                "Fecha Completación",
                "Estado Avance"
            ]
            
            # Si se incluyen preguntas, obtener todas las preguntas de la capacitación
            preguntas_dict = {}  # {pregunta_id: pregunta_texto}
            if include_questions:
                # Obtener todas las lecciones y preguntas de esta capacitación
                modulos = Modulos.objects.filter(idcapacitacion=capacitacion)
                lecciones = Lecciones.objects.filter(idmodulo__in=modulos)
                preguntas = PreguntasLecciones.objects.filter(id_leccion__in=lecciones).order_by('id')
                
                preguntas_abiertas_ids = set()
                for pregunta in preguntas:
                    preguntas_dict[pregunta.id] = pregunta.pregunta
                    if pregunta.tipopregunta == 'pregunta_abierta':
                        preguntas_abiertas_ids.add(pregunta.id)
                
                # Agregar columnas de preguntas y respuestas
                for pregunta_id, pregunta_texto in preguntas_dict.items():
                    headers.append(f"Pregunta: {pregunta_texto}")  # Truncar texto largo
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            # Datos de colaboradores
            row = 7
            for progreso in progreso_qs:
                colaborador = progreso.colaborador
                
                # Solo incluir colaboradores activos
                if colaborador.estadocolaborador != 1:
                    continue
                
                # Obtener datos de las relaciones
                centro_op = colaborador.centroop.nombrecentrop if colaborador.centroop else 'N/A'
                proyecto = colaborador.centroop.id_proyecto.nombreproyecto if colaborador.centroop and colaborador.centroop.id_proyecto else 'N/A'
                unidad_obj = colaborador.centroop.id_proyecto.id_unidad if colaborador.centroop and colaborador.centroop.id_proyecto else None
                unidad_nombre = unidad_obj.nombreunidad if unidad_obj else 'N/A'
                unidad_desc = unidad_obj.descripcionunidad if unidad_obj else 'N/A'
                empresa = unidad_obj.id_empresa.nombre_empresa if unidad_obj and unidad_obj.id_empresa else 'N/A'
                cargo = colaborador.cargocolaborador.nombrecargo if colaborador.cargocolaborador else 'N/A'
                
                ws.cell(row=row, column=1).value = empresa
                ws.cell(row=row, column=2).value = unidad_nombre
                ws.cell(row=row, column=3).value = unidad_desc
                ws.cell(row=row, column=4).value = proyecto
                ws.cell(row=row, column=5).value = centro_op
                ws.cell(row=row, column=6).value = colaborador.cccolaborador
                ws.cell(row=row, column=7).value = cargo
                ws.cell(row=row, column=8).value = colaborador.nombrecolaborador
                ws.cell(row=row, column=9).value = colaborador.apellidocolaborador
                ws.cell(row=row, column=10).value = colaborador.correocolaborador
                ws.cell(row=row, column=11).value = float(progreso.progreso)
                ws.cell(row=row, column=12).value = progreso.fecha_registro.strftime('%d/%m/%Y %H:%M') if progreso.fecha_registro else 'N/A'
                ws.cell(row=row, column=13).value = progreso.fecha_completada.strftime('%d/%m/%Y %H:%M') if progreso.fecha_completada else 'N/A'
                
                # Determinar estado: Completada, No Completado (si pasó fecha fin) o En Progreso
                if progreso.completada == 1:
                    estado = "Completada"
                elif capacitacion.estado == 0:
                    estado = "No Completado"
                else:
                    estado = "En Progreso"
                ws.cell(row=row, column=14).value = estado
                
                # Si se incluyen preguntas, agregar respuestas del colaborador
                if include_questions:
                    col_offset = 14
                    respuestas_colaborador = RespuestasColaboradores.objects.filter(
                        idcolaborador=colaborador
                    ).select_related('idpregunta', 'idrespuesta')
                    
                    # Crear diccionario de respuestas por pregunta_id
                    # Preguntas abiertas muestran el texto escrito; cerradas muestran la opción elegida
                    respuestas_map = {
                        r.idpregunta_id: (
                            r.texto_respuesta if r.idpregunta_id in preguntas_abiertas_ids else r.idrespuesta.valor
                        )
                        for r in respuestas_colaborador
                    }
                    
                    for pregunta_id, pregunta_texto in preguntas_dict.items():
                        # Columna de pregunta
                        #ws.cell(row=row, column=col_offset + 1).value = pregunta_texto
                        # Columna de respuesta
                        respuesta = respuestas_map.get(pregunta_id, '')
                        ws.cell(row=row, column=col_offset + 1).value = respuesta
                        col_offset += 1
                
                # Formato
                num_cols = len(headers)
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    if col == 10:  # Porcentaje
                        cell.alignment = center_alignment
                        cell.number_format = '0.00"%"'
                    elif col in [11, 12]:  # Fechas
                        cell.alignment = center_alignment
                    elif col == 13:  # Estado Avance
                        cell.alignment = center_alignment
                
                row += 1
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 25  # Empresa
            ws.column_dimensions['B'].width = 18  # Unidad
            ws.column_dimensions['C'].width = 25  # Descripción Unidad
            ws.column_dimensions['D'].width = 20  # Proyecto
            ws.column_dimensions['E'].width = 18  # Centro Op
            ws.column_dimensions['F'].width = 15  # Cédula
            ws.column_dimensions['G'].width = 18  # Cargo
            ws.column_dimensions['H'].width = 15  # Nombre
            ws.column_dimensions['I'].width = 15  # Apellido
            ws.column_dimensions['J'].width = 25  # Correo
            ws.column_dimensions['K'].width = 16  # % Completación
            ws.column_dimensions['L'].width = 18  # Fecha Registro
            ws.column_dimensions['M'].width = 18  # Fecha Completación
            ws.column_dimensions['N'].width = 14  # Estado Avance
            
            # Ajustar ancho para columnas de preguntas y respuestas
            #if include_questions:
                #for col_idx in range(14, len(headers) + 1):
                    #col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)
                    #ws.column_dimensions[col_letter].width = 30
            
            # Habilitar filtros (facilita filtrado en Excel)
            try:
                last_row = row - 1
                last_col = len(headers)
                last_col_letter = chr(64 + last_col) if last_col <= 26 else chr(64 + last_col // 26) + chr(64 + last_col % 26)
                ws.auto_filter.ref = f"A6:{last_col_letter}{last_row}"
            except Exception:
                pass

            # Guardar a bytes
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Crear respuesta
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="reporte_capacitacion_{capacitacion_id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response
            
        except Capacitaciones.DoesNotExist:
            return Response(
                {'error': 'Capacitación no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Error al generar reporte de capacitación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @staticmethod
    def _calcular_duracion_total(capacitacion_id):
        """Suma los campos duracion (formato MM:SS o HH:MM) de todas las lecciones de una capacitación."""
        duraciones = Lecciones.objects.filter(
            idmodulo__idcapacitacion_id=capacitacion_id
        ).values_list('duracion', flat=True)
        total_seg = 0
        for dur in duraciones:
            if not dur:
                continue
            try:
                partes = str(dur).strip().split(':')
                if len(partes) == 2:
                    total_seg += int(partes[0]) * 60 + int(partes[1])
                elif len(partes) == 1:
                    total_seg += int(partes[0]) * 60
            except (ValueError, AttributeError):
                pass
        if total_seg == 0:
            return 'N/A'
        horas = total_seg // 3600
        minutos = (total_seg % 3600) // 60
        segundos = total_seg % 60
        if horas > 0:
            return f"{horas}h {minutos:02d}m {segundos:02d}s"
        if minutos > 0:
            return f"{minutos}m {segundos:02d}s"
        return f"{segundos}s"

    def _generar_reporte_rango_fechas(self, fecha_inicio_str, fecha_fin_str):
        """
        Genera reporte de todas las capacitaciones en un rango de fechas.
        Excluye capacitaciones con estado 3 (eliminadas).
        No cuenta colaboradores desactivados.
        """
        try:
            # Parsear fechas
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            except ValueError:
                return Response(
                    {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Obtener capacitaciones en rango (excluyendo estado 3)
            capacitaciones = Capacitaciones.objects.filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin
            ).exclude(estado=3).order_by('fecha_creacion')
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte General"
            
            # Estilos
            header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Título
            ws['A1'] = "REPORTE GENERAL DE CAPACITACIONES"
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
            
            # Headers principales
            headers = [
                "Capacitación",
                "Tipo",
                "Fecha Creación",
                "Duración Total",
                "Total Colaboradores",
                "Total Completados",
                "% Completación General",
                "Empresa",
                "Unidad",
                "Descripción Unidad",
                "Proyecto",
                "Centro Op",
                "Cédula",
                "Cargo",
                "Nombre",
                "Apellido",
                "Correo",
                "% Completación",
                "Estado Avance"
            ]
            
            row = 4
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            # Datos
            row = 5
            for capacitacion in capacitaciones:
                # Calcular duración total de lecciones para esta capacitación
                duracion_total = self._calcular_duracion_total(capacitacion.id)

                # Obtener colaboradores inscritos en esta capacitación (solo activos)
                progreso_qs = progresoCapacitaciones.objects.filter(
                    capacitacion=capacitacion
                ).select_related(
                    'colaborador',
                    'colaborador__centroop',
                    'colaborador__centroop__id_proyecto',
                    'colaborador__centroop__id_proyecto__id_unidad',
                    'colaborador__centroop__id_proyecto__id_unidad__id_empresa'
                ).order_by('colaborador__nombrecolaborador')

                if progreso_qs.exists():
                    # Calcular totales una sola vez por capacitación
                    total_colaboradores = progreso_qs.count()
                    total_completados = progreso_qs.filter(completada=1).count()
                    porcentaje_general = (total_completados / total_colaboradores * 100) if total_colaboradores > 0 else 0

                    for i, progreso in enumerate(progreso_qs):
                        colaborador = progreso.colaborador

                        # Saltar colaboradores desactivados
                        if colaborador.estadocolaborador != 1:
                            continue

                        # Obtener datos de las relaciones
                        centro_op = colaborador.centroop.nombrecentrop if colaborador.centroop else 'N/A'
                        proyecto = colaborador.centroop.id_proyecto.nombreproyecto if colaborador.centroop and colaborador.centroop.id_proyecto else 'N/A'
                        unidad_obj = colaborador.centroop.id_proyecto.id_unidad if colaborador.centroop and colaborador.centroop.id_proyecto else None
                        unidad_nombre = unidad_obj.nombreunidad if unidad_obj else 'N/A'
                        unidad_desc = unidad_obj.descripcionunidad if unidad_obj else 'N/A'
                        empresa = unidad_obj.id_empresa.nombre_empresa if unidad_obj and unidad_obj.id_empresa else 'N/A'
                        cargo = colaborador.cargocolaborador.nombrecargo if colaborador.cargocolaborador else 'N/A'

                        # Mostrar info de capacitación en TODAS las filas
                        ws.cell(row=row, column=1).value = capacitacion.titulo
                        ws.cell(row=row, column=2).value = capacitacion.tipo
                        ws.cell(row=row, column=3).value = capacitacion.fecha_creacion.strftime('%d/%m/%Y') if capacitacion.fecha_creacion else 'N/A'
                        ws.cell(row=row, column=4).value = duracion_total
                        ws.cell(row=row, column=5).value = total_colaboradores
                        ws.cell(row=row, column=6).value = total_completados
                        ws.cell(row=row, column=7).value = porcentaje_general

                        # Datos del colaborador
                        ws.cell(row=row, column=8).value = empresa
                        ws.cell(row=row, column=9).value = unidad_nombre
                        ws.cell(row=row, column=10).value = unidad_desc
                        ws.cell(row=row, column=11).value = proyecto
                        ws.cell(row=row, column=12).value = centro_op
                        ws.cell(row=row, column=13).value = colaborador.cccolaborador
                        ws.cell(row=row, column=14).value = cargo
                        ws.cell(row=row, column=15).value = colaborador.nombrecolaborador
                        ws.cell(row=row, column=16).value = colaborador.apellidocolaborador
                        ws.cell(row=row, column=17).value = colaborador.correocolaborador
                        ws.cell(row=row, column=18).value = float(progreso.progreso)

                        # Determinar estado: Completada, No Completado (si pasó fecha fin) o En Progreso
                        if progreso.completada == 1:
                            estado = "Completada"
                        elif capacitacion.estado == 0:
                            estado = "No Completado"
                        else:
                            estado = "En Progreso"
                        ws.cell(row=row, column=19).value = estado

                        # Aplicar bordes y formato
                        for col in range(1, 20):
                            cell = ws.cell(row=row, column=col)
                            cell.border = border
                            if col in [7, 18]:  # Porcentajes
                                cell.alignment = center_alignment
                                cell.number_format = '0.00"%"'
                            elif col in [4, 19]:  # Duración Total y Estado
                                cell.alignment = center_alignment
                            else:
                                cell.alignment = Alignment(vertical='center')

                        row += 1
                else:
                    # Capacitación sin colaboradores inscritos
                    ws.cell(row=row, column=1).value = capacitacion.titulo
                    ws.cell(row=row, column=2).value = capacitacion.tipo
                    ws.cell(row=row, column=3).value = capacitacion.fecha_creacion.strftime('%d/%m/%Y') if capacitacion.fecha_creacion else 'N/A'
                    ws.cell(row=row, column=4).value = duracion_total
                    ws.cell(row=row, column=5).value = "Sin colaboradores"

                    for col in range(1, 20):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border

                    row += 1
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 25  # Capacitación
            ws.column_dimensions['B'].width = 15  # Tipo
            ws.column_dimensions['C'].width = 16  # Fecha Creación
            ws.column_dimensions['D'].width = 16  # Duración Total
            ws.column_dimensions['E'].width = 18  # Total Colaboradores
            ws.column_dimensions['F'].width = 18  # Total Completados
            ws.column_dimensions['G'].width = 20  # % Completación General
            ws.column_dimensions['H'].width = 15  # Empresa
            ws.column_dimensions['I'].width = 15  # Unidad
            ws.column_dimensions['J'].width = 25  # Descripción Unidad
            ws.column_dimensions['K'].width = 15  # Proyecto
            ws.column_dimensions['L'].width = 18  # Centro Op
            ws.column_dimensions['M'].width = 15  # Cédula
            ws.column_dimensions['N'].width = 18  # Cargo
            ws.column_dimensions['O'].width = 18  # Nombre
            ws.column_dimensions['P'].width = 18  # Apellido
            ws.column_dimensions['Q'].width = 25  # Correo
            ws.column_dimensions['R'].width = 16  # % Completación
            ws.column_dimensions['S'].width = 14  # Estado Avance

            # Habilitar filtros (facilita filtrado en Excel)
            try:
                last_row = row - 1
                ws.auto_filter.ref = f"A4:S{last_row}"
            except Exception:
                pass

            # Guardar a bytes
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Crear respuesta
            response = FileResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="reporte_capacitaciones_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar reporte de rango de fechas: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class InduccionesView(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdmin | IsUsuarioEspecial]
    """
    Listado de inducciones disponibles para crear usuarios temporales.
    GET /capacitaciones/inducciones/ → Devuelve todas las capacitaciones tipo INDUCCIÓN CORPORATIVA (sin importar estado)
    Solo usuarios tipo 3+ (staff) pueden acceder.
    """
    def get(self, request, *args, **kwargs):
            inducciones = Capacitaciones.objects.filter(tipo__in=["INDUCCIÓN CORPORATIVA", "ENCUESTA"]).values('id', 'titulo', 'tipo', 'estado', 'descripcion').order_by('tipo', 'titulo')
            return Response(list(inducciones), status=status.HTTP_200_OK)
