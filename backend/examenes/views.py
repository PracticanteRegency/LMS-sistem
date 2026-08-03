from django.shortcuts import get_object_or_404
from django.core.cache import cache
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from django.utils import timezone
import io
import uuid
import csv
import logging
from django.db.models import F, Prefetch

from usuarios.models import Cargo
from usuarios.permissions import IsUsuarioEspecial, IsSuperAdmin
from analitica.models import Epresa, Unidadnegocio, Proyecto, Centroop
from examenes.models import Examen, ExamenesCargo

from django.core.mail import send_mail, EmailMultiAlternatives
from django.conf import settings

from .models import ExamenesCargo, CorreoExamenEnviado, RegistroExamenes, Examen, ExamenTrabajador

# Tipos de examen válidos
TIPOS_EXAMEN_VALIDOS = ['INGRESO', 'PERIODICO', 'RETIRO', 'ESPECIAL', 'POST_INCAPACIDAD', 'ALTURAS']

from .serializers import (
    CrearExamenSerializer,
    EnviarCorreoSerializer,
    EmpresaConCargosSerializer,
    ReporteCorreoSerializer,
    EnviarCorreoMasivoSerializer,
    ActualizarEstadoExamenesSerializer,
)


class ReporteCorreosEnviadosView(APIView):
    """
    Vista para obtener reporte de correos enviados.
    Retorna lista paginada de correos con información resumida.

    GET: Lista de correos enviados ordenados por fecha descendente.
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def get(self, request):
        """Obtiene lista paginada de correos enviados (con cache)."""
        page = request.query_params.get('page', '1')
        page_size = request.query_params.get('page_size', '25')
        cache_key = f"reporte_correos_page={page}_size={page_size}"

        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached, status=status.HTTP_200_OK, headers={'X-Cache': 'HIT'})

        # Obtener correos con datos relacionados optimizados
        correos = self._get_correos_queryset()

        # Aplicar paginación
        paginated_data = self._paginate_correos(correos, request)

        if paginated_data:
            # Guardar en cache y marcar MISS
            cache.set(cache_key, paginated_data.data, timeout=300)
            paginated_data['X-Cache'] = 'MISS'
            return paginated_data

        # Fallback sin paginación
        serializer = ReporteCorreoSerializer(correos, many=True)
        data = serializer.data
        cache.set(cache_key, data, timeout=300)
        return Response(data, status=status.HTTP_200_OK, headers={'X-Cache': 'MISS'})

    def _get_correos_queryset(self):
        """Construye queryset optimizado de correos.
        
        NOTA: Filtra solo correos enviados correctamente (enviado_correctamente=True).
        Los correos fallidos NO aparecen en el reporte hasta que se corrijan.
        """
        return CorreoExamenEnviado.objects.filter(
            enviado_correctamente=True
        ).select_related(
            'enviado_por'
        ).prefetch_related(
            'trabajadores'
        ).order_by('-fecha_envio')

    def _paginate_correos(self, correos, request):
        """Aplica paginación al queryset de correos."""
        paginator = PageNumberPagination()
        paginator.page_size = 25
        paginator.page_size_query_param = 'page_size'
        paginator.max_page_size = 100
        paginated_correos = paginator.paginate_queryset(
            correos, request, view=self)

        if paginated_correos is not None:
            serializer = ReporteCorreoSerializer(paginated_correos, many=True)
            return paginator.get_paginated_response(serializer.data)

        return None


class DetalleCorreoEnviadoView(APIView):
    """
    Vista para obtener detalles completos de un correo enviado.

    GET: Retorna información detallada de un envío específico.
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def get(self, request, correo_id):
        """Retorna metadata del correo y el listado de trabajadores (RegistroExamenes) asociados, con paginación estándar (count, next, previous, results) y cache."""
        correo = get_object_or_404(
            CorreoExamenEnviado.objects.select_related('enviado_por'),
            id=correo_id
        )

        # Obtener trabajadores del correo
        trabajadores = RegistroExamenes.objects.filter(
            correo_lote=correo
        ).select_related('empresa', 'cargo').order_by('-fecha_registro')

        # Cache por correo_id + paginación
        page = request.query_params.get('page', '1')
        page_size = request.query_params.get('page_size', '25')
        cache_key = f"detalle_correo={correo_id}_page={page}_size={page_size}"

        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached, status=status.HTTP_200_OK, headers={'X-Cache': 'HIT'})

        # Paginar resultados
        paginator = PageNumberPagination()
        paginator.page_size = 25
        paginator.page_size_query_param = 'page_size'
        paginator.max_page_size = 100
        paginated_trabajadores = paginator.paginate_queryset(trabajadores, request, view=self)

        from .serializers import ListarTrabajadoresCorreoSerializer

        if paginated_trabajadores is not None:
            # Serializar página actual
            serializer = ListarTrabajadoresCorreoSerializer(paginated_trabajadores, many=True)
            # Respuesta estándar de DRF: count, next, previous, results
            paginated_response = paginator.get_paginated_response(serializer.data)
            # Agregar metadata del correo
            paginated_response.data.update({
                "correo_id": correo.id,
                "uuid_correo": getattr(correo, 'uuid_correo', None),
                "asunto": correo.asunto,
                "fecha_envio": getattr(correo, 'fecha_envio', None),
                "total_trabajadores": trabajadores.count()
            })
            paginated_response['X-Cache'] = 'MISS'
            cache.set(cache_key, paginated_response.data, timeout=300)
            return paginated_response

        # Fallback sin paginación (poco probable): devolver estructura similar
        serializer = ListarTrabajadoresCorreoSerializer(trabajadores, many=True)
        data = {
            "count": len(serializer.data),
            "next": None,
            "previous": None,
            "results": serializer.data,
            "correo_id": correo.id,
            "uuid_correo": getattr(correo, 'uuid_correo', None),
            "asunto": correo.asunto,
            "fecha_envio": getattr(correo, 'fecha_envio', None),
            "total_trabajadores": trabajadores.count()
        }
        cache.set(cache_key, data, timeout=300)
        return Response(data, status=status.HTTP_200_OK, headers={'X-Cache': 'MISS'})


# ──────────────────────────────────────────────────────────────
#  Funciones auxiliares para emails de Exámenes Médicos
# ──────────────────────────────────────────────────────────────

def crear_email_medico(subject, body, destinatarios, html_body=None):
    """
    Crea un email usando las credenciales de Exámenes Médicos (360 CloudRegency).
    
    Args:
        subject: Asunto del email
        body: Cuerpo en texto plano
        destinatarios: List o tuple de direcciones de email
        html_body: Cuerpo en HTML (opcional)
    
    Returns:
        EmailMultiAlternatives: Email listo para enviar con config médica
    """
    email = EmailMultiAlternatives(
        subject=subject,
        body=body,
        from_email=settings.EMAIL_MEDICAL_FROM_EMAIL,
        to=destinatarios if isinstance(destinatarios, (list, tuple)) else [destinatarios]
    )
    
    if html_body:
        email.attach_alternative(html_body, "text/html")
    
    return email


def enviar_email_medico(subject, body, destinatarios, html_body=None, fail_silently=False):
    """
    Envía un email usando las credenciales de Exámenes Médicos (360 CloudRegency).
    
    Nota: Esta función usa smtplib directamente para aplicar las credenciales médicas.
    
    Args:
        subject: Asunto del email
        body: Cuerpo en texto plano
        destinatarios: List, tuple o string de direcciones de email
        html_body: Cuerpo en HTML (opcional)
        fail_silently: Si True, silencia excepciones
    
    Returns:
        int: Número de emails enviados exitosamente
        
    Raises:
        Exception: Si fail_silently=False y hay error
    """
    import smtplib
    import ssl
    
    logger = logging.getLogger(__name__)
    
    # Asegurar que destinatarios sea una lista
    if isinstance(destinatarios, str):
        destinatarios = [destinatarios]
    else:
        destinatarios = list(destinatarios)
    
    if not destinatarios:
        logger.warning("enviar_email_medico: Lista de destinatarios vacía")
        return 0
    
    try:
        # Crear email con configuración médica
        email = crear_email_medico(subject, body, destinatarios, html_body)
        
        # Enviar usando las credenciales médicas
        timeout = getattr(settings, 'EMAIL_MEDICAL_TIMEOUT', 30)
        
        if settings.EMAIL_MEDICAL_USE_SSL:
            context = ssl.create_default_context()
            server = smtplib.SMTP_SSL(
                settings.EMAIL_MEDICAL_HOST,
                settings.EMAIL_MEDICAL_PORT,
                timeout=timeout,
                context=context
            )
        else:
            server = smtplib.SMTP(
                settings.EMAIL_MEDICAL_HOST,
                settings.EMAIL_MEDICAL_PORT,
                timeout=timeout
            )
            if settings.EMAIL_MEDICAL_USE_TLS:
                server.starttls()
        
        server.login(settings.EMAIL_MEDICAL_HOST_USER, settings.EMAIL_MEDICAL_HOST_PASSWORD)
        server.send_message(email.message())
        server.quit()
        
        logger.info(
            f"enviar_email_medico: Email enviado exitosamente a {len(destinatarios)} "
            f"destinatarios (asunto: {subject})"
        )
        
        return len(destinatarios)
        
    except Exception as e:
        error_msg = f"{type(e).__name__}: {str(e)}"
        logger.error(f"enviar_email_medico: Error enviando email: {error_msg}", exc_info=True)
        
        if not fail_silently:
            raise
        
        return 0


class EnviarCorreoView(APIView):
    """
    Vista para enviar correo individual de convocatoria a exámenes médicos.

    POST: Envía correo a un trabajador con los exámenes seleccionados por el frontend.
    
    JSON esperado:
    {
        "nombre_trabajador": "Juan Pérez",
        "documento_trabajador": "123456789",
        "correo_destino": "juan@email.com",
        "centro_id": 5,
        "cargo_id": 10,
        "tipo_examen": "INGRESO",
        "examenes_ids": [1, 4, 7, 12]
    }
    
    Flujo:
    1. Valida datos y deriva empresa desde centro
    2. Crea CorreoExamenEnviado (lote)
    3. Crea RegistroExamenes (trabajador)
    4. Crea ExamenTrabajador por cada examen (relación M2M)
    5. Crea RegistroExamenesEnviados por cada examen (trazabilidad)
    6. Envía el correo
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def post(self, request):
        """Procesa y envía correo individual de exámenes médicos."""
        # Log raw request for debugging (verificar payload enviado por frontend)
        logger = logging.getLogger(__name__)
        try:
            logger.info(f"Raw request.body: {request.body}")
        except Exception:
            logger.info("Raw request.body: <unavailable>")
        logger.info(f"Raw request.data (parsed): {request.data}")

        # Validar datos de entrada
        serializer = EnviarCorreoSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        # Log payload para debugging (verificar si 'ciudad' llega)
        logger = logging.getLogger(__name__)
        logger.info(f"EnviarCorreo payload validated: {data}")

        # Obtener colaborador autenticado
        enviado_por = self._get_colaborador(request)
        
        # Obtener correo del solicitante (colaborador autenticado)
        correo_solicitante = self._get_correo_colaborador(enviado_por)
        
        # Obtener solicitante extra por ID de colaborador
        solicitante_extra = None
        solicitante_extra_id = data.get('solicitante_extra_id')
        if solicitante_extra_id:
            from usuarios.models import Colaboradores
            try:
                colab_extra = Colaboradores.objects.get(idcolaborador=solicitante_extra_id)
                solicitante_extra = self._get_correo_colaborador(colab_extra)
            except Colaboradores.DoesNotExist:
                logger.warning(f"Colaborador extra con id {solicitante_extra_id} no encontrado")

        # Forzar destinatarios fijos para este endpoint
        correos_destino_fixed = (
            ""
            #"practicante.desarrollogh@regency.com.co,"
            #"coordinador.seleccion@regency.com.co,"
            "operativo@servicompetentes.com,"
            "administrativo@servicompetentes.com"  
        )
        # Limpiar agresivamente: remover espacios, newlines, tabs
        correos_list_fixed = [
            e.strip().replace('\r', '').replace('\n', '').replace('\t', '')
            for e in correos_destino_fixed.split(',') 
            if e and e.strip()
        ]
        
        # Agregar solicitante y solicitante extra a la lista de destinatarios
        correos_con_solicitantes = correos_list_fixed.copy()
        if correo_solicitante:
            correos_con_solicitantes.append(correo_solicitante)
        if solicitante_extra:
            correos_con_solicitantes.append(solicitante_extra)
        
        # Construir el string de correos para almacenar en BD (incluye fijos + solicitantes)
        correos_para_bd = ', '.join(correos_con_solicitantes)
        data['correo_destino'] = correos_para_bd

        # Obtener centro y derivar empresa
        centro = get_object_or_404(
            Centroop.objects.select_related('id_proyecto__id_unidad__id_empresa'),
            idcentrop=data['centro_id']
        )
        empresa = centro.id_proyecto.id_unidad.id_empresa
        cargo = get_object_or_404(Cargo, idcargo=data['cargo_id'])
        tipo_examen = data['tipo_examen']

        # Validar y obtener exámenes desde IDs enviados por frontend
        examenes = self._get_examenes_por_ids(data['examenes_ids'])
        if not examenes:
            return Response(
                {"error": "No se encontraron exámenes válidos con los IDs proporcionados"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Crear registros en BD (CorreoExamenEnviado, RegistroExamenes, ExamenTrabajador, RegistroExamenesEnviados)
        correo_obj, registro_trabajador = self._crear_registros_completos(
            enviado_por=enviado_por,
            data=data,
            empresa=empresa,
            cargo=cargo,
            centro=centro,
            tipo_examen=tipo_examen,
            examenes=examenes,
            colaborador=enviado_por,
            correos_solicitantes={'principal': correo_solicitante, 'extra': solicitante_extra},
            es_masivo=False  # EnviarCorreoView siempre es individual
        )

        # Enviar correo a la lista completa (fijos + solicitante + solicitante extra)
        resultado = self._enviar_correo(
            correo_obj,
            correos_con_solicitantes,
            data['nombre_trabajador'],
            registro_trabajador,
            ciudad=data.get('ciudad'),
            correos_solicitantes={'principal': correo_solicitante, 'extra': solicitante_extra}
        )

        return resultado

    def _get_colaborador(self, request):
        """Obtiene el colaborador del usuario autenticado."""
        colaborador = getattr(request.user, 'idcolaboradoru', None)
        if colaborador is None:
            colaborador = getattr(request.user, 'id_colaboradoru', None)
        return colaborador

    def _get_correo_colaborador(self, colaborador):
        """Obtiene el correo electrónico del colaborador."""
        if colaborador is None:
            return None
        
        # Intentar obtener correo con diferentes nombres de campo
        correo = getattr(colaborador, 'correocolaborador', None) or \
                 getattr(colaborador, 'correo', None) or \
                 getattr(colaborador, 'email', None)
        
        return correo.strip() if correo else None

    def _clear_cache(self):
        """Limpia el cache de reportes de correos y datos de empresas con exámenes."""
        logger = logging.getLogger(__name__)
        try:
            # Limpiar todos los caches de reportes de correos paginados
            # Eliminamos el patrón completo del cache
            cache.delete_pattern("reporte_correos_page=*")
            
            # Limpiar todos los caches de detalles de correo
            cache.delete_pattern("detalle_correo=*")
            
            # Limpiar cache de datos de empresas con exámenes
            cache.delete('cargo_empresa_examenes_data')
            
            logger.info("Cache limpiado: reportes de correos y datos de empresas")
        except Exception as e:
            logger.warning(f"Error al limpiar cache: {str(e)}")

    def _get_medical_email_backend(self):
        """
        Retorna una conexión SMTP configurada con credenciales de Exámenes Médicos.
        
        Returns:
            SMTPConnection: Conexión SMTP con configuración médica (360 CloudRegency)
        """
        from django.core.mail.backends.smtp import EmailBackend
        
        return EmailBackend(
            host=settings.EMAIL_MEDICAL_HOST,
            port=settings.EMAIL_MEDICAL_PORT,
            username=settings.EMAIL_MEDICAL_HOST_USER,
            password=settings.EMAIL_MEDICAL_HOST_PASSWORD,
            use_ssl=settings.EMAIL_MEDICAL_USE_SSL,
            use_tls=settings.EMAIL_MEDICAL_USE_TLS,
            timeout=settings.EMAIL_MEDICAL_TIMEOUT,
            fail_silently=False
        )

    def _get_examenes_por_ids(self, examenes_ids):
        """Obtiene exámenes activos por lista de IDs."""
        return list(Examen.objects.filter(
            id_examen__in=examenes_ids,
            activo=True
        ))

    def _crear_registros_completos(
            self, enviado_por, data, empresa, cargo, centro, tipo_examen, examenes, colaborador=None, correos_solicitantes=None, es_masivo=False):
        """
        Crea todos los registros necesarios en la base de datos:
        1. CorreoExamenEnviado (lote)
        2. RegistroExamenes (trabajador)
        3. ExamenTrabajador (relación M2M por cada examen)
        4. RegistroExamenesEnviados (trazabilidad por cada examen)
        """
        asunto = f"Exámenes médicos – {data['nombre_trabajador']} ({tipo_examen})"

        # 1. Crear CorreoExamenEnviado (genera UUID automáticamente)
        correo_obj = CorreoExamenEnviado(
            enviado_por=enviado_por,
            asunto=asunto,
            cuerpo_correo="",
            correos_destino=data['correo_destino'],
            tipo_examen=tipo_examen,
            enviado_correctamente=False
        )
        correo_obj.save()

        # 2. Crear RegistroExamenes (trabajador)
        # Campo examenes_asignados se llena con nombres para compatibilidad
        nombres_examenes = ", ".join([e.nombre for e in examenes])
        registro_trabajador = RegistroExamenes.objects.create(
            correo_lote=correo_obj,
            nombre_trabajador=data['nombre_trabajador'],
            documento_trabajador=data['documento_trabajador'],
            ciudad=data.get('ciudad'),
            empresa=empresa,
            cargo=cargo,
            centro=centro,
            tipo_examen=tipo_examen,
            examenes_asignados=nombres_examenes,
            estado_trabajador=0  # Pendiente
        )

        # 3. Crear ExamenTrabajador por cada examen (relación M2M - bulk_create)
        examenes_trabajador = [
            ExamenTrabajador(
                registro_examen=registro_trabajador,
                examen=examen
            )
            for examen in examenes
        ]
        ExamenTrabajador.objects.bulk_create(examenes_trabajador)

        # Nota: No creamos ni sincronizamos `RegistroExamenesEnviados` aquí.
        # Usamos `ExamenTrabajador` para representar los exámenes asignados
        # y `RegistroExamenes.estado_trabajador` como la fuente de verdad
        # del estado del trabajador (Pendiente/Completado).

        # 5. Construir cuerpo del correo
        cuerpo_correo = self._construir_cuerpo_correo(
            data=data,
            cargo=cargo,
            empresa=empresa,
            centro=centro,
            tipo_examen=tipo_examen,
            examenes=examenes,
            uuid_correo=correo_obj.uuid_correo,
            uuid_trabajador=registro_trabajador.uuid_trabajador,
            fecha_envio=correo_obj.fecha_envio,
            colaborador=colaborador,
            correos_solicitantes=correos_solicitantes,
            es_masivo=es_masivo  # Pasar el parámetro es_masivo
        )

        # 6. Actualizar correo con cuerpo completo
        correo_obj.cuerpo_correo = cuerpo_correo
        correo_obj.save()

        return correo_obj, registro_trabajador

    def _construir_cuerpo_correo(
            self, data, cargo, empresa, centro, tipo_examen, examenes, uuid_correo, uuid_trabajador, fecha_envio, colaborador=None, correos_solicitantes=None, es_masivo=False):
        """
        Construye el cuerpo del correo según si es individual o masivo.
        
        Args:
            es_masivo: bool - True si hay >1 trabajador, False si es individual
        """
        # Construir lista de exámenes
        lista_examenes = "\n".join([f"- {e.nombre}" for e in examenes])

        # Mapeo de tipos a nombres legibles
        tipos_legibles = {
            'INGRESO': 'Examen de Ingreso',
            'PERIODICO': 'Examen Periódico',
            'RETIRO': 'Examen de Retiro',
            'ESPECIAL': 'Examen Especial',
            'POST_INCAPACIDAD': 'Examen Post-Incapacidad',
            'ALTURAS': 'Examen con énfasis en alturas'
        }
        tipo_legible = tipos_legibles.get(tipo_examen, tipo_examen)

        # Resolver nombre y correo del colaborador (soporta instancia Colaboradores o valor simple)
        nombre_colaborador = None
        correo_colaborador = None
        try:
            if colaborador is not None:
                nombre_colaborador = getattr(colaborador, 'nombrecolaborador', None) or getattr(colaborador, 'nombre', None)
                correo_colaborador = getattr(colaborador, 'correocolaborador', None) or getattr(colaborador, 'correo', None) or getattr(colaborador, 'email', None)
        except Exception:
            nombre_colaborador = None
            correo_colaborador = None

        if not nombre_colaborador:
            nombre_colaborador = data.get('nombre_trabajador') or 'No disponible'
        if not correo_colaborador:
            correo_colaborador = 'No disponible'

        # Construir la línea de correos del solicitante
        correos_solicitantes_text = correo_colaborador
        if correos_solicitantes:
            correos_adicionales = []
            if correos_solicitantes.get('extra'):
                correos_adicionales.append(correos_solicitantes['extra'])
            if correos_adicionales:
                correos_solicitantes_text = f"{correo_colaborador}, {', '.join(correos_adicionales)}"

        # DIFERENTE SEGÚN SI ES INDIVIDUAL O MASIVO
        if es_masivo:
            # Cuerpo para correo masivo (múltiples trabajadores)
            cuerpo = (
                f"Cordial Saludo.\n\n"
                f"Se han programado los siguientes exámenes médicos para los trabajadores en el excel adjunto.\n\n"
                f"---\n"
                f"ID de Seguimiento: {uuid_correo}\n"
                f"Solicitante: {nombre_colaborador}\n"
                f"Correo del solicitante: {correos_solicitantes_text}"
            )
        else:
            # Cuerpo para correo individual (1 trabajador)
            cuerpo = (
                f"Cordial Saludo.\n\n"
                f"Se han programado los siguientes exámenes médicos para el trabajador:\n\n"
                f"Nombre: {data['nombre_trabajador']}\n"
                f"Documento: {data['documento_trabajador']}\n"
                f"Ciudad: {data.get('ciudad', 'No disponible')}\n"
                f"Cargo: {cargo.nombrecargo}\n"
                f"Empresa: {empresa.nombre_empresa}\n"
                f"Centro Operativo: {getattr(centro, 'nombrecentrop', str(centro))}\n"
                f"Tipo de Examen: {tipo_legible}\n\n"
                f"Exámenes requeridos:\n{lista_examenes}\n\n"
                f"---\n"
                f"ID de Lote: {uuid_correo}\n"
                f"ID de Trabajador: {uuid_trabajador}\n"
                f"Solicitante: {nombre_colaborador}\n"
                f"Correo del solicitante: {correos_solicitantes_text}"
            )
        return cuerpo

    def _enviar_correo(self, correo_obj, destinatario, nombre_trabajador, registro_trabajador, ciudad=None, correos_solicitantes=None):
        """Envía el correo y actualiza el estado del registro.

        Acepta `ciudad` únicamente para incluirla en logs y en la respuesta
        (no se guarda en `RegistroExamenes`).
        
        `correos_solicitantes` es un dict con 'principal' y 'extra' para incluir en logs.
        
        Si el envío SMTP falla, programa un reintento automático vía Celery.
        """
        logger = logging.getLogger(__name__)
        try:
            # `destinatario` puede ser un string o una lista de emails
            recipient_list = destinatario if isinstance(destinatario, (list, tuple)) else [destinatario]

            # Filtrar emails vacíos o inválidos y limpiar agresivamente espacios/newlines
            recipient_list = [
                e.strip().replace('\r', '').replace('\n', '').replace('\t', '')
                for e in recipient_list 
                if e and e.strip()
            ]
            if not recipient_list:
                correo_obj.enviado_correctamente = False
                correo_obj.error_envio = "Lista de destinatarios vacía después de filtrar"
                correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])
                logger.error(f"Enviar correo - Sin destinatarios válidos para UUID: {correo_obj.uuid_correo}")
                return Response(
                    {"error": "No hay destinatarios válidos para enviar el correo",
                     "uuid_lote": correo_obj.uuid_correo},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Log del cuerpo para verificación
            logger.info(f"Enviar correo - UUID lote: {correo_obj.uuid_correo}")
            logger.info(f"Enviar correo - Ciudad (desde JSON): {ciudad}")
            logger.info(f"Enviar correo - Destinatarios ({len(recipient_list)}): {recipient_list}")
            if correos_solicitantes:
                logger.info(f"Enviar correo - Solicitante principal: {correos_solicitantes.get('principal')}")
                logger.info(f"Enviar correo - Solicitante extra: {correos_solicitantes.get('extra')}")

            # Verificar conexión SMTP antes de intentar enviar
            from examenes.tasks import verificar_conexion_smtp
            smtp_ok, smtp_error = verificar_conexion_smtp()
            if not smtp_ok:
                logger.warning(
                    f"Enviar correo - SMTP no disponible: {smtp_error}. "
                    f"Programando reintento vía Celery para correo id={correo_obj.id}"
                )
                correo_obj.enviado_correctamente = False
                correo_obj.error_envio = f"SMTP no disponible: {smtp_error}. Reintento programado."
                correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])

                # Programar reintento automático vía Celery (en 60 segundos)
                from examenes.tasks import enviar_correo_examen_task
                enviar_correo_examen_task.apply_async(
                    args=[correo_obj.id],
                    countdown=60
                )

                total_examenes = registro_trabajador.examenes.count()
                return Response(
                    {
                        "mensaje": "Correo registrado. El servidor de correo no está disponible temporalmente, se reintentará automáticamente.",
                        "uuid_lote": correo_obj.uuid_correo,
                        "uuid_trabajador": registro_trabajador.uuid_trabajador,
                        "trabajador": nombre_trabajador,
                        "destinatario": recipient_list,
                        "tipo_examen": correo_obj.tipo_examen,
                        "examenes_asignados": total_examenes,
                        "registro_id": registro_trabajador.id,
                        "ciudad": ciudad,
                        "reintento_programado": True,
                        "advertencia": "El correo se enviará automáticamente cuando el servidor de correo esté disponible."
                    },
                    status=status.HTTP_202_ACCEPTED
                )

            # Generar Excel si hay múltiples trabajadores
            excel_buffer = None
            cantidad_trabajadores = correo_obj.trabajadores.count()
            if cantidad_trabajadores > 1:
                excel_buffer = self._generar_excel_simple(correo_obj)
            
            # Enviar usando EmailMultiAlternatives con credenciales médicas
            email = EmailMultiAlternatives(
                subject=correo_obj.asunto,
                body=correo_obj.cuerpo_correo,
                from_email=settings.EMAIL_MEDICAL_FROM_EMAIL,
                to=recipient_list
            )
            
            # Adjuntar HTML como alternativa
            if '<html' in correo_obj.cuerpo_correo.lower() or '<p>' in correo_obj.cuerpo_correo.lower():
                email.attach_alternative(correo_obj.cuerpo_correo, "text/html")
            
            # Adjuntar Excel si existe
            if excel_buffer:
                email.attach(
                    'Trabajadores_Examenes.xlsx',
                    excel_buffer.getvalue(),
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            
            # Usar backend médico para enviar
            email.connection = self._get_medical_email_backend()
            email.send(fail_silently=False)

            # Marcar como enviado correctamente
            correo_obj.enviado_correctamente = True
            correo_obj.error_envio = None
            correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])

            # Contar exámenes asignados
            total_examenes = registro_trabajador.examenes.count()

            # Limpiar cache de reportes de correos y datos de empresas
            self._clear_cache()

            logger.info(
                f"Enviar correo - ÉXITO - UUID: {correo_obj.uuid_correo}, "
                f"Destinatarios: {len(recipient_list)}, Exámenes: {total_examenes}"
            )

            return Response(
                {
                    "mensaje": "Correo enviado correctamente",
                    "uuid_lote": correo_obj.uuid_correo,
                    "uuid_trabajador": registro_trabajador.uuid_trabajador,
                    "trabajador": nombre_trabajador,
                    "destinatario": recipient_list,
                    "tipo_examen": correo_obj.tipo_examen,
                    "examenes_asignados": total_examenes,
                    "registro_id": registro_trabajador.id,
                    "ciudad": ciudad,
                    "solicitantes_notificados": {
                        "principal": correos_solicitantes.get('principal') if correos_solicitantes else None,
                        "extra": correos_solicitantes.get('extra') if correos_solicitantes else None
                    }
                },
                status=status.HTTP_200_OK
            )

        except Exception as e:
            # Registrar error
            correo_obj.enviado_correctamente = False
            correo_obj.error_envio = f"{type(e).__name__}: {str(e)}"
            correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])

            logger.error(f"Error enviando correo UUID={correo_obj.uuid_correo}: {str(e)}", exc_info=True)

            # Programar reintento automático vía Celery
            try:
                from examenes.tasks import enviar_correo_examen_task
                enviar_correo_examen_task.apply_async(
                    args=[correo_obj.id],
                    countdown=120  # Reintentar en 2 minutos
                )
                logger.info(f"Reintento programado vía Celery para correo id={correo_obj.id}")
                reintento_msg = " Se reintentará automáticamente en 2 minutos."
            except Exception as celery_err:
                logger.error(f"No se pudo programar reintento Celery: {str(celery_err)}")
                reintento_msg = " No se pudo programar reintento automático."

            return Response(
                {
                    "error": f"Error al enviar el correo: {str(e)}.{reintento_msg}",
                    "uuid_lote": correo_obj.uuid_correo,
                    "uuid_trabajador": registro_trabajador.uuid_trabajador,
                    "reintento_programado": "Se reintentará" in reintento_msg
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _generar_excel_simple(self, correo_obj):
        """
        Genera Excel con los trabajadores del correo en el mismo formato que masivo.
        
        Args:
            correo_obj: instancia de CorreoExamenEnviado
            
        Returns:
            BytesIO: buffer del Excel generado
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Obtener trabajadores del lote
        registros = correo_obj.trabajadores.select_related(
            'empresa', 'cargo', 'centro'
        ).prefetch_related('examenes__examen').all()
        
        if not registros.exists():
            return None
        
        # Recopilar todos los exámenes únicos
        examenes_unicos = set()
        trabajadores_data = []
        
        for reg in registros:
            examenes_del_trabajador = []
            for et in reg.examenes.all():
                examenes_del_trabajador.append(et.examen.nombre)
                examenes_unicos.add(et.examen.nombre)
            
            trabajadores_data.append({
                'uuid': reg.uuid_trabajador,
                'nombre': reg.nombre_trabajador,
                'documento': reg.documento_trabajador,
                'empresa': reg.empresa.nombre_empresa if reg.empresa else '',
                'unidad': getattr(reg.empresa, 'id_unidad', None),  # relación si existe
                'proyecto': getattr(reg.centro, 'id_proyecto', None) if reg.centro else None,
                'centro': reg.centro.nombrecentrop if reg.centro else '',
                'ciudad': reg.ciudad or '',
                'cargo': reg.cargo.nombrecargo if reg.cargo else '',
                'tipo_examen': reg.tipo_examen,
                'examenes': examenes_del_trabajador,
            })
        
        # Crear Excel
        nombres_examenes = sorted(list(examenes_unicos))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Trabajadores Examenes"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border_style = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Encabezados (igual que masivo)
        headers = [
            "UUID",
            "Empresa",
            "Unidad",
            "Proyecto",
            "Centro",
            "Ciudad",
            "Cargo",
            "Nombre",
            "Documento",
            "Tipo Examen"
        ] + nombres_examenes
        ws.append(headers)
        
        # Aplicar estilos a encabezados
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border_style
        
        # Agregar datos (intentar obtener unidad y proyecto)
        for trab in trabajadores_data:
            unidad_nombre = ''
            proyecto_nombre = ''
            
            # Intentar obtener unidad desde proyecto del centro
            if trab['proyecto']:
                proyecto_nombre = trab['proyecto'].nombreproyecto
                if hasattr(trab['proyecto'], 'id_unidad'):
                    unidad_nombre = trab['proyecto'].id_unidad.nombreunidad
            
            row_data = [
                trab['uuid'],
                trab['empresa'],
                unidad_nombre,
                proyecto_nombre,
                trab['centro'],
                trab['ciudad'],
                trab['cargo'],
                trab['nombre'],
                trab['documento'],
                trab['tipo_examen']
            ]
            
            # Exámenes con X donde aplica
            examenes_trabajador = set(trab['examenes'])
            for nombre_examen in nombres_examenes:
                row_data.append("X" if nombre_examen in examenes_trabajador else "")
            
            ws.append(row_data)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20  # UUID
        ws.column_dimensions['B'].width = 25  # Empresa
        ws.column_dimensions['C'].width = 20  # Unidad
        ws.column_dimensions['D'].width = 25  # Proyecto
        ws.column_dimensions['E'].width = 20  # Centro
        ws.column_dimensions['F'].width = 15  # Ciudad
        ws.column_dimensions['G'].width = 25  # Cargo
        ws.column_dimensions['H'].width = 25  # Nombre
        ws.column_dimensions['I'].width = 15  # Documento
        ws.column_dimensions['J'].width = 15  # Tipo Examen
        
        # Ajustar ancho de columnas de exámenes (más pequeñas)
        for col_num in range(11, 11 + len(nombres_examenes)):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 5
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


class CargoEmpresaConExamenesView(APIView):
    """
    Vista para obtener empresas con sus cargos y exámenes asociados.
    Implementa cache de 8 horas para optimizar rendimiento.

    GET: Retorna lista de empresas con sus cargos y exámenes activos.
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]
    CACHE_KEY = 'cargo_empresa_examenes_data'
    CACHE_TIMEOUT = 60 * 60 * 8  # 8 horas

    def get(self, request):
        """Obtiene empresas con cargos y exámenes, usando cache cuando está disponible."""
        # Intentar obtener desde cache
        cached_data = self._get_from_cache()
        if cached_data:
            return Response(cached_data, status=status.HTTP_200_OK,
                            headers={'X-Cache': 'HIT'})

        # Generar datos frescos
        data = self._build_empresas_data()

        # Guardar en cache
        self._save_to_cache(data)

        return Response(data, status=status.HTTP_200_OK,
                        headers={'X-Cache': 'MISS'})

    def _get_from_cache(self):
        """Obtiene datos desde cache."""
        return cache.get(self.CACHE_KEY)

    def _save_to_cache(self, data):
        """Guarda datos en cache."""
        cache.set(self.CACHE_KEY, data, self.CACHE_TIMEOUT)

    def _build_empresas_data(self):
        """Construye estructura de datos con empresas, cargos, exámenes y estructura geográfica."""
        # Obtener empresas activas con exámenes activos
        empresas = Epresa.objects.filter(
            estadoempresa=1,
            examenes_por_cargo__examen__activo=True
        ).distinct().annotate(
            id_empresa=F('idempresa')
        ).values('id_empresa', 'nombre_empresa').order_by('nombre_empresa')

        # Serializar cada empresa con sus cargos y exámenes
        empresas_data = [
            EmpresaConCargosSerializer(empresa).data
            for empresa in empresas
        ]

        # Obtener lista de todos los exámenes activos
        examenes_list = list(
            Examen.objects.filter(activo=True)
            .order_by('nombre')
            .values('id_examen', 'nombre')
        )

        # Construir estructura geográfica: empresas → unidades → proyectos → centros
        estructura_geografica = self._build_estructura_geografica()

        return {
            'empresas': empresas_data,
            'examenes': examenes_list,
            'estructura': estructura_geografica
        }

    def _build_estructura_geografica(self):
        """
        Construye jerarquía geográfica de empresas activas DEDUPLICADA:
        Empresa → Unidades → Proyectos → Centros

        Deduplicación por nombre (case-insensitive):
        - Si hay varias unidades con el mismo nombre en una empresa,
          se conserva la de menor idunidad y se fusionan sus proyectos/centros.
        - Si hay varios proyectos con el mismo nombre en una unidad,
          se conserva el de menor idproyecto y se fusionan sus centros.
        - Los centros se deduplicanpor idcentrop.
        """
        from analitica.models import Unidadnegocio, Proyecto

        centros = Centroop.objects.filter(
            estadocentrop=1
        ).select_related(
            'id_proyecto__id_unidad__id_empresa'
        ).order_by('nombrecentrop')

        empresas_dict = {}

        for centro in centros:
            proyecto = centro.id_proyecto
            if not proyecto or not proyecto.estadoproyecto:
                continue
            unidad = proyecto.id_unidad
            if not unidad or not unidad.estadounidad:
                continue
            empresa = unidad.id_empresa
            if not empresa or not empresa.estadoempresa:
                continue

            # ── Empresa ──
            if empresa.idempresa not in empresas_dict:
                empresas_dict[empresa.idempresa] = {
                    'idempresa': empresa.idempresa,
                    'empresa': empresa.nombre_empresa.strip(),
                    'tipo': 'empresa',
                    'unidades': {},       # clave: nombre_normalizado
                    '_unidad_ids': {}     # nombre_normalizado → menor idunidad visto
                }
            emp_dict = empresas_dict[empresa.idempresa]

            # ── Unidad (deduplicar por nombre, conservar menor ID) ──
            nombre_unidad_key = unidad.nombreunidad.strip().upper()

            if nombre_unidad_key not in emp_dict['_unidad_ids']:
                # Primera vez que vemos este nombre de unidad
                emp_dict['_unidad_ids'][nombre_unidad_key] = unidad.idunidad
                emp_dict['unidades'][nombre_unidad_key] = {
                    'idunidad': unidad.idunidad,
                    'unidad': unidad.nombreunidad.strip(),
                    'tipo': 'unidad',
                    'proyectos': {},      # clave: nombre_normalizado
                    '_proy_ids': {}       # nombre_normalizado → menor idproyecto visto
                }
            elif unidad.idunidad < emp_dict['_unidad_ids'][nombre_unidad_key]:
                # Encontramos una con ID menor → actualizar ID representativo
                emp_dict['_unidad_ids'][nombre_unidad_key] = unidad.idunidad
                emp_dict['unidades'][nombre_unidad_key]['idunidad'] = unidad.idunidad
                emp_dict['unidades'][nombre_unidad_key]['unidad'] = unidad.nombreunidad.strip()

            uni_dict = emp_dict['unidades'][nombre_unidad_key]

            # ── Proyecto (deduplicar por nombre, conservar menor ID) ──
            nombre_proy_key = proyecto.nombreproyecto.strip().upper()

            if nombre_proy_key not in uni_dict['_proy_ids']:
                uni_dict['_proy_ids'][nombre_proy_key] = proyecto.idproyecto
                uni_dict['proyectos'][nombre_proy_key] = {
                    'idproyecto': proyecto.idproyecto,
                    'proyecto': proyecto.nombreproyecto.strip(),
                    'tipo': 'proyecto',
                    'centrosop': []
                }
            elif proyecto.idproyecto < uni_dict['_proy_ids'][nombre_proy_key]:
                uni_dict['_proy_ids'][nombre_proy_key] = proyecto.idproyecto
                uni_dict['proyectos'][nombre_proy_key]['idproyecto'] = proyecto.idproyecto
                uni_dict['proyectos'][nombre_proy_key]['proyecto'] = proyecto.nombreproyecto.strip()

            proy_dict = uni_dict['proyectos'][nombre_proy_key]

            # ── Centro (deduplicar por nombre, conservar menor ID) ──
            nombre_centro_key = centro.nombrecentrop.strip().upper()
            
            # Inicializar dict de centros si no existe
            if '_centro_ids' not in proy_dict:
                proy_dict['_centro_ids'] = {}  # nombre_normalizado → menor idcentrop visto
            
            if nombre_centro_key not in proy_dict['_centro_ids']:
                # Primera vez que vemos este nombre de centro
                proy_dict['_centro_ids'][nombre_centro_key] = centro.idcentrop
                proy_dict['centrosop'].append({
                    'idcentrop': centro.idcentrop,
                    'centro_op': centro.nombrecentrop.strip(),
                    'tipo': 'centro_op'
                })
            elif centro.idcentrop < proy_dict['_centro_ids'][nombre_centro_key]:
                # Encontramos uno con ID menor → reemplazar el anterior
                proy_dict['_centro_ids'][nombre_centro_key] = centro.idcentrop
                # Encontrar y actualizar el centro en la lista
                for c in proy_dict['centrosop']:
                    if c['centro_op'].upper() == nombre_centro_key:
                        c['idcentrop'] = centro.idcentrop
                        c['centro_op'] = centro.nombrecentrop.strip()
                        break

        # ── Convertir dicts a listas ordenadas y limpiar claves internas ──
        resultado = []
        for emp_dict in sorted(empresas_dict.values(), key=lambda x: x['empresa']):
            emp_dict.pop('_unidad_ids', None)

            unidades_list = sorted(
                emp_dict['unidades'].values(),
                key=lambda x: x['unidad']
            )
            for uni_dict in unidades_list:
                uni_dict.pop('_proy_ids', None)
                proyectos_list = sorted(
                    uni_dict['proyectos'].values(),
                    key=lambda x: x['proyecto']
                )
                # Limpiar claves internas de centros en cada proyecto
                for proy_dict in proyectos_list:
                    proy_dict.pop('_centro_ids', None)
                uni_dict['proyectos'] = proyectos_list
            emp_dict['unidades'] = unidades_list
            resultado.append(emp_dict)

        return resultado


# ──────────────────────────────────────────────────────────────
#  Funciones sincrónicas para reenvío de correos (sin Celery/Redis)
# ──────────────────────────────────────────────────────────────

def _generar_excel_reenvio(correo_obj):
    """
    Genera Excel con los trabajadores del correo para reenvío.
    Usa el mismo formato que EnviarCorreoView._generar_excel_simple().

    Args:
        correo_obj: instancia de CorreoExamenEnviado

    Returns:
        BytesIO: buffer del Excel generado, o None si no hay trabajadores
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    registros = correo_obj.trabajadores.select_related(
        'empresa', 'cargo', 'centro'
    ).prefetch_related('examenes__examen').all()

    if not registros.exists():
        return None

    # Recopilar exámenes únicos y datos de trabajadores
    examenes_unicos = set()
    trabajadores_data = []

    for reg in registros:
        examenes_del_trabajador = []
        for et in reg.examenes.all():
            examenes_del_trabajador.append(et.examen.nombre)
            examenes_unicos.add(et.examen.nombre)

        trabajadores_data.append({
            'uuid': reg.uuid_trabajador,
            'nombre': reg.nombre_trabajador,
            'documento': reg.documento_trabajador,
            'empresa': reg.empresa.nombre_empresa if reg.empresa else '',
            'unidad': getattr(reg.empresa, 'id_unidad', None),
            'proyecto': getattr(reg.centro, 'id_proyecto', None) if reg.centro else None,
            'centro': reg.centro.nombrecentrop if reg.centro else '',
            'ciudad': reg.ciudad or '',
            'cargo': reg.cargo.nombrecargo if reg.cargo else '',
            'tipo_examen': reg.tipo_examen,
            'examenes': examenes_del_trabajador,
        })

    # Crear Excel
    nombres_examenes = sorted(list(examenes_unicos))

    wb = Workbook()
    ws = wb.active
    ws.title = "Trabajadores Examenes"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Encabezados (mismo formato que EnviarCorreoView._generar_excel_simple)
    headers = [
        "UUID", "Empresa", "Unidad", "Proyecto", "Centro",
        "Ciudad", "Cargo", "Nombre", "Documento", "Tipo Examen"
    ] + nombres_examenes
    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border_style

    # Datos
    for trab in trabajadores_data:
        unidad_nombre = ''
        proyecto_nombre = ''

        if trab['proyecto']:
            proyecto_nombre = trab['proyecto'].nombreproyecto
            if hasattr(trab['proyecto'], 'id_unidad'):
                unidad_nombre = trab['proyecto'].id_unidad.nombreunidad

        row_data = [
            trab['uuid'], trab['empresa'], unidad_nombre, proyecto_nombre,
            trab['centro'], trab['ciudad'], trab['cargo'],
            trab['nombre'], trab['documento'], trab['tipo_examen']
        ]

        examenes_trabajador = set(trab['examenes'])
        for nombre_examen in nombres_examenes:
            row_data.append("X" if nombre_examen in examenes_trabajador else "")

        ws.append(row_data)

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20   # UUID
    ws.column_dimensions['B'].width = 25   # Empresa
    ws.column_dimensions['C'].width = 20   # Unidad
    ws.column_dimensions['D'].width = 25   # Proyecto
    ws.column_dimensions['E'].width = 20   # Centro
    ws.column_dimensions['F'].width = 15   # Ciudad
    ws.column_dimensions['G'].width = 25   # Cargo
    ws.column_dimensions['H'].width = 25   # Nombre
    ws.column_dimensions['I'].width = 15   # Documento
    ws.column_dimensions['J'].width = 15   # Tipo Examen

    for col_num in range(11, 11 + len(nombres_examenes)):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 5

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def reenviar_correo(correo_obj):
    """
    Función sincrónica para reenviar un correo de exámenes.
    Usa los mismos datos almacenados en BD (cuerpo, destinatarios, asunto).
    Si hay >1 trabajador registrado, regenera y adjunta el Excel.

    No usa Celery, Redis ni cache.

    Args:
        correo_obj: instancia de CorreoExamenEnviado

    Returns:
        dict: resultado con claves 'status', 'mensaje', 'correo_id', etc.
    """
    logger = logging.getLogger(__name__)
    logger.info(
        f"[REENVIO] Iniciando reenvío de correo id={correo_obj.id}, "
        f"UUID={correo_obj.uuid_correo}"
    )

    # 1. Obtener destinatarios desde BD y limpiar agresivamente
    destinatarios = [
        e.strip().replace('\r', '').replace('\n', '').replace('\t', '')
        for e in correo_obj.correos_destino.split(',')
        if e and e.strip()
    ]

    if not destinatarios:
        error_msg = "Sin destinatarios válidos en el registro"
        correo_obj.enviado_correctamente = False
        correo_obj.error_envio = error_msg
        correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])
        logger.error(f"[REENVIO] Correo id={correo_obj.id}: {error_msg}")
        return {'status': 'error', 'mensaje': error_msg}

    # 2. Verificar conexión SMTP
    from examenes.tasks import verificar_conexion_smtp
    smtp_ok, smtp_error = verificar_conexion_smtp()
    if not smtp_ok:
        error_msg = f"SMTP no disponible: {smtp_error}"
        correo_obj.enviado_correctamente = False
        correo_obj.error_envio = error_msg
        correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])
        logger.error(f"[REENVIO] Correo id={correo_obj.id}: {error_msg}")
        return {'status': 'error', 'mensaje': error_msg}

    try:
        # 3. Determinar si es masivo (>1 trabajador) y generar Excel
        cantidad_trabajadores = correo_obj.trabajadores.count()
        excel_buffer = None

        if cantidad_trabajadores > 1:
            logger.info(
                f"[REENVIO] Generando Excel para correo masivo id={correo_obj.id} "
                f"({cantidad_trabajadores} trabajadores)"
            )
            excel_buffer = _generar_excel_reenvio(correo_obj)

        # 4. Construir y enviar correo con los datos de BD
        email = EmailMultiAlternatives(
            subject=correo_obj.asunto,
            body=correo_obj.cuerpo_correo,
            from_email=settings.EMAIL_MEDICAL_FROM_EMAIL,
            to=destinatarios
        )

        # Si el cuerpo contiene HTML, adjuntar como alternativa
        if '<html' in correo_obj.cuerpo_correo.lower() or '<p>' in correo_obj.cuerpo_correo.lower():
            email.attach_alternative(correo_obj.cuerpo_correo, "text/html")

        # Adjuntar Excel si es masivo
        if excel_buffer:
            email.attach(
                'Trabajadores_Examenes.xlsx',
                excel_buffer.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        # Usar backend médico para enviar (360 CloudRegency)
        email.connection = EnviarCorreoView()._get_medical_email_backend()
        email.send(fail_silently=False)

        # 5. Marcar como enviado correctamente
        correo_obj.enviado_correctamente = True
        correo_obj.error_envio = None
        correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])

        logger.info(
            f"[REENVIO] Correo id={correo_obj.id} reenviado exitosamente "
            f"a {len(destinatarios)} destinatarios ({cantidad_trabajadores} trabajador(es))"
        )

        return {
            'status': 'ok',
            'correo_id': correo_obj.id,
            'uuid_correo': correo_obj.uuid_correo,
            'destinatarios': len(destinatarios),
            'trabajadores': cantidad_trabajadores,
            'con_excel': excel_buffer is not None
        }

    except Exception as e:
        error_msg = f"{type(e).__name__}: {str(e)}"
        correo_obj.enviado_correctamente = False
        correo_obj.error_envio = error_msg
        correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])
        logger.error(
            f"[REENVIO] Error reenviando correo id={correo_obj.id}: {error_msg}",
            exc_info=True
        )
        return {'status': 'error', 'mensaje': error_msg}


class ReintentarCorreoView(APIView):
    """
    Vista para reintentar manualmente el envío de un correo.

    POST: Reenvía el correo de forma sincrónica (sin Celery/Redis).
    Usa el mismo cuerpo, asunto y destinatarios almacenados en BD.
    Si tiene >1 trabajador asociado, regenera y adjunta el Excel.

    JSON esperado:
    {
        "correo_id": 123
    }
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def post(self, request):
        """Reenvía un correo directamente (sin Celery/Redis)."""
        logger = logging.getLogger(__name__)
        correo_id = request.data.get('correo_id')

        if not correo_id:
            return Response(
                {"error": "Se requiere el campo 'correo_id'"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            correo_obj = CorreoExamenEnviado.objects.get(id=correo_id)
        except CorreoExamenEnviado.DoesNotExist:
            return Response(
                {"error": f"Correo con id={correo_id} no encontrado"},
                status=status.HTTP_404_NOT_FOUND
            )

        logger.info(
            f"Reenvío manual solicitado para correo id={correo_id} "
            f"por usuario {request.user}"
        )

        # Llamar función sincrónica de reenvío
        resultado = reenviar_correo(correo_obj)

        if resultado['status'] == 'ok':
            return Response(
                {
                    "mensaje": "Correo reenviado exitosamente",
                    "uuid_correo": resultado['uuid_correo'],
                    "correo_id": resultado['correo_id'],
                    "trabajadores": resultado['trabajadores'],
                    "destinatarios": resultado['destinatarios'],
                    "con_excel": resultado['con_excel'],
                },
                status=status.HTTP_200_OK
            )
        else:
            return Response(
                {
                    "error": resultado['mensaje'],
                    "correo_id": correo_obj.id,
                    "uuid_correo": correo_obj.uuid_correo,
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def get(self, request):
        """Lista correos fallidos que pueden ser reintentados (últimas 48h)."""
        from datetime import timedelta
        
        hace_48h = timezone.now() - timedelta(hours=48)
        correos_fallidos = CorreoExamenEnviado.objects.filter(
            enviado_correctamente=False,
            fecha_envio__gte=hace_48h
        ).order_by('-fecha_envio').values(
            'id', 'uuid_correo', 'asunto', 'correos_destino',
            'error_envio', 'fecha_envio', 'tipo_examen'
        )[:50]

        return Response(
            {
                "total": len(correos_fallidos),
                "correos_fallidos": list(correos_fallidos)
            },
            status=status.HTTP_200_OK
        )


class DiagnosticoEmailView(APIView):
    """
    Vista de diagnóstico para verificar la configuración de email.
    Solo accesible por superadmins.
    
    GET: Verifica la conexión SMTP y retorna el estado.
    """
    permission_classes = [IsAuthenticated, IsSuperAdmin]

    def get(self, request):
        """Verifica la conexión SMTP y retorna diagnóstico."""
        logger = logging.getLogger(__name__)
        
        from examenes.tasks import verificar_conexion_smtp
        
        diagnostico = {
            "email_backend": settings.EMAIL_BACKEND,
            "email_host": settings.EMAIL_HOST,
            "email_port": settings.EMAIL_PORT,
            "email_use_ssl": settings.EMAIL_USE_SSL,
            "email_use_tls": settings.EMAIL_USE_TLS,
            "email_host_user": settings.EMAIL_HOST_USER[:3] + "***" if settings.EMAIL_HOST_USER else "(vacío)",
            "default_from_email": settings.DEFAULT_FROM_EMAIL,
            "email_timeout": getattr(settings, 'EMAIL_TIMEOUT', 'No configurado'),
        }

        # Verificar conexión SMTP
        smtp_ok, smtp_error = verificar_conexion_smtp()
        diagnostico["smtp_conexion"] = "OK" if smtp_ok else "FALLO"
        if smtp_error:
            diagnostico["smtp_error"] = smtp_error

        # Contar correos recientes
        from datetime import timedelta
        hace_24h = timezone.now() - timedelta(hours=24)
        
        correos_24h = CorreoExamenEnviado.objects.filter(fecha_envio__gte=hace_24h)
        diagnostico["correos_ultimas_24h"] = {
            "total": correos_24h.count(),
            "enviados": correos_24h.filter(enviado_correctamente=True).count(),
            "fallidos": correos_24h.filter(enviado_correctamente=False).count(),
        }

        logger.info(f"Diagnóstico de email ejecutado por {request.user}: SMTP={'OK' if smtp_ok else 'FALLO'}")

        return Response(diagnostico, status=status.HTTP_200_OK)


class ImprimirReporteCorreosView(APIView):
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def get(self, request):
        """
        Genera un reporte en Excel detallado con trabajadores y sus exámenes.

        Query params:
        - fecha_inicio: YYYY-MM-DD (opcional)
        - fecha_fin: YYYY-MM-DD (opcional)
        - empresas: IDs separados por coma o "all" (opcional, default: "all")
        """
        try:
            # Obtener parámetros de filtro
            fecha_inicio = request.GET.get('fecha_inicio')
            fecha_fin = request.GET.get('fecha_fin')
            empresas_param = request.GET.get('empresas', 'all')
            empresa_ids = []

            # Iniciar queryset base - ahora usamos RegistroExamenes
            # OPTIMIZADO: Prefetch de exámenes enviados para evitar N+1 en el loop
            from .models import HistorialEstadoRegistroExamen
            queryset = RegistroExamenes.objects.filter(
                estado_trabajador=1
            ).select_related(
                'correo_lote', 'empresa', 'cargo', 'centro',
                'centro__id_proyecto', 'centro__id_proyecto__id_unidad'
            ).prefetch_related(
                Prefetch(
                    'examenes',
                    queryset=ExamenTrabajador.objects.filter(
                        examen__activo=True
                    ).select_related('examen'),
                    to_attr='examenes_enviados_precargados'
                ),
                Prefetch(
                    'historial_estados',
                    queryset=HistorialEstadoRegistroExamen.objects.select_related('colaborador').order_by('-fecha_cambio'),
                    to_attr='historial_precargado'
                )
            )

            # Validar y aplicar filtro de fechas
            if fecha_inicio:
                try:
                    fecha_inicio_dt = datetime.strptime(
                        fecha_inicio, '%Y-%m-%d')
                    fecha_inicio_dt = timezone.make_aware(fecha_inicio_dt)
                    queryset = queryset.filter(
                        correo_lote__fecha_envio__gte=fecha_inicio_dt)
                except ValueError:
                    return Response(
                        {"error": "Formato de fecha_inicio inválido. Use YYYY-MM-DD"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            if fecha_fin:
                try:
                    fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                    fecha_fin_dt = fecha_fin_dt + timedelta(days=1)
                    fecha_fin_dt = timezone.make_aware(fecha_fin_dt)
                    queryset = queryset.filter(
                        correo_lote__fecha_envio__lt=fecha_fin_dt)
                except ValueError:
                    return Response(
                        {"error": "Formato de fecha_fin inválido. Use YYYY-MM-DD"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Validar y aplicar filtro de empresas
            if empresas_param != 'all':
                try:
                    empresa_ids = [int(id.strip())
                                   for id in empresas_param.split(',')]
                    empresas_existentes = Epresa.objects.filter(
                        idempresa__in=empresa_ids).values_list(
                        'idempresa', flat=True)
                    ids_no_encontrados = set(
                        empresa_ids) - set(empresas_existentes)
                    if ids_no_encontrados:
                        return Response(
                            {
                                "error": f"Las siguientes empresas no existen: {list(ids_no_encontrados)}"},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    queryset = queryset.filter(empresa_id__in=empresa_ids)
                except ValueError:
                    return Response(
                        {"error": "IDs de empresas inválidos. Use números separados por coma"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Obtener registros
            registros = queryset.order_by(
                'empresa__nombre_empresa',
                'cargo__nombrecargo',
                'nombre_trabajador')

            if not registros.exists():
                if fecha_inicio and fecha_fin:
                    rango = f"desde {fecha_inicio} hasta {fecha_fin}"
                elif fecha_inicio:
                    rango = f"desde {fecha_inicio}"
                elif fecha_fin:
                    rango = f"hasta {fecha_fin}"
                else:
                    rango = "sin rango de fechas"

                if empresas_param != 'all':
                    empresas_nombres = Epresa.objects.filter(
                        idempresa__in=empresa_ids).values_list(
                        'nombre_empresa', flat=True)
                    empresas_str = ', '.join(empresas_nombres)
                    msg = (
                        f"No se encontraron registros para las "
                        f"empresas solicitadas en el rango indicado. "
                        f"Empresas: {empresas_str}. Rango: {rango}."
                    )
                    return Response({"error": msg}, status=status.HTTP_404_NOT_FOUND)

            # Optimizar queryset para incluir colaborador que envió el correo
            registros = registros.select_related('correo_lote__enviado_por')

            # Obtener todos los exámenes activos
            examenes_activos = Examen.objects.filter(
                activo=True).order_by('nombre')
            nombres_examenes = [ex.nombre for ex in examenes_activos]
            
            # Deduplicar nombres de exámenes (por si hay duplicados en la BD)
            nombres_examenes = list(dict.fromkeys(nombres_examenes))  # Mantiene orden, elimina duplicados

            # Crear el workbook de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Detallado"

            # Estilos
            from openpyxl.styles import Border, Side
            header_fill = PatternFill(
                start_color="366092",
                end_color="366092",
                fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            center_alignment = Alignment(
                horizontal="center", vertical="center")
            border_style = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # Título del reporte
            num_cols = 15 + len(nombres_examenes)  # 15 columnas base (incluye "Enviado Por", "Completado Por" y "Fecha Completado") + exámenes
            ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
            ws['A1'] = "REPORTE DETALLADO DE EXÁMENES POR TRABAJADOR"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = center_alignment

            # Información de filtros
            row = 3
            if fecha_inicio or fecha_fin:
                ws.merge_cells(f'A{row}:{get_column_letter(num_cols)}{row}')
                periodo = "Período: "
                if fecha_inicio and fecha_fin:
                    periodo += f"desde {fecha_inicio} hasta {fecha_fin}"
                elif fecha_inicio:
                    periodo += f"desde {fecha_inicio}"
                else:
                    periodo += f"hasta {fecha_fin}"
                ws[f'A{row}'] = periodo
                ws[f'A{row}'].font = Font(italic=True)
                row += 1

            if empresas_param != 'all':
                ws.merge_cells(f'A{row}:{get_column_letter(num_cols)}{row}')
                empresas_nombres = Epresa.objects.filter(
                    idempresa__in=empresa_ids).values_list(
                    'nombre_empresa', flat=True)
                ws[f'A{row}'] = f"Empresas: {', '.join(empresas_nombres)}"
                ws[f'A{row}'].font = Font(italic=True)
                row += 1

            row += 1

            # Encabezados
            headers = [
                "UUID Trabajador",
                "UUID Correo",
                "Enviado Por",
                "Empresa",
                "Unidad",
                "Proyecto",
                "Centro",
                "Cargo",
                "Nombre",
                "Cédula",
                "Ciudad",
                "Tipo Examen",
                "Total Exámenes",
                "Completado Por",
                "Fecha Completado"] + nombres_examenes
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border_style

            row += 1
            data_start_row = row

            # Variables para acumular totales durante el procesamiento
            suma_total_examenes = 0
            totales_por_examen = {}  # Dict para guardar totales por columna de examen
            for nombre_ex in nombres_examenes:
                totales_por_examen[nombre_ex] = 0

            # Datos de trabajadores
            for registro in registros:
                # Obtener unidad y proyecto del centro
                centro = registro.centro
                unidad_nombre = (
                    centro.id_proyecto.id_unidad.nombreunidad
                    if centro and getattr(centro, 'id_proyecto', None) and getattr(centro.id_proyecto, 'id_unidad', None)
                    else ''
                )
                proyecto_nombre = (
                    centro.id_proyecto.nombreproyecto
                    if centro and getattr(centro, 'id_proyecto', None) else ''
                )

                # Obtener nombre completo de quien envió el correo
                enviado_por_nombre = ''
                if registro.correo_lote and registro.correo_lote.enviado_por:
                    colaborador = registro.correo_lote.enviado_por
                    nombre_col = getattr(colaborador, 'nombrecolaborador', '')
                    apellido_col = getattr(colaborador, 'apellidocolaborador', '')
                    enviado_por_nombre = f"{nombre_col} {apellido_col}".strip()

                # Colaborador y fecha en que se marcó el examen como Completado
                # (registros completados antes de existir esta auditoría quedan vacíos)
                completado_por_nombre = ''
                fecha_completado_str = ''
                historial = getattr(registro, 'historial_precargado', [])
                if historial:
                    ultimo_cambio = historial[0]
                    col_completo = ultimo_cambio.colaborador
                    nombre_completo = getattr(col_completo, 'nombrecolaborador', '')
                    apellido_completo = getattr(col_completo, 'apellidocolaborador', '')
                    completado_por_nombre = f"{nombre_completo} {apellido_completo}".strip()
                    fecha_completado_str = ultimo_cambio.fecha_cambio.astimezone(ZoneInfo('America/Bogota')).strftime('%d/%m/%Y %H:%M')

                # OPTIMIZADO: Usar exámenes precargados con Prefetch (sin query adicional)
                examenes_enviados = getattr(registro, 'examenes_enviados_precargados', [])

                # Crear set con nombres de exámenes ÚNICOS del trabajador (deduplicado)
                examenes_trabajador = set()
                for ex_env in examenes_enviados:
                    if ex_env.examen and ex_env.examen.nombre in nombres_examenes:
                        examenes_trabajador.add(ex_env.examen.nombre)

                total_examenes_trabajador = len(examenes_trabajador)
                
                # Acumular el total de exámenes
                suma_total_examenes += total_examenes_trabajador
                
                # Acumular totales por examen (una sola vez por examen único)
                for nombre_examen in examenes_trabajador:
                    totales_por_examen[nombre_examen] += 1

                row_data = [
                    registro.uuid_trabajador or '',
                    registro.correo_lote.uuid_correo if registro.correo_lote else '',
                    enviado_por_nombre,
                    registro.empresa.nombre_empresa if registro.empresa else '',
                    unidad_nombre,
                    proyecto_nombre,
                    centro.nombrecentrop if centro else '',
                    registro.cargo.nombrecargo if registro.cargo else '',
                    registro.nombre_trabajador,
                    registro.documento_trabajador,
                    registro.ciudad or '',
                    registro.tipo_examen or '',
                    total_examenes_trabajador,  # Total horizontal por trabajador
                    completado_por_nombre,
                    fecha_completado_str
                ]

                # Agregar X para cada examen activo enviado (una X por examen único)
                for nombre_examen in nombres_examenes:
                    if nombre_examen in examenes_trabajador:
                        row_data.append('X')
                    else:
                        row_data.append('')

                for col_num, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_num, value=value)
                    cell.border = border_style
                    # Centrar desde columna 12 en adelante (Ciudad + Tipo Examen + Total + Exámenes)
                    if col_num >= 12:
                        cell.alignment = center_alignment

                row += 1

            data_end_row = row - 1

            # Fila de totales verticales
            totales_fill = PatternFill(
                start_color="DCE6F1",
                end_color="DCE6F1",
                fill_type="solid")
            
            ws.cell(row=row, column=1, value="TOTALES")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            ws.cell(row=row, column=1).fill = totales_fill
            ws.cell(row=row, column=1).border = border_style

            # Aplicar estilo a celdas vacías de la fila de totales
            for col in range(2, 13):  # Columnas 2-12 (hasta Tipo Examen - ahora con Enviado Por)
                cell = ws.cell(row=row, column=col)
                cell.fill = totales_fill
                cell.border = border_style

            # Total de la columna "Total Exámenes" (suma que ya calculamos)
            col_total_examenes = 13
            cell_total_vertical = ws.cell(row=row, column=col_total_examenes, value=suma_total_examenes)
            cell_total_vertical.font = Font(bold=True, size=11)
            cell_total_vertical.fill = totales_fill
            cell_total_vertical.alignment = center_alignment
            cell_total_vertical.border = border_style

            # "Completado Por" y "Fecha Completado" (14-15) no tienen total numérico,
            # solo se les aplica el mismo estilo de fila para que no queden en blanco
            for col in range(14, 16):
                cell = ws.cell(row=row, column=col)
                cell.fill = totales_fill
                cell.border = border_style

            # Totales verticales por cada columna de examen
            col_inicio_examenes = 16  # Columna donde empiezan los exámenes (después de Completado Por / Fecha Completado)
            gran_total_examenes = 0  # Para el total general
            
            for col_num, nombre_examen in enumerate(nombres_examenes, start=col_inicio_examenes):
                # Usar el total que ya calculamos
                total_col = totales_por_examen[nombre_examen]
                gran_total_examenes += total_col

                cell = ws.cell(row=row, column=col_num, value=total_col)
                cell.font = Font(bold=True, size=11)
                cell.fill = totales_fill
                cell.alignment = center_alignment
                cell.border = border_style

            # Columna adicional para GRAN TOTAL
            col_gran_total = col_inicio_examenes + len(nombres_examenes)
            
            # Agregar encabezado "TOTAL" en la fila de headers
            header_row = data_start_row - 1
            cell_header_total = ws.cell(row=header_row, column=col_gran_total, value="TOTAL")
            cell_header_total.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell_header_total.font = Font(bold=True, color="FFFFFF", size=11)
            cell_header_total.alignment = center_alignment
            cell_header_total.border = border_style

            # Agregar totales horizontales por cada trabajador
            for r in range(data_start_row, data_end_row + 1):
                total_horizontal = 0
                for col_num in range(col_inicio_examenes, col_inicio_examenes + len(nombres_examenes)):
                    if ws.cell(row=r, column=col_num).value == 'X':
                        total_horizontal += 1
                
                cell = ws.cell(row=r, column=col_gran_total, value=total_horizontal)
                cell.alignment = center_alignment
                cell.border = border_style

            # Gran total en la esquina (suma de todos los exámenes)
            cell_gran_total = ws.cell(row=row, column=col_gran_total, value=gran_total_examenes)
            cell_gran_total.font = Font(bold=True, size=12)
            cell_gran_total.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            cell_gran_total.alignment = center_alignment
            cell_gran_total.border = border_style

            # Ajustar anchos de columnas
            ws.column_dimensions['A'].width = 40  # UUID Trabajador
            ws.column_dimensions['B'].width = 30  # UUID Correo
            ws.column_dimensions['C'].width = 25  # Empresa
            ws.column_dimensions['D'].width = 20  # Unidad
            ws.column_dimensions['E'].width = 25  # Proyecto
            ws.column_dimensions['F'].width = 20  # Centro
            ws.column_dimensions['G'].width = 25  # Cargo
            ws.column_dimensions['H'].width = 25  # Nombre
            ws.column_dimensions['I'].width = 15  # Cédula
            ws.column_dimensions['J'].width = 15  # Ciudad
            ws.column_dimensions['K'].width = 18  # Tipo Examen
            ws.column_dimensions['L'].width = 12  # Total Exámenes
            ws.column_dimensions['N'].width = 28  # Completado Por
            ws.column_dimensions['O'].width = 18  # Fecha Completado

            # Columnas de exámenes
            for col_num in range(col_inicio_examenes, col_inicio_examenes + len(nombres_examenes)):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = 6

            # Columna TOTAL
            ws.column_dimensions[get_column_letter(col_gran_total)].width = 8

            # Generar el archivo
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'reporte_trabajadores_examenes_{timestamp}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            wb.save(response)
            return response

        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error interno: {str(e)}", exc_info=True)
            return Response(
                {"error": f"Error interno: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class EnviarCorreoMasivoView(APIView):
    """
    Envía un correo masivo a múltiples trabajadores desde un CSV.
    Crea un registro CorreoExamenEnviado y N registros RegistroExamenes.
    
    Soporta dos formatos de CSV:
    
    FORMATO 1 (Original):
    Empresa,Unidad,Proyecto,Centro,Nombre,CC,Ciudad,Cargo,TipoExamen,Examenes
    
    FORMATO 2 (Nuevo - Exámenes como columnas con 1/0):
    Nombre de empresa;unidad de negocio;PROYECTO;Desc. C.O.;Cedula;Nombre Empleado;
    Cargo;Fecha de Ingreso;TIPO DE EXAMEN;OPTOMETRIA;AUDIOMETRIA;...
    (Los exámenes marcados con "1" se asignan al trabajador)
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    # Mapeo de columnas del CSV formato 2 a nombres internos
    COLUMN_MAPPING = {
        'nombre de empresa': 'empresa',
        'unidad de negocio': 'unidad',
        'proyecto': 'proyecto',
        'desc. c.o.': 'centro',
        'cedula': 'cc',
        'nombre empleado': 'nombre',
        'cargo': 'cargo',
        'ciudad': 'ciudad',
        'fecha de ingreso': 'fecha_ingreso',
        'tipo de examen': 'tipoexamen',
    }

    def _get_medical_email_backend(self):
        """
        Retorna una conexión SMTP configurada con credenciales de Exámenes Médicos.
        
        Returns:
            SMTPConnection: Conexión SMTP con configuración médica (360 CloudRegency)
        """
        from django.core.mail.backends.smtp import EmailBackend
        
        return EmailBackend(
            host=settings.EMAIL_MEDICAL_HOST,
            port=settings.EMAIL_MEDICAL_PORT,
            username=settings.EMAIL_MEDICAL_HOST_USER,
            password=settings.EMAIL_MEDICAL_HOST_PASSWORD,
            use_ssl=settings.EMAIL_MEDICAL_USE_SSL,
            use_tls=settings.EMAIL_MEDICAL_USE_TLS,
            timeout=settings.EMAIL_MEDICAL_TIMEOUT,
            fail_silently=False
        )

    def _detect_csv_format(self, fieldnames):
        """
        Detecta el formato del CSV basándose en las columnas.
        
        Returns:
            tuple: (formato: str, columnas_examenes: list)
            - formato: 'original' o 'columnas_examenes'
            - columnas_examenes: lista de nombres de columnas que son exámenes (solo para formato 2)
        """
        fieldnames_lower = [f.lower().strip() for f in fieldnames]
        
        logger = logging.getLogger(__name__)
        logger.info(f"Detectando formato con fieldnames_lower: {fieldnames_lower}")
        
        # Verificar si tiene la columna 'examenes' (formato original)
        if 'examenes' in fieldnames_lower:
            logger.info("Formato detectado: ORIGINAL (columna 'examenes')")
            return 'original', []
        
        # Verificar si tiene 'tipo de examen' o 'tipo examen' (formato nuevo)
        tipo_examen_idx = None
        for idx, col in enumerate(fieldnames_lower):
            # Buscar variantes de "tipo de examen"
            if col == 'tipo de examen' or col == 'tipo examen' or col == 'tipoexamen':
                tipo_examen_idx = idx
                logger.info(f"Columna 'tipo de examen' encontrada en índice {idx}")
                break
        
        if tipo_examen_idx is not None:
            # Las columnas después de 'tipo de examen' son los exámenes
            columnas_examenes = fieldnames[tipo_examen_idx + 1:]
            # Filtrar columnas vacías
            columnas_examenes = [c for c in columnas_examenes if c and c.strip()]
            logger.info(f"Formato detectado: COLUMNAS_EXAMENES con {len(columnas_examenes)} columnas")
            logger.info(f"Columnas de exámenes: {columnas_examenes}")
            return 'columnas_examenes', columnas_examenes
        
        # Si tiene 'tipoexamen' es el formato original
        if 'tipoexamen' in fieldnames_lower:
            logger.info("Formato detectado: ORIGINAL (columna 'tipoexamen')")
            return 'original', []
        
        logger.info(f"Formato NO RECONOCIDO. Fieldnames: {fieldnames}")
        return 'unknown', []

    def _normalize_row_format2(self, row):
        """
        Normaliza una fila del formato 2 (columnas de exámenes) al formato interno.
        """
        normalized = {}
        for key, value in row.items():
            key_lower = key.lower().strip()
            # Mapear columnas conocidas
            if key_lower in self.COLUMN_MAPPING:
                normalized[self.COLUMN_MAPPING[key_lower]] = value.strip() if isinstance(value, str) else value
            else:
                # Mantener otras columnas (posiblemente exámenes)
                normalized[key] = value
        return normalized

    def _get_examenes_from_columns(self, row, columnas_examenes, examenes_map):
        """
        Obtiene los exámenes marcados con '1' o 'X' en las columnas del CSV.
        Los valores vacíos o sin marcar se ignoran sin generar errores.
        
        Args:
            row: Fila del CSV (dict)
            columnas_examenes: Lista de nombres de columnas que son exámenes
            examenes_map: Diccionario nombre_lower -> objeto Examen
            
        Returns:
            tuple: (examenes_nombres: list, examenes_bd: list, errores: list)
        """
        examenes_nombres = []
        examenes_bd = []
        errores = []
        
        for col_name in columnas_examenes:
            if not col_name or not col_name.strip():
                continue
                
            # Obtener el valor de la columna
            valor = row.get(col_name, '').strip() if isinstance(row.get(col_name), str) else str(row.get(col_name, '')).strip()
            
            # Si el valor es '1' o 'X' (en cualquier caso), el examen está marcado
            if valor.upper() in ['1', 'X']:
                # Buscar el examen en la BD (nombre exacto o similar)
                col_name_clean = col_name.strip()
                examen = examenes_map.get(col_name_clean.lower())
                
                if examen:
                    examenes_nombres.append(examen.nombre)
                    examenes_bd.append(examen)
                else:
                    # Intentar buscar con coincidencia parcial
                    encontrado = False
                    for nombre_bd, examen_obj in examenes_map.items():
                        # Comparar sin acentos y mayúsculas
                        if self._normalize_text(col_name_clean) == self._normalize_text(nombre_bd):
                            examenes_nombres.append(examen_obj.nombre)
                            examenes_bd.append(examen_obj)
                            encontrado = True
                            break
                    
                    if not encontrado:
                        errores.append(f"Examen '{col_name_clean}' no encontrado en la BD")
        
        return examenes_nombres, examenes_bd, errores

    def _normalize_text(self, text):
        """Normaliza texto removiendo acentos y convirtiendo a minúsculas."""
        import unicodedata
        if not text:
            return ''
        # Remover acentos
        nfkd = unicodedata.normalize('NFKD', text)
        text_sin_acentos = ''.join([c for c in nfkd if not unicodedata.combining(c)])
        return text_sin_acentos.lower().strip()

    def _clear_cache(self):
        """Limpia el cache de reportes de correos y datos de empresas con exámenes."""
        logger = logging.getLogger(__name__)
        try:
            # Limpiar todos los caches de reportes de correos paginados
            cache.delete_pattern("reporte_correos_page=*")
            
            # Limpiar todos los caches de detalles de correo
            cache.delete_pattern("detalle_correo=*")
            
            # Limpiar cache de datos de empresas con exámenes
            cache.delete('cargo_empresa_examenes_data')
            
            logger.info("Cache limpiado: reportes de correos y datos de empresas (EnviarCorreoMasivoView)")
        except Exception as e:
            logger.warning(f"Error al limpiar cache: {str(e)}")

    def post(self, request):
        """
        POST: Procesa CSV y envía correo masivo

        Request:
        {
            "archivo_csv": <file>,
            "asunto": "Convocatoria a exámenes",
            "cuerpo_correo": "<p>Estimado participante...</p>"
        }

        CSV esperado:
        nombre,documento,empresa_id,cargo_id
        Juan Pérez,12345678,6,1
        María García,87654321,6,2
        """
        logger = logging.getLogger(__name__)

        try:
            logger.info("=== INICIO EnviarCorreoMasivoView ===")
            logger.info(f"Request data keys: {request.data.keys()}")
            logger.info(f"Request FILES keys: {request.FILES.keys()}")

            serializer = EnviarCorreoMasivoSerializer(data=request.data)
            if not serializer.is_valid():
                logger.error(f"Serializer validation failed: {serializer.errors}")
                return Response(serializer.errors,
                                status=status.HTTP_400_BAD_REQUEST)

            archivo_csv = serializer.validated_data['archivo_csv']
            # asunto y cuerpo_correo ya no se leen del request, se generan
            # automáticamente
            solicitante_extra = None
            solicitante_extra_id = serializer.validated_data.get('solicitante_extra_id')
            if solicitante_extra_id:
                from usuarios.models import Colaboradores as ColabModel
                try:
                    colab_extra = ColabModel.objects.get(idcolaborador=solicitante_extra_id)
                    correo_extra = getattr(colab_extra, 'correocolaborador', None) or \
                                   getattr(colab_extra, 'correo', None) or \
                                   getattr(colab_extra, 'email', None)
                    solicitante_extra = correo_extra.strip() if correo_extra else None
                except ColabModel.DoesNotExist:
                    logger.warning(f"Colaborador extra con id {solicitante_extra_id} no encontrado")

            # Leer CSV con múltiples intentos de codificación y delimitadores
            archivo_csv.seek(0)

            # Intentar diferentes codificaciones
            contenido_csv = None
            for encoding in ['utf-8-sig', 'utf-8',
                             'latin-1', 'iso-8859-1', 'cp1252']:
                try:
                    archivo_csv.seek(0)
                    contenido_csv = archivo_csv.read().decode(encoding)
                    logger.info(
                        f"CSV decodificado exitosamente con "
                        f"encoding: {encoding}"
                    )
                    break
                except (UnicodeDecodeError, AttributeError):
                    continue

            if contenido_csv is None:
                logger.error(
                    "No se pudo decodificar el CSV con ninguna codificación")
                return Response(
                    {
                        "error": (
                            "No se pudo leer el archivo CSV. "
                            "Verifique la codificación del archivo."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Detectar delimitador (coma o punto y coma)
            # Estrategia: contar ocurrencias de delimitadores en la primera línea
            delimiter = ','
            try:
                # Obtener la primera línea del CSV
                primera_linea = contenido_csv.split('\n')[0]
                
                # Contar ocurrencias de posibles delimitadores
                contar_comas = primera_linea.count(',')
                contar_puntoycoma = primera_linea.count(';')
                
                logger.info(f"Primera línea tiene {contar_comas} comas y {contar_puntoycoma} puntos y comas")
                
                # Si hay más puntos y comas que comas, usar punto y coma
                if contar_puntoycoma > contar_comas and contar_puntoycoma > 0:
                    delimiter = ';'
                    logger.info("Delimitador detectado por conteo: ';'")
                else:
                    # Intentar con el sniffer como alternativa
                    sniffer = csv.Sniffer()
                    sample = contenido_csv[:1024]
                    try:
                        dialect = sniffer.sniff(sample, delimiters=',;')
                        delimiter = dialect.delimiter
                        logger.info(f"Delimitador detectado por sniffer: '{delimiter}'")
                    except Exception:
                        logger.info("Sniffer no pudo detectar, usando coma por defecto")
                        delimiter = ','
            except Exception as e:
                logger.warning(f"Error en detección de delimitador: {str(e)}. Usando coma por defecto")
                delimiter = ','
            
            logger.info(f"Delimitador final: '{delimiter}'")

            # Leer CSV con el delimitador detectado
            stream = io.StringIO(contenido_csv)
            reader = csv.DictReader(stream, delimiter=delimiter)

            # Obtener y validar headers ANTES de leer los datos
            if not reader.fieldnames:
                logger.error("CSV sin encabezados")
                return Response(
                    {"error": "El archivo CSV no tiene encabezados"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Mantener nombres originales para detectar formato
            fieldnames_original = [f.strip() for f in reader.fieldnames]
            # Normalizar nombres de columnas a minúsculas para comparación
            fieldnames = [f.lower() for f in fieldnames_original]
            logger.info(f"CSV fieldnames detectados: {fieldnames}")

            # ===================================================================
            # DETECCIÓN AUTOMÁTICA DEL FORMATO DEL CSV
            # ===================================================================
            formato_csv, columnas_examenes = self._detect_csv_format(fieldnames_original)
            logger.info(f"Formato CSV detectado: {formato_csv}")
            logger.info(f"Delimitador usado: '{delimiter}'")
            logger.info(f"Fieldnames originales: {fieldnames_original}")
            logger.info(f"Fieldnames normalizados: {fieldnames}")
            
            if formato_csv == 'columnas_examenes':
                logger.info(f"Columnas de exámenes detectadas: {columnas_examenes}")
                # Formato 2: Columnas de exámenes con 1/0
                # Validar columnas requeridas para formato 2
                expected_format2 = {
                    'nombre de empresa', 'unidad de negocio', 'proyecto', 
                    'desc. c.o.', 'cedula', 'nombre empleado', 'cargo', 
                    'tipo de examen'
                }
                missing = expected_format2 - set(fieldnames)
                if missing:
                    # Intentar con variantes comunes
                    # En caso de que las columnas estén escritas de otra forma
                    logger.warning(f"Columnas esperadas faltantes: {missing}")
                    # Mostrar sugerencia pero no fallar si la mayoría están presentes
                    if len(missing) > 2:  # Si faltan más de 2 columnas críticas
                        logger.error(f"Columnas críticas faltantes para formato 2: {missing}")
                        return Response(
                            {
                                "error": f"El CSV (formato exámenes como columnas) debe contener las columnas: "
                                         f"{', '.join(sorted(expected_format2))}. "
                                         f"Columnas recibidas: {', '.join(fieldnames)}. "
                                         f"Columnas faltantes: {', '.join(missing)}"
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )
            elif formato_csv == 'original':
                logger.info("Formato detectado: ORIGINAL (columna Examenes)")
                # Formato 1: Columna 'Examenes' separados por coma
                expected = {
                    'empresa', 'unidad', 'proyecto', 'centro', 'nombre',
                    'cc', 'ciudad', 'cargo', 'tipoexamen', 'examenes'
                }
                missing = expected - set(fieldnames)
                if missing:
                    logger.error(f"Columnas faltantes. Expected: {expected}, Got: {set(fieldnames)}")
                    return Response(
                        {
                            "error": f"El CSV debe contener las columnas: "
                                     f"{', '.join(sorted(expected))} (insensible a mayúsculas). "
                                     f"Columnas recibidas: {', '.join(fieldnames)}"
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                logger.error(f"Formato de CSV no reconocido. Fieldnames: {fieldnames}")
                return Response(
                    {
                        "error": "Formato de CSV no reconocido. El CSV debe tener: "
                                 "1) Columna 'Examenes' con nombres separados por coma, O "
                                 "2) Columna 'TIPO DE EXAMEN' seguida de columnas de exámenes con valores 1/0. "
                                 f"Columnas recibidas: {', '.join(fieldnames[:20])}..." 
                                 if len(fieldnames) > 20 
                                 else f"Columnas recibidas: {', '.join(fieldnames)}"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            trabajadores_data = list(reader)
            logger.info(f"Total de filas en CSV: {len(trabajadores_data)}")

            if not trabajadores_data:
                return Response(
                    {"error": "El archivo CSV está vacío"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # ===================================================================
            # FASE 1: VALIDACIÓN COMPLETA DEL CSV (NO SE ENVÍA NADA AÚN)
            # ===================================================================
            # OPTIMIZACIÓN: Precargar todos los catálogos antes del loop
            # Esto reduce de N queries por fila a solo 6 queries totales
            
            # Precargar empresas (nombre_lower -> objeto)
            empresas_map = {
                e.nombre_empresa.lower().strip(): e 
                for e in Epresa.objects.all()
            }
            
            # Precargar unidades con su empresa (nombre_lower, empresa_id) -> lista de objetos
            # Puede haber varias unidades con el mismo nombre bajo la misma empresa
            # Se ordenan por ID ascendente para que el de menor ID sea el primero
            from collections import defaultdict
            unidades_map = defaultdict(list)
            for u in Unidadnegocio.objects.select_related('id_empresa').order_by('idunidad').all():
                unidades_map[(u.nombreunidad.lower().strip(), u.id_empresa_id)].append(u)
            
            # Precargar proyectos con su unidad (nombre_lower, unidad_id) -> objeto
            proyectos_map = {
                (p.nombreproyecto.lower().strip(), p.id_unidad_id): p
                for p in Proyecto.objects.select_related('id_unidad').all()
            }
            
            # Precargar centros con su proyecto (nombre_lower, proyecto_id) -> objeto
            centros_map = {
                (c.nombrecentrop.lower().strip(), c.id_proyecto_id): c
                for c in Centroop.objects.select_related('id_proyecto').all()
            }
            
            # Precargar cargos (nombre_lower -> objeto)
            cargos_map = {
                c.nombrecargo.lower().strip(): c
                for c in Cargo.objects.all()
            }
            
            # Precargar exámenes activos (nombre_lower -> objeto)
            examenes_map = {
                e.nombre.lower().strip(): e
                for e in Examen.objects.filter(activo=True)
            }
            
            total_unidades = sum(len(v) for v in unidades_map.values())
            logger.info(f"Catálogos precargados: {len(empresas_map)} empresas, "
                       f"{total_unidades} unidades ({len(unidades_map)} nombres únicos), {len(proyectos_map)} proyectos, "
                       f"{len(centros_map)} centros, {len(cargos_map)} cargos, "
                       f"{len(examenes_map)} exámenes")
            
            # Validar y procesar cada trabajador (resolviendo jerarquía:
            # empresa -> unidad -> proyecto -> centro)
            trabajadores_validos = []
            errores_validacion = []
            
            # Lista de exámenes disponibles en BD para mostrar en mensajes de error
            examenes_disponibles = list(examenes_map.keys())

            for idx, trab in enumerate(trabajadores_data, start=2):
                try:
                    # =========================================================
                    # PROCESAR SEGÚN FORMATO DETECTADO
                    # =========================================================
                    if formato_csv == 'columnas_examenes':
                        # FORMATO 2: Columnas de exámenes con valores 1/0
                        # Normalizar fila usando el mapeo de columnas
                        row = self._normalize_row_format2(trab)
                        
                        empresa_name = (row.get('empresa') or '').strip()
                        unidad_name = (row.get('unidad') or '').strip()
                        proyecto_name = (row.get('proyecto') or '').strip()
                        centro_name = (row.get('centro') or '').strip()
                        nombre = (row.get('nombre') or '').strip()
                        documento = (row.get('cc') or '').strip()
                        cargo_name = (row.get('cargo') or '').strip()
                        ciudad = (row.get('ciudad') or '').strip()
                        tipo_examen = (row.get('tipoexamen') or '').upper().strip()
                        
                        # Obtener exámenes de las columnas marcadas con '1'
                        examenes_nombres, examenes_bd, errores_examenes = self._get_examenes_from_columns(
                            trab, columnas_examenes, examenes_map
                        )
                        
                        # Agregar errores de exámenes no encontrados
                        for error in errores_examenes:
                            errores_validacion.append(f"Línea {idx}: {error}")
                        
                    else:
                        # FORMATO 1 (Original): Columna 'Examenes' separados por coma
                        row = {
                            k.strip().lower(): (
                                v.strip() if isinstance(v, str) else v
                            ) for k, v in trab.items()
                        }

                        empresa_name = (row.get('empresa') or '').strip()
                        unidad_name = (row.get('unidad') or '').strip()
                        proyecto_name = (row.get('proyecto') or '').strip()
                        nombre = (row.get('nombre') or '').strip()
                        documento = (row.get('cc') or '').strip()
                        cargo_name = (row.get('cargo') or '').strip()
                        centro_name = (row.get('centro') or '').strip()
                        ciudad = (row.get('ciudad') or '').strip()
                        tipo_examen = (row.get('tipoexamen') or '').upper().strip()
                        examenes_str = (row.get('examenes') or '').strip()
                        
                        # Parsear exámenes para formato original
                        examenes_nombres = []
                        examenes_bd = []
                        
                        if examenes_str:
                            examenes_nombres = [
                                e.strip() for e in examenes_str.split(',') if e.strip()
                            ]
                    
                    # =========================================================
                    # VALIDACIONES COMUNES PARA AMBOS FORMATOS
                    # =========================================================
                    
                    # Saltar filas vacías (sin nombre ni documento)
                    if not nombre and not documento:
                        continue
                    
                    if not nombre or not documento:
                        errores_validacion.append(
                            f"Línea {idx}: Nombre y/o CC vacío (nombre='{nombre}', cc='{documento}')")
                        continue

                    # Validar tipo de examen
                    tipos_validos = ['INGRESO', 'PERIODICO', 'RETIRO', 'ESPECIAL', 'POST_INCAPACIDAD', 'ALTURAS']
                    if tipo_examen not in tipos_validos:
                        errores_validacion.append(
                            f"Línea {idx}: TipoExamen debe ser uno de "
                            f"{', '.join(tipos_validos)}, "
                            f"recibido: '{tipo_examen}'"
                        )
                        continue

                    # Validar que hay exámenes especificados
                    if not examenes_nombres:
                        if formato_csv == 'columnas_examenes':
                            errores_validacion.append(
                                f"Línea {idx}: No hay exámenes marcados con '1' en las columnas de exámenes")
                        else:
                            errores_validacion.append(
                                f"Línea {idx}: Campo 'Examenes' vacío")
                        continue

                    # OPTIMIZADO: Buscar empresa usando mapa precargado (O(1) en lugar de query)
                    empresa = empresas_map.get(empresa_name.lower())
                    if not empresa:
                        # Intentar búsqueda normalizada (sin acentos)
                        empresa = None
                        for nombre_bd, emp_obj in empresas_map.items():
                            if self._normalize_text(empresa_name) == self._normalize_text(nombre_bd):
                                empresa = emp_obj
                                break
                        if not empresa:
                            errores_validacion.append(
                                f"Línea {idx}: Empresa '{empresa_name}' no encontrada en la BD")
                            continue

                    # OPTIMIZADO: Buscar unidades candidatas usando mapa precargado
                    # Puede haber varias unidades con el mismo nombre bajo la misma empresa
                    # (diferentes descripciones pero mismo nombre). Se prueban todas.
                    unidades_candidatas = unidades_map.get((unidad_name.lower(), empresa.idempresa), [])
                    if not unidades_candidatas:
                        # Intentar búsqueda normalizada (sin acentos)
                        for (nombre_bd, emp_id), uni_list in unidades_map.items():
                            if emp_id == empresa.idempresa and self._normalize_text(unidad_name) == self._normalize_text(nombre_bd):
                                unidades_candidatas = uni_list
                                break
                        if not unidades_candidatas:
                            errores_validacion.append(
                                f"Línea {idx}: Unidad '{unidad_name}' no encontrada para empresa '{empresa.nombre_empresa}'")
                            continue

                    # Buscar proyecto probando TODAS las unidades candidatas
                    # Sin importar cuál unidad duplicada contenga el proyecto,
                    # lo que importa es que empresa + proyecto + centro coincidan.
                    proyecto = None
                    unidad = None
                    for unidad_candidata in unidades_candidatas:
                        # Búsqueda exacta
                        proyecto = proyectos_map.get((proyecto_name.lower(), unidad_candidata.idunidad))
                        if proyecto:
                            unidad = unidad_candidata
                            break
                        # Búsqueda normalizada (sin acentos)
                        for (nombre_bd, uni_id), proy_obj in proyectos_map.items():
                            if uni_id == unidad_candidata.idunidad and self._normalize_text(proyecto_name) == self._normalize_text(nombre_bd):
                                proyecto = proy_obj
                                unidad = unidad_candidata
                                break
                        if proyecto:
                            break

                    if not unidad:
                        # Usar la primera unidad (menor ID) como fallback para el registro
                        unidad = unidades_candidatas[0]

                    if not proyecto:
                        errores_validacion.append(
                            f"Línea {idx}: Proyecto '{proyecto_name}' no encontrado para empresa '{empresa.nombre_empresa}' bajo unidad '{unidad_name}'")
                        continue

                    # OPTIMIZADO: Buscar centro usando mapa precargado
                    centro = centros_map.get((centro_name.lower(), proyecto.idproyecto))
                    if not centro:
                        # Intentar búsqueda normalizada
                        for (nombre_bd, proy_id), cent_obj in centros_map.items():
                            if proy_id == proyecto.idproyecto and self._normalize_text(centro_name) == self._normalize_text(nombre_bd):
                                centro = cent_obj
                                break
                        if not centro:
                            errores_validacion.append(
                                f"Línea {idx}: Centro '{centro_name}' no encontrado para proyecto '{proyecto.nombreproyecto}'"
                            )
                            continue

                    # OPTIMIZADO: Buscar cargo usando mapa precargado
                    cargo = cargos_map.get(cargo_name.lower())
                    if not cargo:
                        # Intentar búsqueda normalizada
                        for nombre_bd, cargo_obj in cargos_map.items():
                            if self._normalize_text(cargo_name) == self._normalize_text(nombre_bd):
                                cargo = cargo_obj
                                break
                        if not cargo:
                            errores_validacion.append(
                                f"Línea {idx}: Cargo '{cargo_name}' no encontrado en la BD")
                            continue

                    # Validar exámenes (solo para formato original, ya que formato 2 lo hizo antes)
                    if formato_csv == 'original':
                        examenes_bd = []
                        examen_invalido = False
                        for examen_nombre in examenes_nombres:
                            examen = examenes_map.get(examen_nombre.lower())
                            if not examen:
                                # Intentar búsqueda normalizada
                                for nombre_bd, ex_obj in examenes_map.items():
                                    if self._normalize_text(examen_nombre) == self._normalize_text(nombre_bd):
                                        examen = ex_obj
                                        break
                                if not examen:
                                    errores_validacion.append(
                                        f"Línea {idx}: Examen '{examen_nombre}' "
                                        f"no encontrado o no está activo. "
                                        f"Exámenes disponibles: {', '.join(sorted(examenes_disponibles)[:10])}..."
                                    )
                                    examen_invalido = True
                                    break
                            examenes_bd.append(examen)
                        
                        if examen_invalido:
                            continue
                    
                    # Validar que examenes_bd tiene objetos válidos
                    if not examenes_bd:
                        errores_validacion.append(
                            f"Línea {idx}: No se encontraron exámenes válidos")
                        continue
                    
                    # Todos los datos son válidos
                    trabajadores_validos.append({
                        'nombre': nombre,
                        'documento': documento,
                        'empresa': empresa,
                        'unidad': unidad,
                        'proyecto': proyecto,
                        'centro': centro,
                        'ciudad': ciudad,
                        'cargo': cargo,
                        'tipo_examen': tipo_examen,
                        'examenes_nombres': [e.nombre for e in examenes_bd],  # Usar nombres de BD
                        'examenes_bd': examenes_bd,
                    })

                except Exception as e:
                    import traceback
                    errores_validacion.append(f"Línea {idx}: Error inesperado - {str(e)}")
                    logger.error(f"Error procesando línea {idx}: {traceback.format_exc()}")

            # ===================================================================
            # VALIDACIÓN COMPLETADA - VERIFICAR ERRORES ANTES DE CONTINUAR
            # ===================================================================
            # Si hay errores de validación, retornarlos SIN crear registros ni
            # enviar correos
            if errores_validacion:
                return Response(
                    {
                        "mensaje": "Errores encontrados en el CSV. No se envió ningún correo.",
                        "errores": errores_validacion,
                        "trabajadores_validos": len(trabajadores_validos),
                        "total_trabajadores": len(trabajadores_data)
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Verificar que al menos haya un trabajador válido
            if not trabajadores_validos:
                return Response(
                    {
                        "mensaje": "No se envió ningún correo.",
                        "error": "Ningún trabajador cumple los requisitos de validación"
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # ===================================================================
            # FASE 2: TODOS LOS TRABAJADORES SON VÁLIDOS - PROCEDER AL ENVÍO
            # ===================================================================
            # Agrupar trabajadores por tipo de examen para saber si será
            # INGRESO, PERIODICO o MIXTO
            tipos_examen_unicos = set(t['tipo_examen']
                                      for t in trabajadores_validos)
            tipo_examen_principal = list(tipos_examen_unicos)[0] if len(
                tipos_examen_unicos) == 1 else 'MIXTO'

            # Generar UUID para el lote
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            unique_id = str(uuid.uuid4())[:8]
            uuid_correo = f"{unique_id}-{timestamp}"

            # Construir el cuerpo del correo con el UUID del lote
            # y datos del colaborador que realiza el envío
            # Obtener el colaborador del usuario autenticado (puede ser objeto o id)
            colaborador = (
                request.user.idcolaboradoru if hasattr(
                    request.user, 'idcolaboradoru') else None
            )

            # Intentar resolver al objeto Colaboradores y extraer nombre/correo
            nombre_colaborador = None
            correo_colaborador = None
            try:
                from usuarios.models import Colaboradores
                if isinstance(colaborador, Colaboradores):
                    colaborador_obj = colaborador
                elif colaborador is not None:
                    # Si viene como id (o FK value), intentar obtener objeto
                    colaborador_obj = Colaboradores.objects.filter(pk=getattr(colaborador, 'idcolaborador', colaborador)).first()
                else:
                    colaborador_obj = None

                if colaborador_obj:
                    nombre_colaborador = getattr(colaborador_obj, 'nombrecolaborador', None)
                    correo_colaborador = getattr(colaborador_obj, 'correocolaborador', None)
            except Exception:
                # Fallback: usar datos del user
                colaborador_obj = None

            if not nombre_colaborador:
                nombre_colaborador = (getattr(request.user, 'get_full_name', lambda: None)() or getattr(request.user, 'username', None))
            if not correo_colaborador:
                correo_colaborador = getattr(request.user, 'email', None)

            # Construir la línea de correos del solicitante
            correos_solicitantes_text = correo_colaborador or 'No disponible'
            if solicitante_extra:
                correos_solicitantes_text = f"{correo_colaborador}, {solicitante_extra}" if correo_colaborador else f"No disponible, {solicitante_extra}"

            # DETECTAR SI ES INDIVIDUAL O MASIVO
            es_masivo = len(trabajadores_validos) > 1
            
            if es_masivo:
                # Cuerpo HTML para correo masivo (múltiples trabajadores)
                cuerpo_final = f"""
<html>
<body>
    <p>Cordial Saludo.</p>
    <p>Se han programado los siguientes exámenes médicos
para los trabajadores en el excel adjunto.</p>
    <br>
    <hr>
    <p><strong>ID de Seguimiento:</strong> {uuid_correo}</p>
    <p><strong>Solicitante:</strong> {nombre_colaborador if nombre_colaborador else 'No disponible'}</p>
    <p><strong>Correo del solicitante:</strong> {correos_solicitantes_text}</p>
</body>
</html>
"""
            else:
                # Cuerpo para correo individual (1 trabajador) - formato plain text
                trab = trabajadores_validos[0]
                tipos_legibles = {
                    'INGRESO': 'Examen de Ingreso',
                    'PERIODICO': 'Examen Periódico',
                    'RETIRO': 'Examen de Retiro',
                    'ESPECIAL': 'Examen Especial',
                    'POST_INCAPACIDAD': 'Examen Post-Incapacidad',
                    'ALTURAS': 'Examen con énfasis en alturas'
                }
                tipo_legible = tipos_legibles.get(trab['tipo_examen'], trab['tipo_examen'])
                
                # Construir lista de exámenes
                lista_examenes = "\n".join([f"- {e.nombre}" for e in trab['examenes_bd']])
                
                cuerpo_final = (
                    f"Cordial Saludo.\n\n"
                    f"Se han programado los siguientes exámenes médicos para el trabajador:\n\n"
                    f"Nombre: {trab['nombre']}\n"
                    f"Documento: {trab['documento']}\n"
                    f"Ciudad: {trab.get('ciudad', 'No disponible')}\n"
                    f"Cargo: {trab['cargo'].nombrecargo}\n"
                    f"Empresa: {trab['empresa'].nombre_empresa}\n"
                    f"Centro Operativo: {trab['centro'].nombrecentrop if trab.get('centro') else 'No disponible'}\n"
                    f"Tipo de Examen: {tipo_legible}\n\n"
                    f"Exámenes requeridos:\n{lista_examenes}\n\n"
                    f"---\n"
                    f"ID de Lote: {uuid_correo}\n"
                    f"Solicitante: {nombre_colaborador if nombre_colaborador else 'No disponible'}\n"
                    f"Correo del solicitante: {correos_solicitantes_text}"
                )

            correos_destino = (
                ""
                #"practicante.desarrollogh@regency.com.co,"
                "operativo@servicompetentes.com,"
                "administrativo@servicompetentes.com"
            )
            # Limpiar agresivamente: remover espacios, newlines, tabs
            correos_list = [
                email.strip().replace('\r', '').replace('\n', '').replace('\t', '')
                for email in correos_destino.split(',')
                if email and email.strip()
            ]
            
            # Agregar solicitante y solicitante extra a la lista de destinatarios
            if correo_colaborador:
                correo_colaborador_clean = correo_colaborador.strip().replace('\r', '').replace('\n', '').replace('\t', '')
                correos_list.append(correo_colaborador_clean)
            if solicitante_extra:
                solicitante_extra_clean = solicitante_extra.strip().replace('\r', '').replace('\n', '').replace('\t', '')
                correos_list.append(solicitante_extra_clean)
            
            # Remover duplicados manteniendo el orden
            correos_list = list(dict.fromkeys(correos_list))
            
            # Validar que hay destinatarios
            if not correos_list:
                logger.error(f"Correo masivo UUID={uuid_correo}: Sin destinatarios válidos (correo_colaborador='{correo_colaborador}', solicitante_extra='{solicitante_extra}')")
                return Response(
                    {
                        "error": "Error: No hay destinatarios válidos para enviar el correo. Verifique el correo del solicitante."
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Actualizar correos_destino para que incluya solicitantes (para guardar en BD)
            correos_destino = ', '.join(correos_list)

            # Obtener el colaborador del usuario autenticado
            colaborador = (
                request.user.idcolaboradoru if hasattr(
                    request.user, 'idcolaboradoru') else None
            )

            # Asunto del correo
            asunto_correo = "Exámenes médicos"

            # Crear registro CorreoExamenEnviado
            correo_lote = CorreoExamenEnviado.objects.create(
                uuid_correo=uuid_correo,
                enviado_por=colaborador,
                asunto=asunto_correo,
                cuerpo_correo=cuerpo_final,
                correos_destino=correos_destino,
                tipo_examen=tipo_examen_principal,
                enviado_correctamente=False
            )

            # Crear registros RegistroExamenes (uno por trabajador)
            # Generar UUIDs manualmente porque bulk_create no ejecuta save()
            registros = []

            for trab in trabajadores_validos:
                uuid_trabajador = str(uuid.uuid4())
                # Guardar exámenes como string separado por coma
                examenes_str = ','.join(trab['examenes_nombres'])

                registro = RegistroExamenes(
                    correo_lote=correo_lote,
                    nombre_trabajador=trab['nombre'],
                    documento_trabajador=trab['documento'],
                    ciudad=trab.get('ciudad'),
                    empresa=trab['empresa'],
                    cargo=trab['cargo'],
                    centro=trab.get('centro'),
                    uuid_trabajador=uuid_trabajador,
                    tipo_examen=trab['tipo_examen'],
                    examenes_asignados=examenes_str
                )
                registros.append(registro)
                # Guardar UUID en trabajadores_validos para Excel
                trab['uuid_trabajador'] = uuid_trabajador
                trab['registro_para_relacion'] = (
                    registro, trab['examenes_bd'])

            # Guardar registros con UUIDs ya generados
            RegistroExamenes.objects.bulk_create(
                registros, ignore_conflicts=True)

            # Crear relaciones ExamenTrabajador (muchos a muchos)
            # Primero obtener los registros guardados
            from examenes.models import ExamenTrabajador
            registros_guardados = {}
            for trab in trabajadores_validos:
                registro_guardado = RegistroExamenes.objects.get(
                    uuid_trabajador=trab['uuid_trabajador'])
                registros_guardados[trab['uuid_trabajador']
                                    ] = registro_guardado

            # Crear relaciones
            examen_trabajador_batch = []
            for trab in trabajadores_validos:
                registro_guardado = registros_guardados[trab['uuid_trabajador']]
                for examen in trab['examenes_bd']:
                    et = ExamenTrabajador(
                        registro_examen=registro_guardado,
                        examen=examen
                    )
                    examen_trabajador_batch.append(et)

            # Guardar relaciones (ignorar duplicados)
            ExamenTrabajador.objects.bulk_create(
                examen_trabajador_batch, ignore_conflicts=True)
            logger.info(
                f"Creadas {
                    len(examen_trabajador_batch)} relaciones ExamenTrabajador")

            # Nota: No creamos `RegistroExamenesEnviados` para trazabilidad en
            # el flujo masivo. La fuente de verdad del estado es
            # `RegistroExamenes.estado_trabajador` y las asignaciones están
            # representadas en `ExamenTrabajador`.

            # Enviar correo SIEMPRE con Excel adjunto
            excel_buffer = None
            try:
                # Generar Excel con formato tabla y exámenes como columnas,
                # separado por tipo
                logger.info(f"Generando Excel para correo masivo UUID={uuid_correo}...")
                excel_buffer = self._generar_excel_por_tipo(
                    trabajadores_validos)
                logger.info(f"Excel generado exitosamente, tamaño: {len(excel_buffer.getvalue())} bytes")
                
                # Verificar conexión SMTP antes de enviar
                from examenes.tasks import verificar_conexion_smtp, enviar_correo_masivo_task
                smtp_ok, smtp_error = verificar_conexion_smtp()

                if not smtp_ok:
                    # SMTP no disponible: programar reintento vía Celery
                    logger.warning(
                        f"SMTP no disponible para correo masivo UUID={uuid_correo}: {smtp_error}. "
                        f"Programando reintento vía Celery."
                    )
                    correo_lote.error_envio = f"SMTP no disponible: {smtp_error}. Reintento programado."
                    correo_lote.save(update_fields=['error_envio'])

                    # Programar reenvío con el Excel adjunto
                    enviar_correo_masivo_task.apply_async(
                        args=[correo_lote.id],
                        kwargs={'excel_bytes': excel_buffer.getvalue()},
                        countdown=60
                    )

                    self._clear_cache()
                    return Response(
                        {
                            "uuid_correo": uuid_correo,
                            "total_trabajadores": len(trabajadores_validos),
                            "enviado_a": correos_list,
                            "estado": "Registrado - envío en proceso",
                            "detalle": (
                                f"Los {len(trabajadores_validos)} trabajadores fueron registrados. "
                                f"El servidor de correo no está disponible temporalmente. "
                                f"Se reintentará el envío automáticamente."
                            ),
                            "reintento_programado": True,
                            "solicitantes_notificados": {
                                "principal": correo_colaborador,
                                "extra": solicitante_extra
                            }
                        },
                        status=status.HTTP_202_ACCEPTED
                    )

                # SMTP disponible: enviar directamente
                email = EmailMultiAlternatives(
                    subject=asunto_correo,
                    body='Por favor, abra este correo en un cliente que soporte HTML.' if es_masivo else cuerpo_final,
                    from_email=settings.EMAIL_MEDICAL_FROM_EMAIL,
                    to=correos_list
                )
                
                # Adjuntar cuerpo como HTML si es masivo
                if es_masivo:
                    email.attach_alternative(cuerpo_final, "text/html")
                
                # SIEMPRE adjuntar Excel (incluso si es 1 sola persona)
                # Esto es consistente con el requerimiento: "si es masivo, envía Excel igual"
                if excel_buffer:
                    email.attach(
                        'Trabajadores_Examenes.xlsx',
                        excel_buffer.getvalue(),
                        'application/vnd.openxmlformats-officedocument.'
                        'spreadsheetml.sheet'
                    )
                
                # Usar backend médico para enviar
                email.connection = self._get_medical_email_backend()
                email.send(fail_silently=False)

                correo_lote.enviado_correctamente = True
                correo_lote.error_envio = None
                correo_lote.save(update_fields=['enviado_correctamente', 'error_envio'])

                logger.info(
                    f"Correo masivo UUID={uuid_correo} enviado exitosamente "
                    f"a {len(correos_list)} destinatarios con {len(trabajadores_validos)} trabajadores"
                )

            except Exception as e:
                error_msg = f"{type(e).__name__}: {str(e)}"
                correo_lote.error_envio = error_msg
                correo_lote.save(update_fields=['error_envio'])

                logger.error(
                    f"Error enviando correo masivo UUID={uuid_correo}: {error_msg}",
                    exc_info=True
                )

                # Intentar programar reintento vía Celery
                reintento_programado = False
                try:
                    from examenes.tasks import enviar_correo_masivo_task
                    enviar_correo_masivo_task.apply_async(
                        args=[correo_lote.id],
                        kwargs={'excel_bytes': excel_buffer.getvalue() if excel_buffer else None},
                        countdown=120
                    )
                    reintento_programado = True
                    logger.info(f"Reintento Celery programado para correo masivo id={correo_lote.id}")
                except Exception as celery_err:
                    logger.error(f"No se pudo programar reintento Celery: {str(celery_err)}")

                return Response(
                    {
                        "error": f"Error al enviar correo: {str(e)}",
                        "uuid_correo": uuid_correo,
                        "trabajadores_registrados": len(registros),
                        "reintento_programado": reintento_programado,
                        "detalle": (
                            "Los trabajadores fueron registrados en la base de datos. "
                            + ("Se reintentará el envío automáticamente." if reintento_programado
                               else "Por favor reintente manualmente.")
                        )
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            # Respuesta exitosa
            # Limpiar cache de reportes y datos de empresas
            self._clear_cache()

            return Response(
                {
                    "uuid_correo": uuid_correo,
                    "total_trabajadores": len(trabajadores_validos),
                    "enviado_a": correos_list,
                    "estado": "Enviado exitosamente",
                    "detalle": (
                        f"Se envió correo a {len(correos_list)} "
                        f"destinatarios con {len(trabajadores_validos)} "
                        f"trabajadores"
                    ),
                    "solicitantes_notificados": {
                        "principal": correo_colaborador,
                        "extra": solicitante_extra
                    }
                },
                status=status.HTTP_201_CREATED
            )

        except Exception as e:
            import traceback
            logger = logging.getLogger(__name__)

            logger.error("=== ERROR EN EnviarCorreoMasivoView ===")
            logger.error(f"Exception type: {type(e).__name__}")
            logger.error(f"Exception message: {str(e)}")
            logger.error(
                f"Traceback:\n{traceback.format_exc()}"
            )

            error_msg = (
                f"Error interno: {str(e)}\n"
                f"{traceback.format_exc()}"
            )
            return Response(
                {"error": error_msg},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _generar_excel_examenes(self, trabajadores):
        """Genera Excel con formato tabla donde los exámenes son columnas con X"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO

        # Obtener solo los exámenes que al menos un trabajador requiere
        examenes_requeridos = set()
        for trab in trabajadores:
            examenes_requeridos.update(trab['examenes'])

        # Ordenar alfabéticamente los exámenes requeridos
        nombres_examenes = sorted(list(examenes_requeridos))

        wb = Workbook()
        ws = wb.active
        ws.title = "Trabajadores Examenes"

        # Estilos
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border_style = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Encabezados: UUID primero, luego datos base con Cargo después de
        # Centro
        headers = [
            "UUID",
            "Empresa",
            "Unidad",
            "Proyecto",
            "Centro",
            "Cargo",
            "Nombre",
            "Documento"] + nombres_examenes
        ws.append(headers)

        # Aplicar estilos a encabezados
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border_style

        # Agregar datos de trabajadores
        for trab in trabajadores:
            row_data = [
                trab.get('uuid_trabajador', ''),  # UUID primero
                trab['empresa'].nombre_empresa,
                trab['unidad'].nombreunidad,
                trab['proyecto'].nombreproyecto,
                (trab['centro'].nombrecentrop if trab.get('centro') else ''),
                trab['cargo'].nombrecargo,  # Cargo después de Centro
                trab['nombre'],
                trab['documento']
            ]

            # Para cada examen, verificar si el cargo de este trabajador lo
            # requiere
            # exámenes que requiere este trabajador
            examenes_trabajador = set(trab['examenes'])

            for nombre_examen in nombres_examenes:
                if nombre_examen in examenes_trabajador:
                    row_data.append('X')
                else:
                    row_data.append('')

            ws.append(row_data)

        # Aplicar bordes y centrado a todas las celdas de datos
        for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_style
                # Centrar las X de los exámenes (ahora desde columna 9 porque
                # agregamos UUID)
                if cell.column >= 9:  # Columnas de exámenes
                    cell.alignment = center_alignment

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40  # UUID (primera columna)
        ws.column_dimensions['B'].width = 25  # Empresa
        ws.column_dimensions['C'].width = 20  # Unidad
        ws.column_dimensions['D'].width = 25  # Proyecto
        ws.column_dimensions['E'].width = 20  # Centro
        ws.column_dimensions['F'].width = 25  # Cargo
        ws.column_dimensions['G'].width = 25  # Nombre
        ws.column_dimensions['H'].width = 15  # Documento

        # Ajustar ancho de columnas de exámenes (más pequeñas)
        for col_num in range(9, 9 + len(nombres_examenes)):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 5

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _generar_excel_por_tipo(self, trabajadores):
        """Genera Excel con una sola hoja con todos los trabajadores, incluyendo tipo de examen"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "Trabajadores Examenes"

        # Estilos comunes
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border_style = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Obtener todos los exámenes únicos de todos los trabajadores
        examenes_requeridos = set()
        for trab in trabajadores:
            examenes_requeridos.update(trab['examenes_nombres'])
        nombres_examenes = sorted(list(examenes_requeridos))

        # Encabezados - incluye "Ciudad" y "Tipo Examen"
        headers = [
            "UUID",
            "Empresa",
            "Unidad",
            "Proyecto",
            "Centro",
            "Ciudad",
            "Cargo",
            "Nombre",
            "Documento",
            "Tipo Examen"] + nombres_examenes
        ws.append(headers)

        # Aplicar estilos a encabezados
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border_style

        # Ordenar trabajadores por tipo de examen
        orden_tipos = ['INGRESO', 'PERIODICO', 'RETIRO', 'ESPECIAL', 'POST_INCAPACIDAD', 'ALTURAS']
        trabajadores_ordenados = sorted(
            trabajadores,
            key=lambda x: orden_tipos.index(x['tipo_examen']) if x['tipo_examen'] in orden_tipos else 999
        )

        # Agregar datos de todos los trabajadores
        for trab in trabajadores_ordenados:
            row_data = [
                trab.get('uuid_trabajador', ''),
                trab['empresa'].nombre_empresa,
                trab['unidad'].nombreunidad,
                trab['proyecto'].nombreproyecto,
                (trab['centro'].nombrecentrop if trab.get('centro') else ''),
                trab.get('ciudad', ''),  # Ciudad
                trab['cargo'].nombrecargo,
                trab['nombre'],
                trab['documento'],
                trab['tipo_examen']
            ]

            # Exámenes con X donde aplica
            examenes_trabajador = set(trab['examenes_nombres'])
            for nombre_examen in nombres_examenes:
                if nombre_examen in examenes_trabajador:
                    row_data.append('X')
                else:
                    row_data.append('')

            ws.append(row_data)

        # Aplicar bordes y centrado a todas las filas de datos
        for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border_style
                if cell.column >= 11:  # Columnas de exámenes empiezan en 11 (después de Ciudad)
                    cell.alignment = center_alignment

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40  # UUID
        ws.column_dimensions['B'].width = 25  # Empresa
        ws.column_dimensions['C'].width = 20  # Unidad
        ws.column_dimensions['D'].width = 25  # Proyecto
        ws.column_dimensions['E'].width = 20  # Centro
        ws.column_dimensions['F'].width = 18  # Ciudad
        ws.column_dimensions['G'].width = 25  # Cargo
        ws.column_dimensions['H'].width = 25  # Nombre
        ws.column_dimensions['I'].width = 15  # Documento
        ws.column_dimensions['J'].width = 18  # Tipo Examen

        # Columnas de exámenes (empiezan en columna 11 = K)
        for col_num in range(11, 11 + len(nombres_examenes)):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 5

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

class ListarTrabajadoresCorreoView(APIView):
    """
    Endpoint para listar trabajadores de un correo con paginación.

    Optimización:
    - Prefetch de `ExamenTrabajador` para evitar N+1 queries
    - Usa `to_attr` para acceso directo sin queries adicionales
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]
    pagination_class = PageNumberPagination

    def _serializar_trabajador(self, trabajador):
        """Serializa un trabajador con sus exámenes de RegistroExamenesEnviados (ya precargados)"""
        # OPTIMIZADO: Usar exámenes precargados con Prefetch
        # Usar ExamenTrabajador precargado para listar exámenes asignados
        examenes_rel = getattr(trabajador, 'examenes_precargados', [])

        examenes_detalle = []
        for ex_rel in examenes_rel:
            examen = getattr(ex_rel, 'examen', None)
            examenes_detalle.append({
                'id': getattr(ex_rel, 'id', None),
                'examen_id': examen.id_examen if examen else None,
                'examen_nombre': examen.nombre if examen else 'N/A',
                # No disponemos de estado por examen cuando no usamos
                # RegistroExamenesEnviados; mostramos estado a nivel de trabajador
                'tipo_examen': trabajador.tipo_examen,
                'estado': None,
                'resultado': None,
                'fecha_envio': getattr(ex_rel, 'fecha_asignacion', None),
                'fecha_completado': None
            })

        return {
            'id': trabajador.id,
            'correo_id': trabajador.correo_lote_id,
            'uuid_trabajador': trabajador.uuid_trabajador,
            'nombre_trabajador': trabajador.nombre_trabajador,
            'documento_trabajador': trabajador.documento_trabajador,
            'cargo_nombre': trabajador.cargo.nombrecargo if trabajador.cargo else None,
            'empresa_nombre': trabajador.empresa.nombre_empresa if trabajador.empresa else None,
            'tipo_examen': trabajador.tipo_examen,
            'examenes_asignados': trabajador.examenes_asignados,
            'examenes': examenes_detalle,
            'total_examenes': len(examenes_detalle),
            # Cuando no se usa RegistroExamenesEnviados, el estado es a nivel
            # de trabajador y se refleja en `estado_trabajador`.
            'examenes_completados': (len(examenes_detalle) if trabajador.estado_trabajador == 1 else 0),
            'examenes_pendientes': (0 if trabajador.estado_trabajador == 1 else len(examenes_detalle)),
            'estado_trabajador': trabajador.estado_trabajador,
            'estado_nombre': "Completado" if trabajador.estado_trabajador == 1 else "Pendiente"
        }

    def get(self, request, correo_id):
        """
        Obtiene lista paginada de trabajadores de un correo

        Parámetros query:
        - page: número de página (default 1)
        """
        try:
            correo = CorreoExamenEnviado.objects.get(id=correo_id)
        except CorreoExamenEnviado.DoesNotExist:
            return Response(
                {"error": "Correo no encontrado"},
                status=status.HTTP_404_NOT_FOUND
            )

        # OPTIMIZADO: Prefetch de exámenes para evitar N+1 queries en serialización

        search = request.query_params.get('search', '').strip()
        trabajadores_qs = RegistroExamenes.objects.filter(
            correo_lote=correo
        ).select_related('empresa', 'cargo').prefetch_related(
            Prefetch(
                'examenes',
                queryset=ExamenTrabajador.objects.select_related('examen'),
                to_attr='examenes_precargados'
            )
        ).order_by('-fecha_registro')

        if search:
            from django.db.models import Q
            trabajadores_qs = trabajadores_qs.filter(
                Q(uuid_trabajador__icontains=search) |
                Q(documento_trabajador__icontains=search)
            )
        trabajadores = trabajadores_qs

        # Cache por correo_id + paginación
        page = request.query_params.get('page', '1')
        page_size = request.query_params.get('page_size', '25')
        cache_key = f"trabajadores_correo_v2={correo_id}_page={page}_size={page_size}"

        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached, status=status.HTTP_200_OK, headers={'X-Cache': 'HIT'})

        # Paginar resultados
        paginator = self.pagination_class()
        paginated_trabajadores = paginator.paginate_queryset(
            trabajadores, request)

        if paginated_trabajadores is not None:
            # Serializar con exámenes individuales
            results = [self._serializar_trabajador(t) for t in paginated_trabajadores]
            
            paginated_response = paginator.get_paginated_response(results)
            # Agregar metadata del correo
            paginated_response.data.update({
                "correo_id": correo.id,
                "uuid_correo": getattr(correo, 'uuid_correo', None),
                "asunto": correo.asunto,
                "tipo_examen_lote": correo.tipo_examen,
                "fecha_envio": getattr(correo, 'fecha_envio', None),
                "total_trabajadores": trabajadores.count(),
                "tipos_examen_disponibles": TIPOS_EXAMEN_VALIDOS
            })
            paginated_response['X-Cache'] = 'MISS'
            cache.set(cache_key, paginated_response.data, timeout=300)
            return paginated_response

        # Fallback sin paginación (estructura equivalente)
        results = [self._serializar_trabajador(t) for t in trabajadores]
        data = {
            "count": len(results),
            "next": None,
            "previous": None,
            "results": results,
            "correo_id": correo.id,
            "uuid_correo": getattr(correo, 'uuid_correo', None),
            "asunto": correo.asunto,
            "tipo_examen_lote": correo.tipo_examen,
            "fecha_envio": getattr(correo, 'fecha_envio', None),
            "total_trabajadores": trabajadores.count(),
            "tipos_examen_disponibles": TIPOS_EXAMEN_VALIDOS
        }
        cache.set(cache_key, data, timeout=300)
        return Response(data, status=status.HTTP_200_OK, headers={'X-Cache': 'MISS'})


class ListarRegistrosPorTipoExamenView(APIView):
    """
    Endpoint para listar registros filtrados por tipo de examen.
    
    Optimización:
    - Prefetch de exámenes enviados para evitar N+1 queries
    - Cache de 5 minutos por tipo de examen
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def get(self, request):
        tipo_examen = request.query_params.get('tipo', '').upper()

        if not tipo_examen or tipo_examen not in TIPOS_EXAMEN_VALIDOS:
            return Response(
                {
                    "error": f"Tipo invalido. Debe ser uno de: {', '.join(TIPOS_EXAMEN_VALIDOS)}"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Intentar obtener de cache
        cache_key = f"registros_tipo_examen_{tipo_examen}"
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached, status=status.HTTP_200_OK, headers={'X-Cache': 'HIT'})

        # OPTIMIZADO: Prefetch de exámenes para evitar N+1 queries
        registros = RegistroExamenes.objects.filter(
            tipo_examen=tipo_examen
        ).select_related('correo_lote', 'empresa', 'cargo').prefetch_related(
            Prefetch(
                'examenes',
                queryset=ExamenTrabajador.objects.select_related('examen'),
                to_attr='examenes_precargados'
            )
        ).order_by('-id')[:50]

        resultados = []
        for reg in registros:
            # OPTIMIZADO: Usar exámenes precargados (sin query adicional)
            examenes_enviados = getattr(reg, 'examenes_precargados', [])

            # Preparar detalle de exámenes a partir de ExamenTrabajador
            examenes_detalle = []
            for ex_rel in examenes_enviados:
                examen = getattr(ex_rel, 'examen', None)
                examenes_detalle.append({
                    'id': getattr(ex_rel, 'id', None),
                    'examen_id': examen.id_examen if examen else None,
                    'nombre': examen.nombre if examen else 'N/A',
                    'tipo_examen': reg.tipo_examen,
                    'estado': None,
                    'resultado': None,
                    'fecha_envio': getattr(ex_rel, 'fecha_asignacion', None),
                    'fecha_completado': None
                })

            resultados.append({
                'id': reg.id,
                'uuid_trabajador': reg.uuid_trabajador,
                'nombre': reg.nombre_trabajador,
                'documento': reg.documento_trabajador,
                'empresa': reg.empresa.nombre_empresa if reg.empresa else None,
                'cargo': reg.cargo.nombrecargo if reg.cargo else None,
                'tipo_examen': reg.tipo_examen,
                'estado_trabajador': 'Completado' if reg.estado_trabajador == 1 else 'Pendiente',
                'examenes': examenes_detalle,
                'total_examenes': len(examenes_detalle),
                'examenes_completados': (len(examenes_detalle) if reg.estado_trabajador == 1 else 0),
                'examenes_pendientes': (0 if reg.estado_trabajador == 1 else len(examenes_detalle))
            })

        response_data = {
            'tipo_examen': tipo_examen,
            'tipos_validos': TIPOS_EXAMEN_VALIDOS,
            'total': len(resultados),
            'registros': resultados
        }
        
        # Guardar en cache (5 minutos)
        cache.set(cache_key, response_data, timeout=300)

        return Response(response_data, status=status.HTTP_200_OK, headers={'X-Cache': 'MISS'})


class ActualizarEstadoExamenesMasivoView(APIView):
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def patch(self, request):
        serializer = ActualizarEstadoExamenesSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        trabajador_ids = serializer.validated_data['trabajador_ids']

        from .models import RegistroExamenes, HistorialEstadoRegistroExamen
        actualizados = []
        no_encontrados = []
        cambios = []
        correos_actualizados = set()  # Recolectar UUIDs únicos de correos actualizados
        colaboradores_actualizados = set()  # Recolectar IDs de colaboradores que han enviado estos correos
        
        for tid in trabajador_ids:
            try:
                reg = RegistroExamenes.objects.get(id=tid)
                estado_anterior = reg.estado_trabajador
                # Se fija explícitamente a Completado (no se invierte el estado actual)
                # para evitar revertir a Pendiente trabajadores que ya estaban completados
                # cuando se seleccionan junto con otros en una acción masiva.
                reg.estado_trabajador = 1
                reg.save()

                # Solo se audita cuando realmente hubo transición a Completado
                if estado_anterior != 1:
                    HistorialEstadoRegistroExamen.objects.create(
                        registro_examen=reg,
                        colaborador=request.user.idcolaboradoru,
                    )

                # No sincronizamos RegistroExamenesEnviados: el estado del trabajador
                # es ahora la fuente de verdad (decoupled). Solo registramos el cambio
                actualizados.append(tid)
                cambios.append({'id': tid, 'de': estado_anterior, 'a': reg.estado_trabajador})
                
                # Recolectar UUID del correo y colaborador para limpiar su caché después
                if reg.correo_lote:
                    if reg.correo_lote.uuid_correo:
                        correos_actualizados.add(reg.correo_lote.uuid_correo)
                    if reg.correo_lote.enviado_por_id:
                        colaboradores_actualizados.add(reg.correo_lote.enviado_por_id)

                # Invalidar cache de listados del correo para reflejar el cambio
                try:
                    correo_id = reg.correo_lote.id
                    # Limpiar trabajadores_correo_v2 - listados de trabajadores del correo
                    for sz in (10, 25, 50, 100):
                        cache.delete(f"trabajadores_correo_v2={correo_id}_page=1_size={sz}")
                        
                    # Limpiar reporte_correos - reporte general de correos
                    for page in range(1, 100):  # Invalida hasta página 100
                        for sz in (10, 25, 50, 100):
                            cache.delete(f"reporte_correos_page={page}_size={sz}")
                            
                except Exception as e:
                    print(f"Error limpiando caché de correo: {e}")
            except RegistroExamenes.DoesNotExist:
                no_encontrados.append(tid)
        
        # Limpiar caché de filtrar-examenes para los colaboradores y correos actualizados
        try:
            # Limpiar lista general de colaboradores
            cache.delete("filtrar_examenes_colaboradores")
            
            # Limpiar caché por colaborador (todas las páginas y tamaños)
            for colaborador_id in colaboradores_actualizados:
                if colaborador_id:
                    for page in range(1, 100):
                        for sz in (10, 25, 50, 100):
                            cache.delete(f"filtrar_examenes_colaborador={colaborador_id}_page={page}_size={sz}")
            
            # Limpiar caché de búsqueda por UUID para los correos actualizados
            # Patrones backend: filtrar_examenes_uuid={uuid_correo}
            for uuid_correo in correos_actualizados:
                # Limpiar patrón del backend (sin paginación, devuelve un correo o múltiples)
                cache.delete(f"filtrar_examenes_uuid={uuid_correo}")
                
                # Limpiar patrones del frontend deduplicados con paginación
                # Patrón: examenes:FiltrarExamenesPorUUID:{uuid}:page={page}:size={size}
                for page in range(1, 20):
                    for sz in (10, 25, 50, 100):
                        cache_key = f"examenes:FiltrarExamenesPorUUID:{uuid_correo}:page={page}:size={sz}"
                        cache.delete(cache_key)
                
                # También limpiar sin paginación por si acaso
                cache_key_simple = f"examenes:FiltrarExamenesPorUUID:{uuid_correo}"
                cache.delete(cache_key_simple)
                
        except Exception as e:
            print(f"Error limpiando caché de filtrar-examenes: {e}")

        return Response({
            'actualizados': actualizados,
            'no_encontrados': no_encontrados,
            'cambios': cambios
        }, status=status.HTTP_200_OK)


# --- ENDPOINT: CrearExamenView ---
class CrearExamenView(APIView):
    """
    Endpoint para crear un examen y asociarlo a empresas y cargos.
    - GET: Devuelve empresas y cargos disponibles.
    - POST: Crea el examen y lo asocia a todas las combinaciones de empresas y cargos recibidos.
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def get(self, request):
        """
        Devuelve la lista de empresas y cargos para selección en el frontend.
        """
        empresas = Epresa.objects.all().values('idempresa', 'nombre_empresa').exclude(estadoempresa = 0)
        cargos = Cargo.objects.all().values('idcargo', 'nombrecargo')
        return Response({
            'empresas': list(empresas),
            'cargos': list(cargos)  
        }, status=status.HTTP_200_OK)

    def post(self, request):
        """
        Crea un examen y lo asocia a las empresas y cargos seleccionados.
        Requiere al menos una empresa y un cargo.
        """
        serializer = CrearExamenSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        nombre = serializer.validated_data['nombre']
        empresas_ids = serializer.validated_data.get('empresas_ids', [])
        cargos_ids = serializer.validated_data.get('cargos_ids', [])
        tipos = serializer.validated_data.get('tipos', ['INGRESO'])

        # Validar que haya al menos una empresa y un cargo
        if not empresas_ids or not cargos_ids:
            return Response({
                'error': 'Debe seleccionar al menos una empresa y un cargo para asociar el examen.'
            }, status=status.HTTP_400_BAD_REQUEST)

        examen = Examen.objects.create(
            nombre=nombre,
        )

        # Asociar examen a todas las combinaciones de empresa, cargo y tipo
        for empresa_id in empresas_ids:
            for cargo_id in cargos_ids:
                for tipo in tipos:
                    ExamenesCargo.objects.create(
                        examen=examen,
                        empresa_id=empresa_id,
                        cargo_id=cargo_id,
                        tipo=tipo
                    )

        # Limpiar cache de datos de empresas con exámenes (se actualizó el catálogo)
        self._clear_cache()

        return Response({
            'id_examen': examen.id_examen,
            'nombre': examen.nombre,
            'empresas_ids': empresas_ids,
            'cargos_ids': cargos_ids,
            'tipos': tipos
        }, status=status.HTTP_201_CREATED)

    def _clear_cache(self):
        """Limpia el cache de datos de empresas con exámenes."""
        logger = logging.getLogger(__name__)
        try:
            # Limpiar cache de datos de empresas con exámenes
            cache.delete('cargo_empresa_examenes_data')
            logger.info("Cache limpiado: datos de empresas con exámenes (CrearExamenView)")
        except Exception as e:
            logger.warning(f"Error al limpiar cache en CrearExamenView: {str(e)}")


class FiltrarExamenesView(APIView):
    """
    Endpoint para filtrar correos enviados por colaboradores.
    
    GET sin parámetro: Devuelve lista de colaboradores que han enviado correos de exámenes.
    GET con parámetro enviado_por_id: Devuelve correos enviados por ese colaborador,
                                       con formato paginado como ReporteCorreosEnviadosView.
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def get(self, request):
        enviado_por_id = request.query_params.get('enviado_por_id', None)
        uuid_correo = request.query_params.get('uuid', None)

        if uuid_correo:
            # Filtrar por UUID del correo
            return self._get_correo_por_uuid(request, uuid_correo)
        elif not enviado_por_id:
            # Devolver lista de colaboradores que han enviado correos
            return self._get_colaboradores_que_enviaron(request)
        else:
            # Devolver correos enviados por un colaborador específico (formato paginado)
            return self._get_correos_por_colaborador(request, enviado_por_id)

    def _get_colaboradores_que_enviaron(self, request):
        """
        Devuelve lista de colaboradores que han enviado correos de exámenes.
        """
        cache_key = "filtrar_examenes_colaboradores"
        
        # Intentar obtener del cache
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached, status=status.HTTP_200_OK, headers={'X-Cache': 'HIT'})

        try:
            # Obtener colaboradores únicos que han enviado correos
            colaboradores_ids = CorreoExamenEnviado.objects.exclude(
                enviado_por_id=None
            ).values_list('enviado_por_id', flat=True).distinct()

            # Importar el modelo de colaboradores
            from usuarios.models import Colaboradores
            colaboradores = Colaboradores.objects.filter(
                idcolaborador__in=colaboradores_ids
            ).values('idcolaborador', 'nombrecolaborador').order_by('nombrecolaborador')

            colaboradores_list = [
                {"id": c["idcolaborador"], "nombre": c["nombrecolaborador"]}
                for c in colaboradores
            ]

            response_data = {
                "total": len(colaboradores_list),
                "colaboradores": colaboradores_list
            }

            # Guardar en cache (5 minutos)
            cache.set(cache_key, response_data, timeout=300)

            return Response(response_data, status=status.HTTP_200_OK, headers={'X-Cache': 'MISS'})

        except Exception as e:
            return Response(
                {"error": f"Error obteniendo colaboradores: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _get_correos_por_colaborador(self, request, enviado_por_id):
        """
        Devuelve correos enviados por un colaborador específico con formato paginado
        igual a ReporteCorreosEnviadosView.
        """
        # Cache con parámetros de paginación
        page = request.query_params.get('page', '1')
        page_size = request.query_params.get('page_size', '25')
        cache_key = f"filtrar_examenes_colaborador={enviado_por_id}_page={page}_size={page_size}"

        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached, status=status.HTTP_200_OK, headers={'X-Cache': 'HIT'})

        # Validar que el colaborador existe
        try:
            from usuarios.models import Colaboradores
            colaborador = Colaboradores.objects.get(idcolaborador=enviado_por_id)
        except Exception:
            return Response(
                {"error": f"Colaborador con id {enviado_por_id} no encontrado"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Obtener correos del colaborador con relaciones optimizadas
        correos_queryset = CorreoExamenEnviado.objects.filter(
            enviado_por_id=enviado_por_id
        ).select_related('enviado_por').order_by('-fecha_envio')

        # Aplicar paginación
        paginator = PageNumberPagination()
        paginator.page_size = 25
        paginator.page_size_query_param = 'page_size'
        paginator.max_page_size = 100
        
        paginated_correos = paginator.paginate_queryset(correos_queryset, request, view=self)

        if paginated_correos is not None:
            # Serializar página actual
            serializer = ReporteCorreoSerializer(paginated_correos, many=True)
            # Respuesta estándar de DRF con paginación
            paginated_response = paginator.get_paginated_response(serializer.data)
            # Agregar metadata del colaborador
            paginated_response.data.update({
                "enviado_por_id": enviado_por_id,
                "nombre_colaborador": colaborador.nombrecolaborador,
                "total_correos": correos_queryset.count()
            })
            paginated_response['X-Cache'] = 'MISS'
            cache.set(cache_key, paginated_response.data, timeout=300)
            return paginated_response

        # Fallback sin paginación (poco probable)
        serializer = ReporteCorreoSerializer(correos_queryset, many=True)
        data = {
            "count": len(serializer.data),
            "next": None,
            "previous": None,
            "results": serializer.data,
            "enviado_por_id": enviado_por_id,
            "nombre_colaborador": colaborador.nombrecolaborador,
            "total_correos": correos_queryset.count()
        }
        cache.set(cache_key, data, timeout=300)
        return Response(data, status=status.HTTP_200_OK, headers={'X-Cache': 'MISS'})

    def _get_correo_por_uuid(self, request, uuid_correo):
        """
        Devuelve correos filtrados por UUID o nombre del trabajador.
        Si no encuentra por UUID, busca por nombre del trabajador.
        Retorna formato compatible con ReporteCorreosEnviadosView.
        """
        cache_key = f"filtrar_examenes_uuid={uuid_correo}"

        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached, status=status.HTTP_200_OK, headers={'X-Cache': 'HIT'})

        try:
            # Primero intentar buscar por UUID del correo
            try:
                correo = CorreoExamenEnviado.objects.select_related('enviado_por').prefetch_related('trabajadores').get(
                    uuid_correo=uuid_correo
                )
                
                # Serializar el correo encontrado
                serializer = ReporteCorreoSerializer(correo)
                response_data = {
                    "found": True,
                    "correo": serializer.data,
                    "uuid": uuid_correo,
                    "search_type": "uuid"
                }

                # Guardar en cache
                cache.set(cache_key, response_data, timeout=300)
                return Response(response_data, status=status.HTTP_200_OK, headers={'X-Cache': 'MISS'})
                
            except CorreoExamenEnviado.DoesNotExist:
                # Si no existe por UUID, buscar por nombre del trabajador
                pass

            # Buscar correos que contengan trabajadores con ese nombre
            from django.db.models import Q
            correos = CorreoExamenEnviado.objects.filter(
                trabajadores__nombre_trabajador__icontains=uuid_correo
            ).select_related('enviado_por').prefetch_related('trabajadores').distinct().order_by('-fecha_envio')

            if not correos.exists():
                return Response(
                    {"found": False, "error": f"No se encontró correo con UUID o nombre de trabajador: {uuid_correo}"},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Si hay múltiples correos, retornar formato de lista paginada
            if correos.count() > 1:
                # Aplicar paginación
                page = request.query_params.get('page', '1')
                page_size = request.query_params.get('page_size', '25')
                
                paginator = PageNumberPagination()
                paginator.page_size = int(page_size)
                paginator.page_size_query_param = 'page_size'
                paginator.max_page_size = 100
                
                paginated_correos = paginator.paginate_queryset(correos, request, view=self)

                if paginated_correos is not None:
                    serializer = ReporteCorreoSerializer(paginated_correos, many=True)
                    paginated_response = paginator.get_paginated_response(serializer.data)
                    paginated_response.data.update({
                        "search_type": "trabajador",
                        "search_term": uuid_correo,
                        "total_correos": correos.count()
                    })
                    cache.set(cache_key, paginated_response.data, timeout=300)
                    return paginated_response

            # Si hay un solo correo, retornar formato unitario
            serializer = ReporteCorreoSerializer(correos.first())
            response_data = {
                "found": True,
                "correo": serializer.data,
                "search_term": uuid_correo,
                "search_type": "trabajador",
                "total_found": 1
            }

            cache.set(cache_key, response_data, timeout=300)
            return Response(response_data, status=status.HTTP_200_OK, headers={'X-Cache': 'MISS'})

        except Exception as e:
            return Response(
                {"error": f"Error buscando correo: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ListarExamenesDisponiblesView(APIView):
    """
    Lista todos los exámenes activos disponibles en la base de datos.
    Útil para verificar qué nombres de exámenes usar en el CSV.
    
    GET: Retorna lista de exámenes con id y nombre.
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def get(self, request):
        """Lista exámenes activos para validar columnas del CSV."""
        examenes = Examen.objects.filter(activo=True).order_by('nombre')
        
        data = {
            "total": examenes.count(),
            "examenes": [
                {
                    "id": e.id_examen,
                    "nombre": e.nombre,
                    "nombre_normalizado": e.nombre.lower().strip()
                }
                for e in examenes
            ],
            "nota": "Los nombres de columnas en el CSV deben coincidir exactamente con 'nombre' o 'nombre_normalizado'"
        }
        
        return Response(data, status=status.HTTP_200_OK)


class ValidarCSVExamenesView(APIView):
    """
    Valida un archivo CSV de exámenes SIN enviar correo.
    Útil para verificar que el CSV está bien formateado antes de enviarlo.
    
    POST: Recibe archivo CSV y retorna validación detallada.
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def post(self, request):
        """
        Valida CSV sin enviar correo.
        
        Request:
        {
            "archivo_csv": <file>
        }
        
        Response:
        {
            "formato_detectado": "columnas_examenes" | "original",
            "columnas_examenes_csv": ["OPTOMETRIA", "AUDIOMETRIA", ...],
            "examenes_encontrados_bd": [...],
            "examenes_no_encontrados": [...],
            "trabajadores_validos": 24,
            "trabajadores_con_errores": 2,
            "errores": [...],
            "validacion_exitosa": true/false
        }
        """
        import io
        import csv
        logger = logging.getLogger(__name__)
        
        # Validar archivo
        archivo_csv = request.FILES.get('archivo_csv')
        if not archivo_csv:
            return Response(
                {"error": "Se requiere archivo_csv"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not archivo_csv.name.endswith('.csv'):
            return Response(
                {"error": "El archivo debe ser CSV (.csv)"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Decodificar archivo
            archivo_csv.seek(0)
            contenido_csv = None
            encoding_usado = None
            
            for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'iso-8859-1', 'cp1252']:
                try:
                    archivo_csv.seek(0)
                    contenido_csv = archivo_csv.read().decode(encoding)
                    encoding_usado = encoding
                    break
                except (UnicodeDecodeError, AttributeError):
                    continue
            
            if contenido_csv is None:
                return Response(
                    {"error": "No se pudo decodificar el CSV"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Detectar delimitador
            delimiter = ','
            try:
                # Obtener la primera línea del CSV
                primera_linea = contenido_csv.split('\n')[0]
                
                # Contar ocurrencias de posibles delimitadores
                contar_comas = primera_linea.count(',')
                contar_puntoycoma = primera_linea.count(';')
                
                # Si hay más puntos y comas que comas, usar punto y coma
                if contar_puntoycoma > contar_comas and contar_puntoycoma > 0:
                    delimiter = ';'
                else:
                    # Intentar con el sniffer como alternativa
                    sniffer = csv.Sniffer()
                    try:
                        sample = contenido_csv[:1024]
                        dialect = sniffer.sniff(sample, delimiters=',;')
                        delimiter = dialect.delimiter
                    except Exception:
                        delimiter = ','
            except Exception as e:
                logger.warning(f"Error en detección de delimitador: {str(e)}")
                delimiter = ','
            
            # Leer CSV
            stream = io.StringIO(contenido_csv)
            reader = csv.DictReader(stream, delimiter=delimiter)
            
            if not reader.fieldnames:
                return Response(
                    {"error": "CSV sin encabezados"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            fieldnames_original = [f.strip() for f in reader.fieldnames]
            fieldnames_lower = [f.lower() for f in fieldnames_original]
            
            # Detectar formato
            formato_csv = 'unknown'
            columnas_examenes = []
            
            if 'examenes' in fieldnames_lower:
                formato_csv = 'original'
            elif 'tipo de examen' in fieldnames_lower:
                formato_csv = 'columnas_examenes'
                tipo_idx = fieldnames_lower.index('tipo de examen')
                columnas_examenes = [c for c in fieldnames_original[tipo_idx + 1:] if c and c.strip()]
            elif 'tipoexamen' in fieldnames_lower:
                formato_csv = 'original'
            
            # Cargar exámenes de BD
            examenes_bd = {
                e.nombre.lower().strip(): e.nombre
                for e in Examen.objects.filter(activo=True)
            }
            
            # Validar columnas de exámenes
            examenes_encontrados = []
            examenes_no_encontrados = []
            
            if formato_csv == 'columnas_examenes':
                for col in columnas_examenes:
                    col_lower = col.lower().strip()
                    if col_lower in examenes_bd:
                        examenes_encontrados.append({
                            "columna_csv": col,
                            "nombre_bd": examenes_bd[col_lower]
                        })
                    else:
                        examenes_no_encontrados.append(col)
            
            # Contar trabajadores
            trabajadores_data = list(reader)
            filas_no_vacias = [
                t for t in trabajadores_data
                if any(v and v.strip() for v in t.values())
            ]
            
            # Construir respuesta
            response_data = {
                "formato_detectado": formato_csv,
                "encoding_usado": encoding_usado,
                "delimitador": delimiter,
                "columnas_csv": fieldnames_original,
                "total_filas": len(trabajadores_data),
                "filas_no_vacias": len(filas_no_vacias),
            }
            
            if formato_csv == 'columnas_examenes':
                response_data.update({
                    "columnas_examenes_csv": columnas_examenes,
                    "examenes_encontrados_bd": examenes_encontrados,
                    "examenes_no_encontrados": examenes_no_encontrados,
                    "total_examenes_csv": len(columnas_examenes),
                    "examenes_validos": len(examenes_encontrados),
                })
            
            response_data["validacion_exitosa"] = (
                formato_csv != 'unknown' and 
                (formato_csv == 'original' or len(examenes_no_encontrados) == 0)
            )
            
            if not response_data["validacion_exitosa"]:
                response_data["sugerencia"] = (
                    "Verifique que los nombres de columnas de exámenes coincidan "
                    "exactamente con los nombres en la base de datos. "
                    "Use el endpoint /api/examenes/listar-examenes/ para ver los nombres correctos."
                )
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            import traceback
            logger.error(f"Error validando CSV: {traceback.format_exc()}")
            return Response(
                {"error": f"Error procesando CSV: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        

class GestionarExamenesCargoView(APIView):
    """
    Endpoint para gestionar (ver, agregar, eliminar) exámenes asignados a un cargo
    en una empresa específica, filtrados por tipo de examen.

    GET: Retorna exámenes asignados a un cargo en una empresa, filtrados por tipo.
         Params: empresa_id, cargo_id, tipo (opcional)
    POST: Agrega exámenes a un cargo en una empresa para un tipo.
         Body: { empresa_id, cargo_id, tipo, examenes_ids: [1,2,3] }
    DELETE: Elimina exámenes de un cargo en una empresa para un tipo.
         Body: { empresa_id, cargo_id, tipo, examenes_ids: [1,2,3] }
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperAdmin]

    def get(self, request):
        empresa_id = request.query_params.get('empresa_id')
        cargo_id = request.query_params.get('cargo_id')
        tipo = request.query_params.get('tipo', '').upper()

        if not empresa_id:
            return Response(
                {"error": "Se requiere empresa_id"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            empresa = Epresa.objects.get(idempresa=empresa_id)
        except Epresa.DoesNotExist:
            return Response({"error": "Empresa no encontrada"}, status=status.HTTP_404_NOT_FOUND)

        # Si no se proporciona cargo_id, devolver solo los cargos asignados a esa empresa
        if not cargo_id:
            # Obtener cargos únicos que tienen asignaciones en esa empresa
            cargos_asignados = Cargo.objects.filter(
                examenescargo__empresa=empresa
            ).distinct().values('idcargo', 'nombrecargo').order_by('nombrecargo')
            
            return Response({
                'empresa': {'id': empresa.idempresa, 'nombre': empresa.nombre_empresa},
                'cargos_disponibles': list(cargos_asignados),
            }, status=status.HTTP_200_OK)

        try:
            cargo = Cargo.objects.get(idcargo=cargo_id)
        except Cargo.DoesNotExist:
            return Response({"error": "Cargo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        # Filtrar asignaciones
        queryset = ExamenesCargo.objects.filter(
            empresa=empresa,
            cargo=cargo,
            examen__activo=True
        ).select_related('examen')

        if tipo and tipo in TIPOS_EXAMEN_VALIDOS:
            queryset = queryset.filter(tipo=tipo)

        # Agrupar por tipo
        asignaciones_por_tipo = {}
        for ec in queryset.order_by('tipo', 'examen__nombre'):
            if ec.tipo not in asignaciones_por_tipo:
                asignaciones_por_tipo[ec.tipo] = []
            asignaciones_por_tipo[ec.tipo].append({
                'id_asignacion': ec.id,
                'id_examen': ec.examen.id_examen,
                'nombre_examen': ec.examen.nombre,
            })

        # Todos los exámenes activos disponibles (para el selector)
        todos_examenes = list(
            Examen.objects.filter(activo=True)
            .order_by('nombre')
            .values('id_examen', 'nombre')
        )

        return Response({
            'empresa': {'id': empresa.idempresa, 'nombre': empresa.nombre_empresa},
            'cargo': {'id': cargo.idcargo, 'nombre': cargo.nombrecargo},
            'asignaciones_por_tipo': asignaciones_por_tipo,
            'todos_examenes': todos_examenes,
            'tipos_validos': TIPOS_EXAMEN_VALIDOS,
        }, status=status.HTTP_200_OK)

    def post(self, request):
        """Agregar exámenes a un cargo en una empresa para un tipo."""
        empresa_id = request.data.get('empresa_id')
        cargo_id = request.data.get('cargo_id')
        tipo = (request.data.get('tipo') or '').upper()
        examenes_ids = request.data.get('examenes_ids', [])

        if not empresa_id or not cargo_id or not tipo or not examenes_ids:
            return Response(
                {"error": "Se requieren empresa_id, cargo_id, tipo y examenes_ids"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if tipo not in TIPOS_EXAMEN_VALIDOS:
            return Response(
                {"error": f"Tipo inválido. Debe ser: {', '.join(TIPOS_EXAMEN_VALIDOS)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            empresa = Epresa.objects.get(idempresa=empresa_id)
            cargo = Cargo.objects.get(idcargo=cargo_id)
        except Epresa.DoesNotExist:
            return Response({"error": "Empresa no encontrada"}, status=status.HTTP_404_NOT_FOUND)
        except Cargo.DoesNotExist:
            return Response({"error": "Cargo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        examenes = Examen.objects.filter(id_examen__in=examenes_ids, activo=True)
        if not examenes.exists():
            return Response(
                {"error": "No se encontraron exámenes válidos con los IDs proporcionados"},
                status=status.HTTP_400_BAD_REQUEST
            )

        creados = []
        ya_existentes = []
        for examen in examenes:
            obj, created = ExamenesCargo.objects.get_or_create(
                empresa=empresa,
                cargo=cargo,
                examen=examen,
                tipo=tipo
            )
            if created:
                creados.append(examen.nombre)
            else:
                ya_existentes.append(examen.nombre)

        # Invalidar cache
        cache.delete('cargo_empresa_examenes_data')

        return Response({
            'mensaje': f"Se agregaron {len(creados)} exámenes al cargo '{cargo.nombrecargo}' "
                       f"en empresa '{empresa.nombre_empresa}' para tipo '{tipo}'",
            'creados': creados,
            'ya_existentes': ya_existentes,
        }, status=status.HTTP_201_CREATED)

    def delete(self, request):
        """Eliminar exámenes de un cargo en una empresa para un tipo."""
        empresa_id = request.data.get('empresa_id')
        cargo_id = request.data.get('cargo_id')
        tipo = (request.data.get('tipo') or '').upper()
        examenes_ids = request.data.get('examenes_ids', [])

        if not empresa_id or not cargo_id or not tipo or not examenes_ids:
            return Response(
                {"error": "Se requieren empresa_id, cargo_id, tipo y examenes_ids"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if tipo not in TIPOS_EXAMEN_VALIDOS:
            return Response(
                {"error": f"Tipo inválido. Debe ser: {', '.join(TIPOS_EXAMEN_VALIDOS)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            empresa = Epresa.objects.get(idempresa=empresa_id)
            cargo = Cargo.objects.get(idcargo=cargo_id)
        except Epresa.DoesNotExist:
            return Response({"error": "Empresa no encontrada"}, status=status.HTTP_404_NOT_FOUND)
        except Cargo.DoesNotExist:
            return Response({"error": "Cargo no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        eliminados = ExamenesCargo.objects.filter(
            empresa=empresa,
            cargo=cargo,
            tipo=tipo,
            examen_id__in=examenes_ids
        )
        count = eliminados.count()
        nombres_eliminados = list(eliminados.values_list('examen__nombre', flat=True))
        eliminados.delete()

        # Invalidar cache
        cache.delete('cargo_empresa_examenes_data')

        return Response({
            'mensaje': f"Se eliminaron {count} exámenes del cargo '{cargo.nombrecargo}' "
                       f"en empresa '{empresa.nombre_empresa}' para tipo '{tipo}'",
            'eliminados': nombres_eliminados,
        }, status=status.HTTP_200_OK)