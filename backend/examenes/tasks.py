"""
Celery tasks para el módulo de exámenes médicos.

Tareas:
- verificar_conexion_smtp: Verifica conexión al servidor SMTP (capacitaciones)
- verificar_conexion_smtp_medico: Verifica conexión al servidor SMTP médico
- enviar_correo_examen_task: Reenvía un correo individual
- enviar_correo_masivo_task: Reenvía un correo masivo con Excel adjunto
"""
from __future__ import absolute_import, unicode_literals
import logging
import smtplib
import ssl
from io import BytesIO

from celery import shared_task
from django.conf import settings
from django.core.mail import EmailMultiAlternatives

logger = logging.getLogger(__name__)


def verificar_conexion_smtp():
    """
    Verifica que el servidor SMTP de CAPACITACIONES esté disponible y acepte conexiones.
    
    Returns:
        tuple: (bool, str|None) — (True, None) si OK, (False, error_msg) si falla
    """
    try:
        timeout = getattr(settings, 'EMAIL_TIMEOUT', 30)
        
        if settings.EMAIL_USE_SSL:
            context = ssl.create_default_context()
            server = smtplib.SMTP_SSL(
                settings.EMAIL_HOST,
                settings.EMAIL_PORT,
                timeout=timeout,
                context=context
            )
        else:
            server = smtplib.SMTP(
                settings.EMAIL_HOST,
                settings.EMAIL_PORT,
                timeout=timeout
            )
            if settings.EMAIL_USE_TLS:
                server.starttls()
        
        server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
        server.quit()
        
        logger.info("Verificación SMTP (capacitaciones) exitosa")
        return True, None
        
    except smtplib.SMTPAuthenticationError as e:
        error = f"Error de autenticación SMTP (capacitaciones): {str(e)}"
        logger.error(error)
        return False, error
    except (smtplib.SMTPConnectError, ConnectionRefusedError) as e:
        error = f"No se pudo conectar al servidor SMTP (capacitaciones): {str(e)}"
        logger.error(error)
        return False, error
    except TimeoutError as e:
        error = f"Timeout conectando al servidor SMTP (capacitaciones): {str(e)}"
        logger.error(error)
        return False, error
    except Exception as e:
        error = f"Error verificando SMTP (capacitaciones): {type(e).__name__}: {str(e)}"
        logger.error(error)
        return False, error


def verificar_conexion_smtp_medico():
    """
    Verifica que el servidor SMTP de EXÁMENES MÉDICOS (360 CloudRegency) 
    esté disponible y acepte conexiones.
    
    Returns:
        tuple: (bool, str|None) — (True, None) si OK, (False, error_msg) si falla
    """
    try:
        timeout = getattr(settings, 'EMAIL_MEDICAL_TIMEOUT', 30)
        
        if settings.EMAIL_MEDICAL_USE_SSL:
            context = ssl.create_default_context()
            server = smtplib.SMTP_SSL(
                settings.EMAIL_MEDICAL_HOST,
                settings.EMAIL_MEDICAL_PORT,
                timeout=timeout,
                context=context
            )
        else:
            server = smtplib.SMTP(
                settings.EMAIL_MEDICAL_HOST,
                settings.EMAIL_MEDICAL_PORT,
                timeout=timeout
            )
            if settings.EMAIL_MEDICAL_USE_TLS:
                server.starttls()
        
        server.login(settings.EMAIL_MEDICAL_HOST_USER, settings.EMAIL_MEDICAL_HOST_PASSWORD)
        server.quit()
        
        logger.info("Verificación SMTP (exámenes médicos) exitosa")
        return True, None
        
    except smtplib.SMTPAuthenticationError as e:
        error = f"Error de autenticación SMTP (exámenes médicos): {str(e)}"
        logger.error(error)
        return False, error
    except (smtplib.SMTPConnectError, ConnectionRefusedError) as e:
        error = f"No se pudo conectar al servidor SMTP (exámenes médicos): {str(e)}"
        logger.error(error)
        return False, error
    except TimeoutError as e:
        error = f"Timeout conectando al servidor SMTP (exámenes médicos): {str(e)}"
        logger.error(error)
        return False, error
    except Exception as e:
        error = f"Error verificando SMTP (exámenes médicos): {type(e).__name__}: {str(e)}"
        logger.error(error)
        return False, error


@shared_task(
    bind=True,
    name='examenes.tasks.enviar_correo_examen_task',
    max_retries=3,
    default_retry_delay=60,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=300,
    acks_late=True,
)
def enviar_correo_examen_task(self, correo_id):
    """
    Tarea Celery para (re)enviar un correo de exámenes.
    Puede ser individual (1 trabajador sin Excel) o con múltiples trabajadores (con Excel).
    
    Reconstruye el correo a partir del registro CorreoExamenEnviado en BD,
    que ya tiene asunto, cuerpo y destinatarios almacenados.
    
    Si hay >1 trabajador, regenera automáticamente el Excel.
    
    Args:
        correo_id: ID del registro CorreoExamenEnviado
    """
    from examenes.models import CorreoExamenEnviado
    
    logger.info(f"[TASK] Iniciando envío de correo id={correo_id}")
    
    try:
        correo_obj = CorreoExamenEnviado.objects.get(id=correo_id)
    except CorreoExamenEnviado.DoesNotExist:
        logger.error(f"[TASK] Correo id={correo_id} no encontrado en BD")
        return {'status': 'error', 'mensaje': f'Correo id={correo_id} no encontrado'}
    
    # Obtener destinatarios y limpiar agresivamente
    destinatarios = [
        e.strip().replace('\r', '').replace('\n', '').replace('\t', '')
        for e in correo_obj.correos_destino.split(',')
        if e and e.strip()
    ]
    
    if not destinatarios:
        error_msg = "Sin destinatarios válidos"
        correo_obj.error_envio = error_msg
        correo_obj.save(update_fields=['error_envio'])
        logger.error(f"[TASK] Correo id={correo_id}: {error_msg}")
        return {'status': 'error', 'mensaje': error_msg}
    
    # Verificar SMTP antes de enviar
    smtp_ok, smtp_error = verificar_conexion_smtp()
    if not smtp_ok:
        logger.warning(f"[TASK] SMTP no disponible para correo id={correo_id}: {smtp_error}")
        correo_obj.error_envio = f"SMTP no disponible: {smtp_error}"
        correo_obj.save(update_fields=['error_envio'])
        # Celery reintentará automáticamente por autoretry_for
        raise Exception(f"SMTP no disponible: {smtp_error}")
    
    try:
        # Determinar si necesita Excel (>1 trabajador)
        cantidad_trabajadores = correo_obj.trabajadores.count()
        excel_bytes = None
        if cantidad_trabajadores > 1:
            logger.info(f"[TASK] Generando Excel para correo id={correo_id} ({cantidad_trabajadores} trabajadores)")
            excel_bytes = _regenerar_excel_desde_bd(correo_obj)
        
        # Enviar correo con los datos almacenados en BD
        email = EmailMultiAlternatives(
            subject=correo_obj.asunto,
            body=correo_obj.cuerpo_correo,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=destinatarios
        )
        
        # Si el cuerpo tiene HTML, adjuntar como alternativa
        if '<html' in correo_obj.cuerpo_correo.lower() or '<p>' in correo_obj.cuerpo_correo.lower():
            email.attach_alternative(correo_obj.cuerpo_correo, "text/html")
        
        # Adjuntar Excel si existe
        if excel_bytes:
            email.attach(
                'Trabajadores_Examenes.xlsx',
                excel_bytes,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        email.send(fail_silently=False)
        
        # Marcar como enviado
        correo_obj.enviado_correctamente = True
        correo_obj.error_envio = None
        correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])
        
        logger.info(
            f"[TASK] Correo id={correo_id} enviado exitosamente "
            f"a {len(destinatarios)} destinatarios (trabajadores: {cantidad_trabajadores})"
        )
        
        # Limpiar cache
        _limpiar_cache_reportes()
        
        return {
            'status': 'ok',
            'correo_id': correo_id,
            'uuid_correo': correo_obj.uuid_correo,
            'destinatarios': len(destinatarios),
            'trabajadores': cantidad_trabajadores,
            'con_excel': excel_bytes is not None
        }
        
    except Exception as e:
        error_msg = f"{type(e).__name__}: {str(e)}"
        correo_obj.enviado_correctamente = False
        correo_obj.error_envio = error_msg
        correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])
        
        logger.error(f"[TASK] Error enviando correo id={correo_id}: {error_msg}")
        raise  # Celery reintentará automáticamente


@shared_task(
    bind=True,
    name='examenes.tasks.enviar_correo_masivo_task',
    max_retries=3,
    default_retry_delay=120,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=600,
    acks_late=True,
)
def enviar_correo_masivo_task(self, correo_id, excel_bytes=None):
    """
    Tarea Celery para (re)enviar un correo masivo de exámenes con Excel adjunto.
    
    Si excel_bytes es None (reintento manual), regenera el Excel
    a partir de los registros de trabajadores en BD.
    
    Args:
        correo_id: ID del registro CorreoExamenEnviado
        excel_bytes: bytes del Excel a adjuntar (opcional, se regenera si no viene)
    """
    from examenes.models import CorreoExamenEnviado, RegistroExamenes, ExamenTrabajador
    
    logger.info(f"[TASK] Iniciando envío de correo masivo id={correo_id}")
    
    try:
        correo_obj = CorreoExamenEnviado.objects.get(id=correo_id)
    except CorreoExamenEnviado.DoesNotExist:
        logger.error(f"[TASK] Correo masivo id={correo_id} no encontrado en BD")
        return {'status': 'error', 'mensaje': f'Correo id={correo_id} no encontrado'}
    
    # Obtener destinatarios y limpiar agresivamente
    destinatarios = [
        e.strip().replace('\r', '').replace('\n', '').replace('\t', '')
        for e in correo_obj.correos_destino.split(',')
        if e and e.strip()
    ]
    
    if not destinatarios:
        error_msg = "Sin destinatarios válidos"
        correo_obj.error_envio = error_msg
        correo_obj.save(update_fields=['error_envio'])
        logger.error(f"[TASK] Correo masivo id={correo_id}: {error_msg}")
        return {'status': 'error', 'mensaje': error_msg}
    
    # Verificar SMTP antes de enviar
    smtp_ok, smtp_error = verificar_conexion_smtp()
    if not smtp_ok:
        logger.warning(f"[TASK] SMTP no disponible para correo masivo id={correo_id}: {smtp_error}")
        correo_obj.error_envio = f"SMTP no disponible: {smtp_error}"
        correo_obj.save(update_fields=['error_envio'])
        raise Exception(f"SMTP no disponible: {smtp_error}")
    
    try:
        # Si no hay Excel, regenerarlo desde BD
        if excel_bytes is None:
            logger.info(f"[TASK] Regenerando Excel para correo masivo id={correo_id}")
            excel_bytes = _regenerar_excel_desde_bd(correo_obj)
        elif isinstance(excel_bytes, (list, memoryview)):
            excel_bytes = bytes(excel_bytes)
        
        # Enviar correo
        email = EmailMultiAlternatives(
            subject=correo_obj.asunto,
            body='Por favor, abra este correo en un cliente que soporte HTML.',
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=destinatarios
        )
        email.attach_alternative(correo_obj.cuerpo_correo, "text/html")
        
        # Adjuntar Excel si existe
        if excel_bytes:
            email.attach(
                'Trabajadores_Examenes.xlsx',
                excel_bytes,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        email.send(fail_silently=False)
        
        # Marcar como enviado
        correo_obj.enviado_correctamente = True
        correo_obj.error_envio = None
        correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])
        
        logger.info(
            f"[TASK] Correo masivo id={correo_id} enviado exitosamente "
            f"a {len(destinatarios)} destinatarios"
        )
        
        # Limpiar cache
        _limpiar_cache_reportes()
        
        return {
            'status': 'ok',
            'correo_id': correo_id,
            'uuid_correo': correo_obj.uuid_correo,
            'destinatarios': len(destinatarios),
            'con_excel': excel_bytes is not None
        }
        
    except Exception as e:
        error_msg = f"{type(e).__name__}: {str(e)}"
        correo_obj.enviado_correctamente = False
        correo_obj.error_envio = error_msg
        correo_obj.save(update_fields=['enviado_correctamente', 'error_envio'])
        
        logger.error(f"[TASK] Error enviando correo masivo id={correo_id}: {error_msg}")
        raise


def _regenerar_excel_desde_bd(correo_obj):
    """
    Regenera el archivo Excel a partir de los registros de trabajadores
    almacenados en BD para un correo masivo.
    
    Args:
        correo_obj: instancia de CorreoExamenEnviado
        
    Returns:
        bytes: contenido del Excel generado, o None si no hay trabajadores
    """
    from examenes.models import RegistroExamenes, ExamenTrabajador
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    registros = RegistroExamenes.objects.filter(
        correo_lote=correo_obj
    ).select_related('empresa', 'cargo', 'centro').prefetch_related('examenes__examen')
    
    if not registros.exists():
        logger.warning(f"No hay trabajadores para correo id={correo_obj.id}")
        return None
    
    # Recopilar todos los exámenes únicos
    examenes_unicos = set()
    trabajadores_data = []
    
    for reg in registros:
        examenes_del_trabajador = []
        for et in reg.examenes.all():
            examenes_del_trabajador.append(et.examen.nombre)
            examenes_unicos.add(et.examen.nombre)
        
        trabajadores_data.append({
            'uuid': reg.uuid_trabajador,
            'nombre': reg.nombre_trabajador,
            'documento': reg.documento_trabajador,
            'empresa': reg.empresa.nombre_empresa if reg.empresa else '',
            'cargo': reg.cargo.nombrecargo if reg.cargo else '',
            'centro': reg.centro.nombrecentrop if reg.centro else '',
            'ciudad': reg.ciudad or '',
            'tipo_examen': reg.tipo_examen,
            'examenes': examenes_del_trabajador,
        })
    
    # Crear Excel
    nombres_examenes = sorted(list(examenes_unicos))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Trabajadores Examenes"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ["UUID", "Empresa", "Centro", "Cargo", "Ciudad", "Nombre",
               "Documento", "Tipo Examen"] + nombres_examenes
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border_style
    
    # Datos
    for trab in trabajadores_data:
        row = [
            trab['uuid'], trab['empresa'], trab['centro'],
            trab['cargo'], trab['ciudad'], trab['nombre'],
            trab['documento'], trab['tipo_examen']
        ]
        # Agregar marcas de exámenes
        for examen_nombre in nombres_examenes:
            row.append("X" if examen_nombre in trab['examenes'] else "")
        ws.append(row)
    
    # Ajustar anchos de columna
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = 18
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    logger.info(
        f"Excel regenerado: {len(trabajadores_data)} trabajadores, "
        f"{len(nombres_examenes)} exámenes"
    )
    
    return buffer.getvalue()


def _limpiar_cache_reportes():
    """Limpia el cache de reportes de correos."""
    try:
        from django.core.cache import cache
        cache.delete_pattern("reporte_correos_page=*")
        cache.delete_pattern("detalle_correo=*")
        cache.delete('cargo_empresa_examenes_data')
        logger.info("[TASK] Cache de reportes limpiado")
    except Exception as e:
        logger.warning(f"[TASK] Error limpiando cache: {str(e)}")
