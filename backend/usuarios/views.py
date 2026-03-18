import json
import csv
import io
import os
from io import BytesIO
from django.http import JsonResponse, FileResponse
from django.db import transaction
from rest_framework.response import Response
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from usuarios.permissions import IsSuperAdmin, IsAdminUser, IsUsuarioEspecial, IsSuperUserOrAdmin
from usuarios.models import Colaboradores, Usuarios, Cargo, Niveles, Regional
from analitica.models import Centroop, Proyecto, Unidadnegocio, Epresa
from capacitaciones.models import Capacitaciones, progresoCapacitaciones, Modulos, Lecciones, progresolecciones
from capacitaciones.serializers import CapacitacionProgresoSerializer
from usuarios.serializers import ColaboradorListadoSerializer, cargosSerializer, nivelesSerializer, regionalesSerializer
from django.db.models import Count, Q, Prefetch, OuterRef, Subquery, IntegerField, Case, When, Max
from django.db.models.functions import Coalesce
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


class Perfil(APIView):
    """
    Vista de perfil del colaborador con sus capacitaciones.
    
    Optimización:
    - Usa annotate() para calcular totales de lecciones en DB
    - Elimina queries N+1 en el loop de capacitaciones
    - Precarga relaciones con select_related
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, id=None):
        # Si viene un id en la ruta, usar ese colaborador; de lo contrario, el del token
        if id is not None:
            colaborador = Colaboradores.objects.select_related(
                'centroop__id_proyecto__id_unidad__id_empresa',
                'nivelcolaborador',
                'regionalcolab',
                'cargocolaborador'
            ).filter(idcolaborador=id).first()
        else:
            colaborador = request.user.idcolaboradoru
            if colaborador:
                # Recargar con relaciones
                colaborador = Colaboradores.objects.select_related(
                    'centroop__id_proyecto__id_unidad__id_empresa',
                    'nivelcolaborador',
                    'regionalcolab',
                    'cargocolaborador'
                ).filter(idcolaborador=colaborador.idcolaborador).first()

        if not colaborador:
            return Response(
                {"error": "El usuario no tiene colaborador asociado"},
                status=400
            )

        # Subquery para contar lecciones completadas por capacitación
        # Contar lecciones donde el progreso está completado (completada=1)
        lecciones_completadas_subq = progresolecciones.objects.filter(
            idcolaborador=colaborador,
            idleccion__idmodulo__idcapacitacion=OuterRef('capacitacion_id'),
            completada=1  # Usar completada=1 como indicador de completación
        ).values('idleccion__idmodulo__idcapacitacion').annotate(
            count=Count('id_progreso', distinct=True)
        ).values('count')

        # FIX: Eliminar duplicados si un colaborador está múltiples veces en la misma capacitación
        # Usar Max('id') + GROUP BY capacitacion (compatible con MySQL)
        latest_ids = (
            progresoCapacitaciones.objects
            .filter(colaborador=colaborador)
            .exclude(capacitacion__estado=3)
            .values('capacitacion')
            .annotate(max_id=Max('id'))
            .values_list('max_id', flat=True)
        )
        
        # Filtrar solo los registros de progreso más recientes
        progresos = (
            progresoCapacitaciones.objects
            .filter(id__in=list(latest_ids))
            .select_related('capacitacion')
            .annotate(
                total_lecciones=Count(
                    'capacitacion__modulos__lecciones',
                    distinct=True
                ),
                lecciones_completadas=Coalesce(
                    Subquery(lecciones_completadas_subq, output_field=IntegerField()),
                    0
                ),
                completada_orden=Case(
                    When(completada=0, then=0),  # Incompletas primero
                    When(completada=1, then=1),  # Completadas después
                    output_field=IntegerField()
                )
            )
            .order_by('completada_orden', '-fecha_registro')
        )

        capacitaciones_totales = progresos.count()
        capacitaciones_completadas = progresos.filter(completada=1).count()

        # Construir datos de capacitaciones sin queries adicionales
        capacitaciones_data = [
            {
                "id_capacitacion": prog.capacitacion.id,
                "nombre_capacitacion": prog.capacitacion.titulo,
                "completada": bool(prog.completada),
                "progreso": float(prog.progreso) if prog.progreso is not None else 0.0,
                "lecciones_completadas": prog.lecciones_completadas,
                "estado_capacitacion": prog.capacitacion.estado,
                "total_lecciones": prog.total_lecciones,
                "fecha_completacion": prog.fecha_completada.isoformat() if getattr(prog, 'fecha_completada', None) else None
            }
            for prog in progresos
        ]

        # Acceder a relaciones precargadas
        centro = getattr(colaborador, 'centroop', None)
        proyecto = getattr(centro, 'id_proyecto', None) if centro else None
        unidad = getattr(proyecto, 'id_unidad', None) if proyecto else None
        empresa = getattr(unidad, 'id_empresa', None) if unidad else None

        data = {
            "id_colaborador": colaborador.idcolaborador,
            "nombre_colaborador": colaborador.nombrecolaborador,
            "apellido_colaborador": colaborador.apellidocolaborador,
            "correo_colaborador": colaborador.correocolaborador,
            "telefo_colaborador": colaborador.telefocolaborador,

            "nombre_centroOP": getattr(centro, 'nombrecentrop', None),
            "nombre_empresa": getattr(empresa, 'nombre_empresa', None),
            "nombre_nivel": getattr(colaborador.nivelcolaborador, 'nombrenivel', None) if colaborador.nivelcolaborador_id else None,
            "nombre_regional": getattr(colaborador.regionalcolab, 'nombreregional', None) if colaborador.regionalcolab_id else None,
            "nombre_cargo": getattr(colaborador.cargocolaborador, 'nombrecargo', None) if colaborador.cargocolaborador_id else None,
            "nombre_proyecto": getattr(proyecto, 'nombreproyecto', None) if proyecto else None,
            "nombre_unidad": getattr(unidad, 'nombreunidad', None) if unidad else None,

            "capacitaciones_totales": capacitaciones_totales,
            "capacitaciones_completadas": capacitaciones_completadas,
            "capacitaciones": capacitaciones_data
        }

        return Response(data)
    
    def patch(self, request, id=None):
        """
        Alterna el estado del colaborador (0 <-> 1). Requiere id de colaborador.
        """
        if id is None:
            return Response({"error": "Se requiere el id del colaborador"}, status=400)

        colaborador = Colaboradores.objects.filter(idcolaborador=id).first()
        if not colaborador:
            return Response({"error": "Colaborador no encontrado"}, status=404)

        # Alternar estado: si es 1 pasa a 0, si es 0 pasa a 1
        if hasattr(colaborador, 'estadocolaborador'):
            nuevo_estado = 0 if colaborador.estadocolaborador == 1 else 1
            colaborador.estadocolaborador = nuevo_estado
            colaborador.save()

            # Cambiar también el estado del usuario relacionado si existe
            usuario = Usuarios.objects.filter(idcolaboradoru=colaborador).first()
            if usuario:
                if nuevo_estado == 0:
                    usuario.estadousuario = 0
                    usuario.save()
                # Si quieres reactivar el usuario cuando el colaborador se activa, descomenta:
                # elif nuevo_estado == 1:
                #     usuario.estadousuario = 1
                #     usuario.save()

            return Response({
                "id_colaborador": colaborador.idcolaborador,
                "nuevo_estado_colaborador": colaborador.estadocolaborador,
                "nuevo_estado_usuario": usuario.estadousuario if usuario else None
            })
        else:
            return Response({"error": "El colaborador no tiene campo 'estadocolaborador'"}, status=400)



class Register(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdmin, IsAdminUser]

    def post(self, request, *args, **kwargs):

        payload = request.data if hasattr(request, 'data') else None
        if not payload:
            try:
                payload = json.loads(request.body.decode('utf-8'))
            except Exception:
                return JsonResponse({'error': 'JSON inválido'}, status=400)

        # Validar campos raíz requeridos
        required_root = ['usuario', 'password', 'idcolaborador']
        if any(key not in payload for key in required_root):
            return JsonResponse({'error': 'Faltan campos requeridos'}, status=400)

        colab_data = payload.get('idcolaborador') or {}
        required_colab = [
            'cc_colaborador', 'nombre_colaborador', 'apellido_colaborador',
            'cargo_colaborador', 'correo_colaborador', 'nivel_colaborador',
            'regional_colab', 'centroOP'
        ]
        if any(key not in colab_data for key in required_colab):
            return JsonResponse({'error': 'Faltan datos del colaborador'}, status=400)

        # VALIDACIONES TEMPRANAS - Verificar ANTES DE TODO
        # 1. Validar que el usuario NO exista (por nombre de usuario)
        usuario_nombre = payload.get('usuario', '').strip()
        if not usuario_nombre:
            return JsonResponse({'error': 'El usuario no puede estar vacío'}, status=400)
        if Usuarios.objects.filter(usuario=usuario_nombre).exists():
            return JsonResponse({'error': f'El usuario "{usuario_nombre}" ya existe en la base de datos'}, status=400)

        # 2. Validar que la cédula NO exista (validar colaborador por cédula)
        cc_colaborador = colab_data.get('cc_colaborador', '').strip()
        if not cc_colaborador:
            return JsonResponse({'error': 'La cédula del colaborador no puede estar vacía'}, status=400)
        if Colaboradores.objects.filter(cccolaborador=cc_colaborador).exists():
            return JsonResponse({'error': f'El colaborador con cédula "{cc_colaborador}" ya existe en la base de datos'}, status=400)

        try:
            # CREAR NUEVO COLABORADOR
            colaborador = Colaboradores.objects.create(
                cccolaborador=colab_data['cc_colaborador'],
                nombrecolaborador=colab_data['nombre_colaborador'],
                apellidocolaborador=colab_data['apellido_colaborador'],
                cargocolaborador_id=colab_data['cargo_colaborador'],
                correocolaborador=colab_data.get('correo_colaborador', ''),
                telefocolaborador=colab_data.get('telefo_colaborador', ''),
                nivelcolaborador_id=colab_data['nivel_colaborador'],
                regionalcolab_id=colab_data['regional_colab'],
                centroop_id=colab_data['centroOP'],
            )

            # CREAR NUEVO USUARIO
            user = Usuarios(
                usuario=payload['usuario'],
                tipousuario=int(payload.get('is_staff', 0)),
                idcolaboradoru=colaborador,
                estadousuario=1,
            )
            user.set_password(payload['password'])
            user.save()

            return JsonResponse({
                'mensaje': 'Usuario y colaborador creados correctamente',
                'usuario_id': user.id,
                'colaborador_id': colaborador.idcolaborador,
            }, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        

    def get(self, request, colaborador_id):
        colaborador = (
            Colaboradores.objects
            .select_related(
                'cargocolaborador',
                'nivelcolaborador',
                'regionalcolab',
                'centroop__id_proyecto__id_unidad__id_empresa'
            )
            .filter(idcolaborador=colaborador_id)
            .first()
        )
        if not colaborador:
            return Response({"error": "Colaborador no encontrado"}, status=404)

        cargo = colaborador.cargocolaborador
        nivel = colaborador.nivelcolaborador
        region = colaborador.regionalcolab
        centroop = colaborador.centroop
        proyecto = centroop.id_proyecto if centroop else None
        unidad = proyecto.id_unidad if proyecto else None
        empresa = unidad.id_empresa if unidad else None

        data = {
            "idcolaborador": colaborador.idcolaborador,
            "nombre": colaborador.nombrecolaborador,
            "apellido": colaborador.apellidocolaborador,
            "correo": colaborador.correocolaborador,
            "telefono": colaborador.telefocolaborador,
            "cargo": cargo.idcargo if cargo else None,
            "nivel": nivel.idnivel if nivel else None,
            "region": region.idregional if region else None,
            "centroop_id": centroop.idcentrop if centroop else None,
            "centroop": centroop.nombrecentrop if centroop else None,
        }
        return Response(data)
    
    def put(self, request, colaborador_id):
        colaborador = Colaboradores.objects.filter(idcolaborador=colaborador_id).first()
        if not colaborador:
            return Response({"error": "Colaborador no encontrado"}, status=404)

        data = request.data
        # Mapear campos del frontend a los del modelo
        mapeo = {
            'nombre': 'nombrecolaborador',
            'apellido': 'apellidocolaborador',
            'correo': 'correocolaborador',
            'telefono': 'telefocolaborador',
            'cargo': 'cargocolaborador_id',
            'nivel': 'nivelcolaborador_id',
            'region': 'regionalcolab_id',
            'centroop_id': 'centroop_id',
        }
        for campo_front, campo_modelo in mapeo.items():
            if campo_front in data:
                setattr(colaborador, campo_modelo, data[campo_front])

        # Si el frontend envía estadocolaborador, también actualizarlo
        if 'estadocolaborador' in data:
            colaborador.estadocolaborador = data['estadocolaborador']

        colaborador.save()
        serializer = ColaboradorListadoSerializer(colaborador)
        return Response(serializer.data)
    

class RegisterTemporal(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdmin, IsUsuarioEspecial]

    def post(self, request, *args, **kwargs):
        payload = request.data if hasattr(request, 'data') else None
        # Normalizar payload: si viene vacío, intentar parsear body
        if not payload:
            try:
                payload = json.loads(request.body.decode('utf-8'))
            except Exception:
                return JsonResponse({'error': 'JSON inválido'}, status=400)

        # 'is_staff' no es requerido para RegisterTemporal; se ignorará y se usará 0
        required_root = ['usuario', 'password', 'idcolaborador']

        # Si `idcolaborador` llegó como JSON string (p.e. multipart/form-data), intentar parsearlo
        if isinstance(payload.get('idcolaborador'), str):
            try:
                payload['idcolaborador'] = json.loads(payload['idcolaborador'])
            except Exception:
                pass

        if any(key not in payload for key in required_root):
            try:
                print('RegisterTemporal - payload keys:', list(payload.keys()))
            except Exception:
                pass
            return JsonResponse({'error': 'Faltan campos requeridos'}, status=400)

        # VALIDACIÓN TEMPRANA: Verificar usuario antes de cualquier otra validación
        usuario_nombre = payload.get('usuario', '').strip()
        if not usuario_nombre:
            return JsonResponse({'error': 'El usuario no puede estar vacío'}, status=400)
        if Usuarios.objects.filter(usuario=usuario_nombre).exists():
            return JsonResponse({'error': f'El usuario {usuario_nombre} ya existe en la base de datos'}, status=400)

        colab_data = payload.get('idcolaborador') or {}
        required_colab_min = [
            'cc_colaborador', 'nombre_colaborador', 'apellido_colaborador'
        ]
        if any(key not in colab_data for key in required_colab_min):
            return JsonResponse({'error': 'Faltan datos mínimos del colaborador'}, status=400)

        # Validar cédula
        cc_colaborador = colab_data.get('cc_colaborador', '').strip()
        if not cc_colaborador:
            return JsonResponse({'error': 'La cédula del colaborador no puede estar vacía'}, status=400)
        if Colaboradores.objects.filter(cccolaborador=cc_colaborador).exists():
            return JsonResponse({'error': f'La cédula {cc_colaborador} ya existe en la base de datos'}, status=400)

        try:
            colaborador = Colaboradores.objects.create(
                cccolaborador=colab_data['cc_colaborador'],
                nombrecolaborador=colab_data['nombre_colaborador'],
                apellidocolaborador=colab_data['apellido_colaborador'],
                cargocolaborador_id= 118,
                correocolaborador=colab_data.get('correo_colaborador', ''),
                telefocolaborador=colab_data.get('telefo_colaborador', ''),
                nivelcolaborador_id=5,
                regionalcolab_id=1,
                centroop_id=1,
            )

            user = Usuarios(
                usuario=payload['usuario'],
                tipousuario=0,
                idcolaboradoru=colaborador,
                estadousuario=1,
            )
            user.set_password(payload['password'])
            user.save()

            # Si viene capacitacion_id, registrar el colaborador a esa capacitación
            capacitacion_id = payload.get('capacitacion_id')
            if capacitacion_id:
                try:
                    from capacitaciones.models import Capacitaciones, progresoCapacitaciones
                    capacitacion = Capacitaciones.objects.get(id=capacitacion_id)
                    progresoCapacitaciones.objects.get_or_create(
                        capacitacion_id=capacitacion_id,
                        colaborador_id=colaborador.idcolaborador,
                        defaults={'completada': 0, 'progreso': 0}
                    )
                except Capacitaciones.DoesNotExist:
                    return JsonResponse({'error': f'Capacitación {capacitacion_id} no encontrada'}, status=404)
                except Exception as e:
                    return JsonResponse({'error': f'Error registrando a capacitación: {str(e)}'}, status=500)

            return JsonResponse({
                'mensaje': 'Usuario temporal creado' + (' y registrado a capacitación' if capacitacion_id else ''),
                'usuario_id': user.id,
                'colaborador_id': colaborador.idcolaborador,
            }, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

class ListaUsuarios(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdmin, IsAdminUser]

    def get(self, request, *args, **kwargs):
        try:
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 10))
            if page < 1:
                page = 1
            if page_size < 1 or page_size > 100:
                page_size = 10

            search = request.GET.get('search', '').strip()

            base_qs = (
                Colaboradores.objects
                .filter(estadocolaborador=1)
                .select_related('cargocolaborador')
                .annotate(
                    total_capacitaciones=Count(
                        'progresocapacitaciones',
                        filter=Q(progresocapacitaciones__capacitacion__estado__in=[0, 1]),
                        distinct=True
                    ),
                    completadas=Count(
                        'progresocapacitaciones',
                        filter=Q(
                            progresocapacitaciones__capacitacion__estado__in=[0, 1],
                            progresocapacitaciones__completada=1
                        ),
                        distinct=True
                    ),
                )
                .prefetch_related(
                    Prefetch(
                        'progresocapacitaciones_set',
                        queryset=(
                            progresoCapacitaciones.objects
                            .filter(capacitacion__estado__in=[0, 1])
                            .select_related('capacitacion')
                        ),
                        to_attr='progresos_activos'
                    ),
                    Prefetch(
                        'progresocapacitaciones_set',
                        queryset=(
                            progresoCapacitaciones.objects
                            .filter(capacitacion__estado__in=[0, 1], completada=1)
                            .select_related('capacitacion')
                        ),
                        to_attr='progresos_activos_completados'
                    ),
                )
                .order_by('idcolaborador')
            )

            if search:
                base_qs = base_qs.filter(
                    Q(nombrecolaborador__icontains=search) |
                    Q(apellidocolaborador__icontains=search) |
                    Q(cccolaborador__icontains=search)
                )

            total = base_qs.count()
            start = (page - 1) * page_size
            end = start + page_size
            items = list(base_qs[start:end])

            results = ColaboradorListadoSerializer(items, many=True).data

            response = {
                'count': total,
                'page': page,
                'page_size': page_size,
                'results': results,
            }
            return Response(response)
        except Exception as e:
            return Response({'error': str(e)}, status=500)

class PerfilCapacitacionView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, colaborador_id=None, capacitacion_id=None):

        colaborador = Colaboradores.objects.filter(
            idcolaborador=colaborador_id
        ).first()

        if not colaborador:
            return Response({"error": "Colaborador no encontrado"}, status=404)

        capacitacion = Capacitaciones.objects.filter(pk=capacitacion_id).first()
        if not capacitacion:
            return Response({"error": "Capacitación no encontrada"}, status=404)

        # Serializar detalle completo de la capacitación con progreso y estructura
        serializer = CapacitacionProgresoSerializer(
            capacitacion,
            context={"colaborador": colaborador}
        )
        return Response(serializer.data)
    
class CargoNivelRegionalView(APIView):
    """
    Vista para obtener listas de Cargo, Niveles y Regionales.
    """
    permission_classes = [IsAuthenticated, IsUsuarioEspecial | IsSuperUserOrAdmin ]

    def get(self, request):

        cargos = Cargo.objects.filter(estadocargo=1).order_by('nombrecargo')
        niveles = Niveles.objects.filter(estadonivel=1).order_by('nombrenivel')
        regionales = Regional.objects.filter(estadoregional=1).order_by('nombreregional')

        cargos_data = cargosSerializer(cargos, many=True).data
        niveles_data = nivelesSerializer(niveles, many=True).data
        regionales_data = regionalesSerializer(regionales, many=True).data

        return Response({
            "cargos": cargos_data,
            "niveles": niveles_data,
            "regionales": regionales_data
        })

class FiltrarUsuariosView(APIView):
    """
    Vista para filtrar usuarios por nombre o CC.
    """
    permission_classes = [IsAuthenticated, IsSuperAdmin, IsAdminUser]

    def get(self, request):
        query = request.GET.get('q', '').strip()
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))
        if page < 1:
            page = 1
        if page_size < 1 or page_size > 100:
            page_size = 10

        base_qs = (
            Colaboradores.objects
            .exclude(estadocolaborador=3)
            .annotate(
                total_capacitaciones=Count(
                    'progresocapacitaciones',
                    filter=Q(progresocapacitaciones__capacitacion__estado__in=[0, 1]),
                    distinct=True
                ),
                completadas=Count(
                    'progresocapacitaciones',
                    filter=Q(
                        progresocapacitaciones__capacitacion__estado__in=[0, 1],
                        progresocapacitaciones__completada=1
                    ),
                    distinct=True
                ),
            )
            .order_by('idcolaborador')
        )
        if query:
            base_qs = base_qs.filter(
                Q(nombrecolaborador__icontains=query) |
                Q(apellidocolaborador__icontains=query) |
                Q(cccolaborador__icontains=query)
            )

        total = base_qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        items = list(base_qs[start:end])
        results = ColaboradorListadoSerializer(items, many=True).data

        response = {
            'count': total,
            'page': page,
            'page_size': page_size,
            'results': results,
        }
        return Response(response)


class CambiarEstadoUsuarioView(APIView):
    """
    Vista para activar o desactivar usuarios (uno o múltiples).
    Solo SuperAdmin puede realizar esta acción.
    
    Soporta dos formatos:
    1. PATCH individual: /usuarios/cambiar-estado-usuario/<colaborador_id>/
       Body: { "estado": 1 o 0 }
    
    2. POST masivo: /usuarios/cambiar-estado-usuario/
       Body: { "colaborador_ids": [1, 2, 3], "estado": 1 o 0 }
    """
    permission_classes = [IsAuthenticated, IsSuperAdmin]

    def patch(self, request, colaborador_id=None):
        """
        PATCH para desactivar/activar un usuario individual.
        PATCH /usuarios/cambiar-estado-usuario/<colaborador_id>/
        Body: { "estado": 1 o 0 }
        """
        try:
            if colaborador_id is None:
                return Response(
                    {"error": "colaborador_id es requerido en la URL"},
                    status=400
                )
            
            nuevo_estado = request.data.get('estado')
            
            if nuevo_estado is None:
                return Response(
                    {"error": "El campo 'estado' es requerido (0 o 1)"},
                    status=400
                )
            
            if nuevo_estado not in [0, 1]:
                return Response(
                    {"error": "El estado debe ser 0 (inactivo) o 1 (activo)"},
                    status=400
                )
            
            # Obtener el usuario por su colaborador
            usuario = Usuarios.objects.filter(
                idcolaboradoru__idcolaborador=colaborador_id
            ).first()
            
            if not usuario:
                return Response(
                    {"error": "Usuario no encontrado"},
                    status=404
                )
            
            usuario.estadousuario = nuevo_estado
            usuario.save()
            
            # Sincronizar estado en el modelo Colaboradores
            colaborador = getattr(usuario, 'idcolaboradoru', None)
            nuevo_estado_colaborador = None
            if colaborador:
                try:
                    colaborador.estadocolaborador = nuevo_estado
                    colaborador.save()
                    nuevo_estado_colaborador = colaborador.estadocolaborador
                except Exception:
                    nuevo_estado_colaborador = None
            
            return Response({
                "mensaje": "Estado del usuario actualizado correctamente",
                "usuario_id": usuario.id,
                "colaborador_id": usuario.idcolaboradoru.idcolaborador if getattr(usuario, 'idcolaboradoru', None) else None,
                "nuevo_estado_usuario": usuario.estadousuario,
                "nuevo_estado_colaborador": nuevo_estado_colaborador
            }, status=200)
            
        except Exception as e:
            return Response(
                {"error": f"Error al actualizar estado: {str(e)}"},
                status=500
            )

    def post(self, request):
        """
        POST para desactivar/activar múltiples usuarios.
        POST /usuarios/cambiar-estado-usuario/
        Body: { "colaborador_ids": [1, 2, 3, ...], "estado": 1 o 0 }
        o
        Body: { "cedulas": ["1234567", "7654321", ...], "estado": 1 o 0 }
        
        Retorna:
        {
            "mensaje": "...",
            "total": 3,
            "actualizados": 3,
            "no_encontrados": 0,
            "detalles_encontrados": [...],
            "detalles_no_encontrados": [
                {"identificador": "1234567", "tipo": "cedula", "error": "No encontrado"}
            ]
        }
        """
        try:
            colaborador_ids = request.data.get('colaborador_ids', [])
            cedulas = request.data.get('cedulas', [])
            nuevo_estado = request.data.get('estado')
            
            # Validaciones
            if not (isinstance(colaborador_ids, list) or isinstance(cedulas, list)) or (len(colaborador_ids) == 0 and len(cedulas) == 0):
                return Response(
                    {"error": "El campo 'colaborador_ids' o 'cedulas' es requerido y debe ser una lista no vacía"},
                    status=400
                )
            
            if nuevo_estado is None:
                return Response(
                    {"error": "El campo 'estado' es requerido (0 o 1)"},
                    status=400
                )
            
            if nuevo_estado not in [0, 1]:
                return Response(
                    {"error": "El estado debe ser 0 (inactivo) o 1 (activo)"},
                    status=400
                )
            
            # Procesar cada colaborador
            encontrados = []
            no_encontrados = []
            actualizados = 0
            
            # Procesar por ID de colaborador
            for colaborador_id in colaborador_ids:
                try:
                    # Obtener colaborador primero
                    colaborador = Colaboradores.objects.filter(
                        idcolaborador=colaborador_id
                    ).first()
                    
                    if not colaborador:
                        no_encontrados.append({
                            "identificador": colaborador_id,
                            "tipo": "colaborador_id",
                            "error": "Colaborador no encontrado"
                        })
                        continue
                    
                    # Obtener usuario del colaborador
                    usuario = Usuarios.objects.filter(idcolaboradoru=colaborador).first()
                    
                    if not usuario:
                        no_encontrados.append({
                            "identificador": colaborador_id,
                            "tipo": "colaborador_id",
                            "error": "Usuario no encontrado para este colaborador"
                        })
                        continue
                    
                    # Actualizar usuario
                    usuario.estadousuario = nuevo_estado
                    usuario.save()
                    
                    # Sincronizar estado en colaborador
                    colaborador.estadocolaborador = nuevo_estado
                    colaborador.save()
                    
                    # Construir nombre completo
                    nombre_completo = f"{colaborador.nombrecolaborador} {colaborador.apellidocolaborador}".strip()
                    
                    encontrados.append({
                        "colaborador_id": colaborador_id,
                        "usuario_id": usuario.id,
                        "nombre": nombre_completo,
                        "cedula": colaborador.cccolaborador,
                        "estado": nuevo_estado,
                        "success": True
                    })
                    actualizados += 1
                    
                except Exception as e:
                    no_encontrados.append({
                        "identificador": colaborador_id,
                        "tipo": "colaborador_id",
                        "error": str(e)
                    })
            
            # Procesar por cédula
            for cedula in cedulas:
                try:
                    # Buscar colaborador por cédula (cccolaborador)
                    colaborador = Colaboradores.objects.filter(cccolaborador=str(cedula)).first()
                    
                    if not colaborador:
                        no_encontrados.append({
                            "identificador": str(cedula),
                            "tipo": "cedula",
                            "error": "Cédula no encontrada"
                        })
                        continue
                    
                    # Obtener usuario del colaborador
                    usuario = Usuarios.objects.filter(idcolaboradoru=colaborador).first()
                    
                    if not usuario:
                        no_encontrados.append({
                            "identificador": str(cedula),
                            "tipo": "cedula",
                            "error": "Usuario no encontrado para esta cédula"
                        })
                        continue
                    
                    # Actualizar usuario
                    usuario.estadousuario = nuevo_estado
                    usuario.save()
                    
                    # Sincronizar estado en colaborador
                    colaborador.estadocolaborador = nuevo_estado
                    colaborador.save()
                    
                    # Construir nombre completo
                    nombre_completo = f"{colaborador.nombrecolaborador} {colaborador.apellidocolaborador}".strip()
                    
                    encontrados.append({
                        "colaborador_id": colaborador.idcolaborador,
                        "usuario_id": usuario.id,
                        "nombre": nombre_completo,
                        "cedula": str(cedula),
                        "estado": nuevo_estado,
                        "success": True
                    })
                    actualizados += 1
                    
                except Exception as e:
                    no_encontrados.append({
                        "identificador": str(cedula),
                        "tipo": "cedula",
                        "error": str(e)
                    })
            
            return Response({
                "mensaje": f"Procesamiento completado: {actualizados} activados/desactivados, {len(no_encontrados)} no encontrados",
                "total": len(colaborador_ids) + len(cedulas),
                "actualizados": actualizados,
                "no_encontrados": len(no_encontrados),
                "detalles_encontrados": encontrados,
                "detalles_no_encontrados": no_encontrados
            }, status=200)
            
        except Exception as e:
            return Response(
                {"error": f"Error al procesar estados: {str(e)}"},
                status=500
            )


class ActualizarRolUsuarioView(APIView):
    """
    Vista para cambiar el rol de un usuario.
    Solo SuperAdmin puede realizar esta acción.
    """
    permission_classes = [IsAuthenticated, IsSuperAdmin]

    def patch(self, request, colaborador_id):
        """
        PATCH /usuarios/actualizar-rol-usuario/<colaborador_id>/
        Body: { "tipousuario": 0, 1, 2, 3 o 4 }
        
        Tipos de usuario:
        - 0: Usuario Normal/Colaborador
        - 1: Administrador
        - 2: Lectura Admin
        - 3: Usuario Especial
        - 4: Super Admin
        """
        try:
            nuevo_rol = request.data.get('tipousuario')
            
            if nuevo_rol is None:
                return Response(
                    {"error": "El campo 'tipousuario' es requerido"},
                    status=400
                )
            
            if nuevo_rol not in [0, 1, 2, 3, 4]:
                return Response(
                    {"error": "El tipousuario debe ser 0, 1, 2, 3 o 4"},
                    status=400
                )
            
            # Obtener el usuario por su colaborador
            usuario = Usuarios.objects.filter(
                idcolaboradoru__idcolaborador=colaborador_id
            ).first()
            
            if not usuario:
                return Response(
                    {"error": "Usuario no encontrado"},
                    status=404
                )
            
            rol_anterior = usuario.tipousuario
            usuario.tipousuario = nuevo_rol
            usuario.save()
            
            return Response({
                "mensaje": "Rol del usuario actualizado correctamente",
                "usuario_id": usuario.id,
                "colaborador_id": usuario.idcolaboradoru.idcolaborador,
                "rol_anterior": rol_anterior,
                "nuevo_rol": usuario.tipousuario
            }, status=200)
            
        except Exception as e:
            return Response(
                {"error": f"Error al actualizar rol: {str(e)}"},
                status=500
            )


class DatosCargoView(APIView):
    """
    CRUD para gestionar Cargos.
    Administrador (1), Usuario Especial (3) y SuperAdmin (4) pueden manipular.
    """
    permission_classes = [IsAuthenticated, IsSuperUserOrAdmin | IsUsuarioEspecial]

    def check_permission(self, request):
        """Verifica si el usuario tiene permisos para manipular cargos"""
        tipo_usuario = getattr(request.user, 'tipousuario', None)
        # Administrador (1), Usuario Especial (3) o SuperAdmin (4)
        return tipo_usuario in [1, 3, 4]

    def get(self, request):
        """GET /usuarios/Cargo/ - Obtener todos los cargos activos"""
        try:
            cargos = Cargo.objects.filter(estadocargo=1).order_by('idcargo')
            serializer = cargosSerializer(cargos, many=True)
            return Response({
                "count": cargos.count(),
                "results": serializer.data
            }, status=200)
        except Exception as e:
            return Response(
                {"error": f"Error al obtener cargos: {str(e)}"},
                status=500
            )

    def post(self, request):
        """POST /usuarios/Cargo/ - Crear un nuevo cargo"""
        if not self.check_permission(request):
            return Response(
                {"error": "Solo SuperAdmin e IsUsuarioEspecial pueden crear cargos"},
                status=403
            )
        
        try:
            nombre_cargo = request.data.get('nombrecargo')
            
            if not nombre_cargo:
                return Response(
                    {"error": "El campo 'nombrecargo' es requerido"},
                    status=400
                )
            
            # Verificar si el cargo ya existe
            if Cargo.objects.filter(nombrecargo=nombre_cargo).exists():
                return Response(
                    {"error": "El cargo ya existe"},
                    status=400
                )
            
            cargo = Cargo.objects.create(
                nombrecargo=nombre_cargo,
                estadocargo=1
            )
            
            serializer = cargosSerializer(cargo)
            return Response({
                "mensaje": "Cargo creado correctamente",
                "data": serializer.data
            }, status=201)
            
        except Exception as e:
            return Response(
                {"error": f"Error al crear cargo: {str(e)}"},
                status=500
            )

    def put(self, request):
        """PUT /usuarios/Cargo/ - Actualizar un cargo"""
        if not self.check_permission(request):
            return Response(
                {"error": "Solo SuperAdmin e IsUsuarioEspecial pueden actualizar cargos"},
                status=403
            )
        
        try:
            cargo_id = request.data.get('idcargo')
            nombre_cargo = request.data.get('nombrecargo')
            
            if not cargo_id:
                return Response(
                    {"error": "El campo 'idcargo' es requerido"},
                    status=400
                )
            
            if not nombre_cargo:
                return Response(
                    {"error": "El campo 'nombrecargo' es requerido"},
                    status=400
                )
            
            cargo = Cargo.objects.filter(idcargo=cargo_id).first()
            
            if not cargo:
                return Response(
                    {"error": "Cargo no encontrado"},
                    status=404
                )
            
            cargo.nombrecargo = nombre_cargo
            cargo.save()
            
            serializer = cargosSerializer(cargo)
            return Response({
                "mensaje": "Cargo actualizado correctamente",
                "data": serializer.data
            }, status=200)
            
        except Exception as e:
            return Response(
                {"error": f"Error al actualizar cargo: {str(e)}"},
                status=500
            )

    def delete(self, request):
        """DELETE /usuarios/Cargo/ - Eliminar (desactivar) un cargo"""
        if not self.check_permission(request):
            return Response(
                {"error": "Solo SuperAdmin e IsUsuarioEspecial pueden eliminar cargos"},
                status=403
            )
        
        try:
            cargo_id = request.data.get('idcargo')
            
            if not cargo_id:
                return Response(
                    {"error": "El campo 'idcargo' es requerido"},
                    status=400
                )
            
            cargo = Cargo.objects.filter(idcargo=cargo_id).first()
            
            if not cargo:
                return Response(
                    {"error": "Cargo no encontrado"},
                    status=404
                )
            
            cargo.estadocargo = 0
            cargo.save()
            
            return Response({
                "mensaje": "Cargo desactivado correctamente",
                "cargo_id": cargo.idcargo
            }, status=200)
            
        except Exception as e:
            return Response(
                {"error": f"Error al eliminar cargo: {str(e)}"},
                status=500
            )


class DatosNivelView(APIView):
    """
    CRUD para gestionar Niveles.
    Administrador (1), Usuario Especial (3) y SuperAdmin (4) pueden manipular.
    """
    permission_classes = [IsAuthenticated, IsSuperUserOrAdmin | IsUsuarioEspecial]

    def check_permission(self, request):
        """Verifica si el usuario tiene permisos para manipular niveles"""
        tipo_usuario = getattr(request.user, 'tipousuario', None)
        # Administrador (1), Usuario Especial (3) o SuperAdmin (4)
        return tipo_usuario in [1, 3, 4]

    def get(self, request):
        """GET /usuarios/Nivel/ - Obtener todos los niveles activos"""
        try:
            niveles = Niveles.objects.filter(estadonivel=1).order_by('idnivel')
            serializer = nivelesSerializer(niveles, many=True)
            return Response({
                "count": niveles.count(),
                "results": serializer.data
            }, status=200)
        except Exception as e:
            return Response(
                {"error": f"Error al obtener niveles: {str(e)}"},
                status=500
            )

    def post(self, request):
        """POST /usuarios/Nivel/ - Crear un nuevo nivel"""
        if not self.check_permission(request):
            return Response(
                {"error": "Solo SuperAdmin e IsUsuarioEspecial pueden crear niveles"},
                status=403
            )
        
        try:
            nombre_nivel = request.data.get('nombrenivel')
            
            if not nombre_nivel:
                return Response(
                    {"error": "El campo 'nombrenivel' es requerido"},
                    status=400
                )
            
            # Verificar si el nivel ya existe
            if Niveles.objects.filter(nombrenivel=nombre_nivel).exists():
                return Response(
                    {"error": "El nivel ya existe"},
                    status=400
                )
            
            nivel = Niveles.objects.create(
                nombrenivel=nombre_nivel,
                estadonivel=1
            )
            
            serializer = nivelesSerializer(nivel)
            return Response({
                "mensaje": "Nivel creado correctamente",
                "data": serializer.data
            }, status=201)
            
        except Exception as e:
            return Response(
                {"error": f"Error al crear nivel: {str(e)}"},
                status=500
            )

    def put(self, request):
        """PUT /usuarios/Nivel/ - Actualizar un nivel"""
        if not self.check_permission(request):
            return Response(
                {"error": "Solo SuperAdmin e IsUsuarioEspecial pueden actualizar niveles"},
                status=403
            )
        
        try:
            nivel_id = request.data.get('idnivel')
            nombre_nivel = request.data.get('nombrenivel')
            
            if not nivel_id:
                return Response(
                    {"error": "El campo 'idnivel' es requerido"},
                    status=400
                )
            
            if not nombre_nivel:
                return Response(
                    {"error": "El campo 'nombrenivel' es requerido"},
                    status=400
                )
            
            nivel = Niveles.objects.filter(idnivel=nivel_id).first()
            
            if not nivel:
                return Response(
                    {"error": "Nivel no encontrado"},
                    status=404
                )
            
            nivel.nombrenivel = nombre_nivel
            nivel.save()
            
            serializer = nivelesSerializer(nivel)
            return Response({
                "mensaje": "Nivel actualizado correctamente",
                "data": serializer.data
            }, status=200)
            
        except Exception as e:
            return Response(
                {"error": f"Error al actualizar nivel: {str(e)}"},
                status=500
            )

    def delete(self, request):
        """DELETE /usuarios/Nivel/ - Eliminar (desactivar) un nivel"""
        if not self.check_permission(request):
            return Response(
                {"error": "Solo SuperAdmin e IsUsuarioEspecial pueden eliminar niveles"},
                status=403
            )
        
        try:
            nivel_id = request.data.get('idnivel')
            
            if not nivel_id:
                return Response(
                    {"error": "El campo 'idnivel' es requerido"},
                    status=400
                )
            
            nivel = Niveles.objects.filter(idnivel=nivel_id).first()
            
            if not nivel:
                return Response(
                    {"error": "Nivel no encontrado"},
                    status=404
                )
            
            nivel.estadonivel = 0
            nivel.save()
            
            return Response({
                "mensaje": "Nivel desactivado correctamente",
                "nivel_id": nivel.idnivel
            }, status=200)
            
        except Exception as e:
            return Response(
                {"error": f"Error al eliminar nivel: {str(e)}"},
                status=500
            )


class DatosRegionView(APIView):
    """
    CRUD para gestionar Regionales.
    Administrador (1), Usuario Especial (3) y SuperAdmin (4) pueden manipular.
    """
    permission_classes = [IsAuthenticated, IsSuperUserOrAdmin | IsUsuarioEspecial]

    def check_permission(self, request):
        """Verifica si el usuario tiene permisos para manipular regionales"""
        tipo_usuario = getattr(request.user, 'tipousuario', None)
        # Administrador (1), Usuario Especial (3) o SuperAdmin (4)
        return tipo_usuario in [1, 3, 4]

    def get(self, request):
        """GET /usuarios/Region/ - Obtener todas las regionales activas"""
        try:
            regionales = Regional.objects.filter(estadoregional=1).order_by('idregional')
            serializer = regionalesSerializer(regionales, many=True)
            return Response({
                "count": regionales.count(),
                "results": serializer.data
            }, status=200)
        except Exception as e:
            return Response(
                {"error": f"Error al obtener regionales: {str(e)}"},
                status=500
            )

    def post(self, request):
        """POST /usuarios/Region/ - Crear una nueva regional"""
        if not self.check_permission(request):
            return Response(
                {"error": "Solo SuperAdmin e IsUsuarioEspecial pueden crear regionales"},
                status=403
            )
        
        try:
            nombre_regional = request.data.get('nombreregional')
            
            if not nombre_regional:
                return Response(
                    {"error": "El campo 'nombreregional' es requerido"},
                    status=400
                )
            
            # Verificar si la regional ya existe
            if Regional.objects.filter(nombreregional=nombre_regional).exists():
                return Response(
                    {"error": "La regional ya existe"},
                    status=400
                )
            
            regional = Regional.objects.create(
                nombreregional=nombre_regional,
                estadoregional=1
            )
            
            serializer = regionalesSerializer(regional)
            return Response({
                "mensaje": "Regional creada correctamente",
                "data": serializer.data
            }, status=201)
            
        except Exception as e:
            return Response(
                {"error": f"Error al crear regional: {str(e)}"},
                status=500
            )

    def put(self, request):
        """PUT /usuarios/Region/ - Actualizar una regional"""
        if not self.check_permission(request):
            return Response(
                {"error": "Solo SuperAdmin e IsUsuarioEspecial pueden actualizar regionales"},
                status=403
            )
        
        try:
            regional_id = request.data.get('idregional')
            nombre_regional = request.data.get('nombreregional')
            
            if not regional_id:
                return Response(
                    {"error": "El campo 'idregional' es requerido"},
                    status=400
                )
            
            if not nombre_regional:
                return Response(
                    {"error": "El campo 'nombreregional' es requerido"},
                    status=400
                )
            
            regional = Regional.objects.filter(idregional=regional_id).first()
            
            if not regional:
                return Response(
                    {"error": "Regional no encontrada"},
                    status=404
                )
            
            regional.nombreregional = nombre_regional
            regional.save()
            
            serializer = regionalesSerializer(regional)
            return Response({
                "mensaje": "Regional actualizada correctamente",
                "data": serializer.data
            }, status=200)
            
        except Exception as e:
            return Response(
                {"error": f"Error al actualizar regional: {str(e)}"},
                status=500
            )

    def delete(self, request):
        """DELETE /usuarios/Region/ - Eliminar (desactivar) una regional"""
        if not self.check_permission(request):
            return Response(
                {"error": "Solo SuperAdmin e IsUsuarioEspecial pueden eliminar regionales"},
                status=403
            )
        
        try:
            regional_id = request.data.get('idregional')
            
            if not regional_id:
                return Response(
                    {"error": "El campo 'idregional' es requerido"},
                    status=400
                )
            
            regional = Regional.objects.filter(idregional=regional_id).first()
            
            if not regional:
                return Response(
                    {"error": "Regional no encontrada"},
                    status=404
                )
            
            regional.estadoregional = 0
            regional.save()
            
            return Response({
                "mensaje": "Regional desactivada correctamente",
                "regional_id": regional.idregional
            }, status=200)
            
        except Exception as e:
            return Response(
                {"error": f"Error al eliminar regional: {str(e)}"},
                status=500
            )
        

class RegistrarMasivoView(APIView):
    """
    Vista para registrar múltiples usuarios a través de un archivo csv utf-8.
    Solo SuperAdmin y admin puede realizar esta acción.
    
    CSV esperado (separador ;):
    cédula;Nombre;Correo;Número;Región;Nivel;Empresa;Unidad;Descripción Unidad;Proyecto;Centro;Cargo
    
    - La cédula se usa como usuario y contraseña.
    - tipousuario por defecto es 0 (colaborador normal).
    - En la columna Nombre vienen apellidos y nombre juntos:
      las dos primeras palabras son apellidos, las siguientes son nombres.
    - Empresa, Unidad, Descripción Unidad, Proyecto y Centro se usan para filtrar el CentroOp.
      La Descripción Unidad ayuda a identificar la unidad de negocio correcta.
    - Cargo, Nivel y Región se buscan por nombre en la BD.
    """
    permission_classes = [IsAuthenticated, IsSuperUserOrAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def _separar_nombre(self, nombre_completo):
        """
        Separa el nombre completo en apellidos y nombres.
        Las dos primeras palabras son apellidos, el resto son nombres.
        """
        partes = nombre_completo.strip().split()
        if len(partes) >= 3:
            apellidos = ' '.join(partes[:2])
            nombres = ' '.join(partes[2:])
        elif len(partes) == 2:
            apellidos = partes[0]
            nombres = partes[1]
        else:
            apellidos = nombre_completo.strip()
            nombres = ''
        return apellidos, nombres

    def _buscar_centro_op(self, empresa_nombre, unidad_nombre, unidad_descripcion, proyecto_nombre, centro_nombre):
        """
        Busca el CentroOp filtrando por la jerarquía completa:
        Empresa -> Unidad (nombre Y descripción) -> Proyecto -> Centro
        
        Cuando ambos nombre y descripción de unidad están presentes,
        se usa AND para máxima precisión.
        """
        try:
            from django.db.models import Q
            
            # Construir query para buscar la unidad de negocio
            unidad_query = Q(estadounidad=1)
            
            # Si ambas están presentes, usar AND para filtrado estricto
            if unidad_nombre and unidad_descripcion:
                unidad_query = unidad_query & Q(
                    nombreunidad__iexact=unidad_nombre.strip()
                ) & Q(
                    descripcionunidad__iexact=unidad_descripcion.strip()
                )
            elif unidad_nombre:
                unidad_query = unidad_query & Q(nombreunidad__iexact=unidad_nombre.strip())
            elif unidad_descripcion:
                unidad_query = unidad_query & Q(descripcionunidad__iexact=unidad_descripcion.strip())
            
            centro = Centroop.objects.filter(
                nombrecentrop__iexact=centro_nombre.strip(),
                estadocentrop=1,
                id_proyecto__nombreproyecto__iexact=proyecto_nombre.strip(),
                id_proyecto__estadoproyecto=1,
                id_proyecto__id_unidad__id_empresa__nombre_empresa__iexact=empresa_nombre.strip(),
                id_proyecto__id_unidad__id_empresa__estadoempresa=1,
                id_proyecto__id_unidad__in=Unidadnegocio.objects.filter(unidad_query)
            ).first()
            
            return centro
        except Exception:
            return None

    def _validar_jerarquia_centro_op(self, empresa_nombre, unidad_nombre, unidad_descripcion, proyecto_nombre, centro_nombre):
        """
        Valida cada nivel de la jerarquía por separado y retorna errores detallados.
        Retorna (centro_op, error_detallado).
        Si centro_op es None, error_detallado contiene qué falló exactamente.
        """
        from django.db.models import Q
        errores = []

        # 1. Validar Empresa
        empresa = Epresa.objects.filter(
            nombre_empresa__iexact=empresa_nombre.strip(),
            estadoempresa=1
        ).first()
        if not empresa:
            empresas_disponibles = list(
                Epresa.objects.filter(estadoempresa=1)
                .values_list('nombre_empresa', flat=True)[:10]
            )
            return None, f"Empresa '{empresa_nombre}' no encontrada. Disponibles: {', '.join(empresas_disponibles) if empresas_disponibles else 'ninguna'}"

        # 2. Validar Unidad de Negocio (estricto: nombre Y descripción)
        unidad_query = Q(estadounidad=1, id_empresa=empresa)
        if unidad_nombre and unidad_descripcion:
            unidad_query = unidad_query & Q(
                nombreunidad__iexact=unidad_nombre.strip()
            ) & Q(
                descripcionunidad__iexact=unidad_descripcion.strip()
            )
        elif unidad_nombre:
            unidad_query = unidad_query & Q(nombreunidad__iexact=unidad_nombre.strip())
        elif unidad_descripcion:
            unidad_query = unidad_query & Q(descripcionunidad__iexact=unidad_descripcion.strip())

        unidad = Unidadnegocio.objects.filter(unidad_query).first()
        if not unidad:
            unidades_disponibles = list(
                Unidadnegocio.objects.filter(estadounidad=1, id_empresa=empresa)
                .values_list('nombreunidad', 'descripcionunidad')
            )[:10]
            unidades_str = '; '.join([f"{u[0]} ({u[1]})" for u in unidades_disponibles]) if unidades_disponibles else 'ninguna'
            return None, (
                f"Unidad de Negocio no encontrada para Empresa '{empresa_nombre}'. "
                f"Buscando nombre='{unidad_nombre}', descripción='{unidad_descripcion}'. "
                f"Unidades disponibles en esta empresa: {unidades_str}"
            )

        # 3. Validar Proyecto
        proyecto = Proyecto.objects.filter(
            nombreproyecto__iexact=proyecto_nombre.strip(),
            estadoproyecto=1,
            id_unidad=unidad
        ).first()
        if not proyecto:
            proyectos_disponibles = list(
                Proyecto.objects.filter(estadoproyecto=1, id_unidad=unidad)
                .values_list('nombreproyecto', flat=True)[:10]
            )
            return None, (
                f"Proyecto '{proyecto_nombre}' no encontrado para Unidad '{unidad.nombreunidad}' ({unidad.descripcionunidad}). "
                f"Proyectos disponibles: {', '.join(proyectos_disponibles) if proyectos_disponibles else 'ninguno'}"
            )

        # 4. Validar Centro de Operación
        centro = Centroop.objects.filter(
            nombrecentrop__iexact=centro_nombre.strip(),
            estadocentrop=1,
            id_proyecto=proyecto
        ).first()
        if not centro:
            centros_disponibles = list(
                Centroop.objects.filter(estadocentrop=1, id_proyecto=proyecto)
                .values_list('nombrecentrop', flat=True)[:10]
            )
            return None, (
                f"Centro de operación '{centro_nombre}' no encontrado para Proyecto '{proyecto.nombreproyecto}'. "
                f"Centros disponibles: {', '.join(centros_disponibles) if centros_disponibles else 'ninguno'}"
            )

        return centro, None

    def _buscar_cargo(self, nombre_cargo):
        return Cargo.objects.filter(
            nombrecargo__iexact=nombre_cargo.strip(),
            estadocargo=1
        ).first()

    def _buscar_nivel(self, nombre_nivel):
        return Niveles.objects.filter(
            nombrenivel__iexact=nombre_nivel.strip(),
            estadonivel=1
        ).first()

    def _buscar_regional(self, nombre_regional):
        return Regional.objects.filter(
            nombreregional__iexact=nombre_regional.strip(),
            estadoregional=1
        ).first()

    def post(self, request):
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response(
                {"error": "Se requiere un archivo CSV. Envíelo con el campo 'archivo'."},
                status=400
            )

        # Validar extensión
        if not archivo.name.lower().endswith('.csv'):
            return Response(
                {"error": "El archivo debe ser de tipo .csv"},
                status=400
            )

        try:
            # Leer archivo CSV con encoding utf-8
            try:
                contenido = archivo.read().decode('utf-8-sig')
            except UnicodeDecodeError:
                try:
                    archivo.seek(0)
                    contenido = archivo.read().decode('latin-1')
                except Exception:
                    return Response(
                        {"error": "No se pudo leer el archivo. Asegúrese de que esté en formato UTF-8."},
                        status=400
                    )

            # Detectar el delimitador
            primera_linea = contenido.split('\n')[0]
            if ';' in primera_linea:
                delimitador = ';'
            elif ',' in primera_linea:
                delimitador = ','
            else:
                delimitador = ';'

            reader = csv.DictReader(io.StringIO(contenido), delimiter=delimitador)

            # Normalizar nombres de columnas (quitar espacios, tildes, minúsculas)
            if reader.fieldnames is None:
                return Response(
                    {"error": "El archivo CSV está vacío o no tiene encabezados."},
                    status=400
                )

            # Mapeo flexible de columnas
            columnas_normalizadas = {}
            for col in reader.fieldnames:
                col_limpio = col.strip().lower()
                col_limpio = col_limpio.replace('é', 'e').replace('á', 'a').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
                columnas_normalizadas[col.strip()] = col_limpio

            # Mapeo de columnas esperadas - con prioridad de coincidencia exacta
            mapeo_columnas = {
                'cedula': None, 'nombre': None, 'correo': None, 'numero': None,
                'region': None, 'nivel': None, 'empresa': None, 'unidad': None,
                'descripcion': None, 'proyecto': None, 'centro': None, 'cargo': None
            }

            # Primera pasada: buscar coincidencias exactas o muy específicas
            # Priorizar las búsquedas más específicas primero (ej: "descripcion unidad" antes de solo "unidad")
            busquedas_prioritarias = [
                ('descripcion', ['descripcion unidad', 'desc unidad']),
                ('unidad', ['nombreunidad', 'unidad']),
                ('cedula', ['cedula', 'cc', 'id']),
                ('numero', ['numero', 'telefono', 'celular']),
                ('region', ['region', 'regional']),
                ('nivel', ['nivel']),
                ('empresa', ['empresa']),
                ('proyecto', ['proyecto']),
                ('centro', ['centro']),
                ('cargo', ['cargo']),
                ('nombre', ['nombre']),
                ('correo', ['correo', 'email']),
            ]
            
            for esperada, patrones in busquedas_prioritarias:
                for col_original, col_norm in columnas_normalizadas.items():
                    if mapeo_columnas[esperada] is None:
                        # Buscar por coincidencia exacta primero
                        if col_norm in patrones or any(patron == col_norm for patron in patrones):
                            mapeo_columnas[esperada] = col_original
                            break
                        # Si no hay exacta, buscar por substring (pero solo si no fue ya asignada otra columna a este campo)
                        if any(patron in col_norm for patron in patrones):
                            mapeo_columnas[esperada] = col_original
                            break

            # Verificar columnas mínimas requeridas
            requeridas = ['cedula', 'nombre']
            faltantes = [c for c in requeridas if mapeo_columnas[c] is None]
            if faltantes:
                return Response(
                    {"error": f"Faltan columnas requeridas en el CSV: {', '.join(faltantes)}. Columnas encontradas: {', '.join(reader.fieldnames)}"},
                    status=400
                )

            # =====================================================
            # PASO 1: VALIDAR TODAS LAS FILAS (SIN CREAR NADA)
            # =====================================================
            filas_datos = []
            errores_validacion = []
            cedulas_en_csv = {}  # Para detectar duplicados dentro del CSV

            for num_fila, fila in enumerate(reader, start=2):
                # Limpiar valores
                fila_limpia = {k.strip(): (v.strip() if v else '') for k, v in fila.items()}

                cedula = fila_limpia.get(mapeo_columnas.get('cedula', ''), '').strip()
                nombre_completo = fila_limpia.get(mapeo_columnas.get('nombre', ''), '').strip()
                correo = fila_limpia.get(mapeo_columnas.get('correo', ''), '').strip() if mapeo_columnas.get('correo') else ''
                telefono = fila_limpia.get(mapeo_columnas.get('numero', ''), '').strip() if mapeo_columnas.get('numero') else ''
                region_nombre = fila_limpia.get(mapeo_columnas.get('region', ''), '').strip() if mapeo_columnas.get('region') else ''
                nivel_nombre = fila_limpia.get(mapeo_columnas.get('nivel', ''), '').strip() if mapeo_columnas.get('nivel') else ''
                empresa_nombre = fila_limpia.get(mapeo_columnas.get('empresa', ''), '').strip() if mapeo_columnas.get('empresa') else ''
                unidad_nombre = fila_limpia.get(mapeo_columnas.get('unidad', ''), '').strip() if mapeo_columnas.get('unidad') else ''
                unidad_descripcion = fila_limpia.get(mapeo_columnas.get('descripcion', ''), '').strip() if mapeo_columnas.get('descripcion') else ''
                proyecto_nombre = fila_limpia.get(mapeo_columnas.get('proyecto', ''), '').strip() if mapeo_columnas.get('proyecto') else ''
                centro_nombre = fila_limpia.get(mapeo_columnas.get('centro', ''), '').strip() if mapeo_columnas.get('centro') else ''
                cargo_nombre = fila_limpia.get(mapeo_columnas.get('cargo', ''), '').strip() if mapeo_columnas.get('cargo') else ''

                # Saltar filas vacías o incompletas (sin contar como error)
                if not cedula or not nombre_completo:
                    continue

                # Verificar cédula duplicada dentro del mismo CSV
                if cedula in cedulas_en_csv:
                    errores_validacion.append({
                        "fila": num_fila,
                        "cedula": cedula,
                        "error": f"Cédula duplicada en el CSV: {cedula} ya aparece en la fila {cedulas_en_csv[cedula]}"
                    })
                    continue
                cedulas_en_csv[cedula] = num_fila

                # VALIDACIÓN TEMPRANA: Verificar existencia del usuario/colaborador ANTES de todo
                # Verificar si el colaborador ya existe por cédula
                if Colaboradores.objects.filter(cccolaborador=cedula).exists():
                    errores_validacion.append({
                        "fila": num_fila,
                        "cedula": cedula,
                        "error": f"Colaborador ya existe: La cédula {cedula} ya está registrada en la base de datos"
                    })
                    continue

                # Verificar si el usuario ya existe
                if Usuarios.objects.filter(usuario=cedula).exists():
                    errores_validacion.append({
                        "fila": num_fila,
                        "cedula": cedula,
                        "error": f"Usuario ya existe: El usuario con cédula {cedula} ya está registrado"
                    })
                    continue

                # Separar nombre: primeras 2 palabras = apellidos, resto = nombres
                apellidos, nombres = self._separar_nombre(nombre_completo)

                # Validar que los campos de centro_op vengan completos (todos o ninguno)
                campos_centro = [empresa_nombre, unidad_nombre, proyecto_nombre, centro_nombre]
                campos_centro_presentes = [c for c in campos_centro if c]
                if campos_centro_presentes and len(campos_centro_presentes) < 4:
                    faltantes_centro = []
                    if not empresa_nombre: faltantes_centro.append('Empresa')
                    if not unidad_nombre: faltantes_centro.append('Unidad')
                    if not proyecto_nombre: faltantes_centro.append('Proyecto')
                    if not centro_nombre: faltantes_centro.append('Centro')
                    errores_validacion.append({
                        "fila": num_fila,
                        "cedula": cedula,
                        "error": f"Datos incompletos para Centro de Operación. Faltan: {', '.join(faltantes_centro)}. Debe proporcionar todos los campos (Empresa, Unidad, Proyecto, Centro) o ninguno."
                    })
                    continue

                # Buscar CentroOp por jerarquía con validación detallada
                centro_op = None
                if empresa_nombre and unidad_nombre and proyecto_nombre and centro_nombre:
                    centro_op, error_centro = self._validar_jerarquia_centro_op(
                        empresa_nombre, unidad_nombre, unidad_descripcion, proyecto_nombre, centro_nombre
                    )
                    if error_centro:
                        errores_validacion.append({
                            "fila": num_fila,
                            "cedula": cedula,
                            "error": error_centro
                        })
                        continue

                # Buscar Cargo
                cargo_obj = None
                if cargo_nombre:
                    cargo_obj = self._buscar_cargo(cargo_nombre)
                    if not cargo_obj:
                        errores_validacion.append({
                            "fila": num_fila,
                            "cedula": cedula,
                            "error": f"Cargo no encontrado: {cargo_nombre}"
                        })
                        continue

                # Buscar Nivel
                nivel_obj = None
                if nivel_nombre:
                    nivel_obj = self._buscar_nivel(nivel_nombre)
                    if not nivel_obj:
                        errores_validacion.append({
                            "fila": num_fila,
                            "cedula": cedula,
                            "error": f"Nivel no encontrado: {nivel_nombre}"
                        })
                        continue

                # Buscar Regional
                regional_obj = None
                if region_nombre:
                    regional_obj = self._buscar_regional(region_nombre)
                    if not regional_obj:
                        errores_validacion.append({
                            "fila": num_fila,
                            "cedula": cedula,
                            "error": f"Regional no encontrada: {region_nombre}"
                        })
                        continue

                # Si llegamos aquí, la fila es válida
                filas_datos.append({
                    "num_fila": num_fila,
                    "cedula": cedula,
                    "nombre": nombres,
                    "apellido": apellidos,
                    "correo": correo,
                    "telefono": telefono,
                    "centro_op": centro_op,
                    "cargo_obj": cargo_obj,
                    "nivel_obj": nivel_obj,
                    "regional_obj": regional_obj
                })

            # =====================================================
            # Si hay errores de validación, rechazar TODO
            # =====================================================
            if errores_validacion:
                return Response({
                    "error": "Validación fallida. No se creó ningún registro.",
                    "total_errores": len(errores_validacion),
                    "detalles_errores": errores_validacion
                }, status=400)

            # =====================================================
            # PASO 2: CREAR TODOS LOS REGISTROS EN UNA TRANSACCIÓN
            # =====================================================
            if not filas_datos:
                return Response({
                    "error": "El archivo CSV no contiene filas válidas para procesar",
                    "total_filas": 0,
                    "creados": 0
                }, status=400)

            resultados = []
            try:
                with transaction.atomic():
                    for fila_data in filas_datos:
                        try:
                            # Crear Colaborador
                            colaborador = Colaboradores.objects.create(
                                cccolaborador=fila_data['cedula'],
                                nombrecolaborador=fila_data['nombre'],
                                apellidocolaborador=fila_data['apellido'],
                                centroop=fila_data['centro_op'],
                                cargocolaborador=fila_data['cargo_obj'],
                                correocolaborador=fila_data['correo'],
                                telefocolaborador=fila_data['telefono'],
                                nivelcolaborador=fila_data['nivel_obj'],
                                regionalcolab=fila_data['regional_obj'],
                                estadocolaborador=1,
                            )

                            # Crear Usuario (cédula como usuario y contraseña, tipousuario=0)
                            usuario = Usuarios(
                                usuario=fila_data['cedula'],
                                tipousuario=0,
                                idcolaboradoru=colaborador,
                                estadousuario=1,
                            )
                            usuario.set_password(fila_data['cedula'])
                            usuario.save()

                            resultados.append({
                                "fila": fila_data['num_fila'],
                                "cedula": fila_data['cedula'],
                                "nombre": fila_data['nombre'],
                                "apellido": fila_data['apellido'],
                                "usuario_id": usuario.id,
                                "colaborador_id": colaborador.idcolaborador,
                                "success": True
                            })

                        except Exception as e:
                            # Si hay error dentro de la transacción, se hará rollback automático
                            raise Exception(f"Error en fila {fila_data['num_fila']} (cédula {fila_data['cedula']}): {str(e)}")

                # Si llegamos aquí, la transacción fue exitosa
                return Response({
                    "mensaje": f"Todos los {len(resultados)} usuarios fueron registrados exitosamente",
                    "total_creados": len(resultados),
                    "detalles": resultados
                }, status=201)

            except Exception as e:
                # Rollback automático de la transacción
                return Response({
                    "error": f"Error durante la creación de registros. Ningún usuario fue creado. Detalles: {str(e)}",
                    "total_intentados": len(filas_datos),
                    "creados": 0
                }, status=500)

        except Exception as e:
            return Response(
                {"error": f"Error al procesar el archivo CSV: {str(e)}"},
                status=500
            )

    def put(self, request):
        """
        Actualización masiva de usuarios existentes a través de un archivo CSV.
        Busca al colaborador por cédula y actualiza sus datos.
        NO actualiza el estado del colaborador.
        
        Validación estricta: los valores del CSV deben coincidir exactamente
        con los registros existentes en la base de datos.
        
        CSV esperado (mismo formato que POST):
        cédula;Nombre;Correo;Número;Región;Nivel;Empresa;Unidad;Descripción Unidad;Proyecto;Centro;Cargo
        """
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response(
                {"error": "Se requiere un archivo CSV. Envíelo con el campo 'archivo'."},
                status=400
            )

        if not archivo.name.lower().endswith('.csv'):
            return Response(
                {"error": "El archivo debe ser de tipo .csv"},
                status=400
            )

        try:
            # Leer archivo CSV
            try:
                contenido = archivo.read().decode('utf-8-sig')
            except UnicodeDecodeError:
                try:
                    archivo.seek(0)
                    contenido = archivo.read().decode('latin-1')
                except Exception:
                    return Response(
                        {"error": "No se pudo leer el archivo. Asegúrese de que esté en formato UTF-8."},
                        status=400
                    )

            # Detectar delimitador
            primera_linea = contenido.split('\n')[0]
            if ';' in primera_linea:
                delimitador = ';'
            elif ',' in primera_linea:
                delimitador = ','
            else:
                delimitador = ';'

            reader = csv.DictReader(io.StringIO(contenido), delimiter=delimitador)

            if reader.fieldnames is None:
                return Response(
                    {"error": "El archivo CSV está vacío o no tiene encabezados."},
                    status=400
                )

            # Mapeo flexible de columnas (mismo que POST)
            columnas_normalizadas = {}
            for col in reader.fieldnames:
                col_limpio = col.strip().lower()
                col_limpio = col_limpio.replace('é', 'e').replace('á', 'a').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
                columnas_normalizadas[col.strip()] = col_limpio

            mapeo_columnas = {
                'cedula': None, 'nombre': None, 'correo': None, 'numero': None,
                'region': None, 'nivel': None, 'empresa': None, 'unidad': None,
                'descripcion': None, 'proyecto': None, 'centro': None, 'cargo': None
            }

            busquedas_prioritarias = [
                ('descripcion', ['descripcion unidad', 'desc unidad']),
                ('unidad', ['nombreunidad', 'unidad']),
                ('cedula', ['cedula', 'cc', 'id']),
                ('numero', ['numero', 'telefono', 'celular']),
                ('region', ['region', 'regional']),
                ('nivel', ['nivel']),
                ('empresa', ['empresa']),
                ('proyecto', ['proyecto']),
                ('centro', ['centro']),
                ('cargo', ['cargo']),
                ('nombre', ['nombre']),
                ('correo', ['correo', 'email']),
            ]

            for esperada, patrones in busquedas_prioritarias:
                for col_original, col_norm in columnas_normalizadas.items():
                    if mapeo_columnas[esperada] is None:
                        if col_norm in patrones or any(patron == col_norm for patron in patrones):
                            mapeo_columnas[esperada] = col_original
                            break
                        if any(patron in col_norm for patron in patrones):
                            mapeo_columnas[esperada] = col_original
                            break

            # Verificar columna mínima: cédula
            if mapeo_columnas['cedula'] is None:
                return Response(
                    {"error": f"Falta la columna 'cédula' en el CSV. Columnas encontradas: {', '.join(reader.fieldnames)}"},
                    status=400
                )

            # =====================================================
            # PASO 1: VALIDAR TODAS LAS FILAS
            # Separa en: filas_existentes (actualizar) y filas_nuevas (crear)
            # =====================================================
            filas_existentes = []
            filas_nuevas = []
            errores_validacion = []
            cedulas_en_csv = {}  # Para detectar duplicados dentro del CSV

            for num_fila, fila in enumerate(reader, start=2):
                fila_limpia = {k.strip(): (v.strip() if v else '') for k, v in fila.items()}

                cedula = fila_limpia.get(mapeo_columnas.get('cedula', ''), '').strip()
                nombre_completo = fila_limpia.get(mapeo_columnas.get('nombre', ''), '').strip()
                correo = fila_limpia.get(mapeo_columnas.get('correo', ''), '').strip() if mapeo_columnas.get('correo') else ''
                telefono = fila_limpia.get(mapeo_columnas.get('numero', ''), '').strip() if mapeo_columnas.get('numero') else ''
                region_nombre = fila_limpia.get(mapeo_columnas.get('region', ''), '').strip() if mapeo_columnas.get('region') else ''
                nivel_nombre = fila_limpia.get(mapeo_columnas.get('nivel', ''), '').strip() if mapeo_columnas.get('nivel') else ''
                empresa_nombre = fila_limpia.get(mapeo_columnas.get('empresa', ''), '').strip() if mapeo_columnas.get('empresa') else ''
                unidad_nombre = fila_limpia.get(mapeo_columnas.get('unidad', ''), '').strip() if mapeo_columnas.get('unidad') else ''
                unidad_descripcion = fila_limpia.get(mapeo_columnas.get('descripcion', ''), '').strip() if mapeo_columnas.get('descripcion') else ''
                proyecto_nombre = fila_limpia.get(mapeo_columnas.get('proyecto', ''), '').strip() if mapeo_columnas.get('proyecto') else ''
                centro_nombre = fila_limpia.get(mapeo_columnas.get('centro', ''), '').strip() if mapeo_columnas.get('centro') else ''
                cargo_nombre = fila_limpia.get(mapeo_columnas.get('cargo', ''), '').strip() if mapeo_columnas.get('cargo') else ''

                # Saltar filas vacías
                if not cedula:
                    continue

                # Verificar cédula duplicada dentro del mismo CSV
                if cedula in cedulas_en_csv:
                    errores_validacion.append({
                        "fila": num_fila,
                        "cedula": cedula,
                        "error": f"Cédula duplicada en el CSV: {cedula} ya aparece en la fila {cedulas_en_csv[cedula]}"
                    })
                    continue
                cedulas_en_csv[cedula] = num_fila

                # Separar nombre si viene
                apellidos, nombres = None, None
                if nombre_completo:
                    apellidos, nombres = self._separar_nombre(nombre_completo)

                # Verificar si el colaborador existe
                colaborador = Colaboradores.objects.filter(cccolaborador=cedula).first()
                es_nuevo = colaborador is None

                # Para colaboradores NUEVOS, nombre es requerido
                if es_nuevo and not nombre_completo:
                    errores_validacion.append({
                        "fila": num_fila,
                        "cedula": cedula,
                        "error": f"El campo 'Nombre' es obligatorio para crear un nuevo colaborador con cedula {cedula}"
                    })
                    continue

                # Verificar que tiene usuario asociado (solo para existentes)
                usuario = None
                if not es_nuevo:
                    usuario = Usuarios.objects.filter(idcolaboradoru=colaborador).first()
                    if not usuario:
                        errores_validacion.append({
                            "fila": num_fila,
                            "cedula": cedula,
                            "error": f"El colaborador con cedula {cedula} no tiene usuario asociado en el sistema"
                        })
                        continue
                if nombre_completo:
                    apellidos, nombres = self._separar_nombre(nombre_completo)

                # Validar que los campos de centro_op vengan completos (todos o ninguno)
                campos_centro = [empresa_nombre, unidad_nombre, proyecto_nombre, centro_nombre]
                campos_centro_presentes = [c for c in campos_centro if c]
                if campos_centro_presentes and len(campos_centro_presentes) < 4:
                    faltantes_centro = []
                    if not empresa_nombre: faltantes_centro.append('Empresa')
                    if not unidad_nombre: faltantes_centro.append('Unidad')
                    if not proyecto_nombre: faltantes_centro.append('Proyecto')
                    if not centro_nombre: faltantes_centro.append('Centro')
                    errores_validacion.append({
                        "fila": num_fila,
                        "cedula": cedula,
                        "error": f"Datos incompletos para Centro de Operación. Faltan: {', '.join(faltantes_centro)}. Debe proporcionar todos los campos (Empresa, Unidad, Proyecto, Centro) o ninguno."
                    })
                    continue

                # Buscar CentroOp con validación estricta por jerarquía
                centro_op = None
                if empresa_nombre and unidad_nombre and proyecto_nombre and centro_nombre:
                    centro_op, error_centro = self._validar_jerarquia_centro_op(
                        empresa_nombre, unidad_nombre, unidad_descripcion, proyecto_nombre, centro_nombre
                    )
                    if error_centro:
                        errores_validacion.append({
                            "fila": num_fila,
                            "cedula": cedula,
                            "error": error_centro
                        })
                        continue

                # Buscar Cargo (validación estricta)
                cargo_obj = None
                if cargo_nombre:
                    cargo_obj = self._buscar_cargo(cargo_nombre)
                    if not cargo_obj:
                        cargos_disponibles = list(
                            Cargo.objects.filter(estadocargo=1)
                            .values_list('nombrecargo', flat=True)[:15]
                        )
                        errores_validacion.append({
                            "fila": num_fila,
                            "cedula": cedula,
                            "error": f"Cargo '{cargo_nombre}' no encontrado en la BD. Cargos disponibles: {', '.join(cargos_disponibles) if cargos_disponibles else 'ninguno'}"
                        })
                        continue

                # Buscar Nivel (validación estricta)
                nivel_obj = None
                if nivel_nombre:
                    nivel_obj = self._buscar_nivel(nivel_nombre)
                    if not nivel_obj:
                        niveles_disponibles = list(
                            Niveles.objects.filter(estadonivel=1)
                            .values_list('nombrenivel', flat=True)[:15]
                        )
                        errores_validacion.append({
                            "fila": num_fila,
                            "cedula": cedula,
                            "error": f"Nivel '{nivel_nombre}' no encontrado en la BD. Niveles disponibles: {', '.join(niveles_disponibles) if niveles_disponibles else 'ninguno'}"
                        })
                        continue

                # Buscar Regional (validación estricta)
                regional_obj = None
                if region_nombre:
                    regional_obj = self._buscar_regional(region_nombre)
                    if not regional_obj:
                        regionales_disponibles = list(
                            Regional.objects.filter(estadoregional=1)
                            .values_list('nombreregional', flat=True)[:15]
                        )
                        errores_validacion.append({
                            "fila": num_fila,
                            "cedula": cedula,
                            "error": f"Regional '{region_nombre}' no encontrada en la BD. Regionales disponibles: {', '.join(regionales_disponibles) if regionales_disponibles else 'ninguna'}"
                        })
                        continue

                # Agregar a la lista correspondiente (nuevas o existentes)
                fila_obj = {
                    "num_fila": num_fila,
                    "cedula": cedula,
                    "nombre": nombres,
                    "apellido": apellidos,
                    "correo": correo,
                    "telefono": telefono,
                    "centro_op": centro_op,
                    "cargo_obj": cargo_obj,
                    "nivel_obj": nivel_obj,
                    "regional_obj": regional_obj,
                }

                if es_nuevo:
                    filas_nuevas.append(fila_obj)
                else:
                    fila_obj['colaborador'] = colaborador
                    filas_existentes.append(fila_obj)

            # =====================================================
            # Si hay errores de validación, rechazar TODO
            # =====================================================
            if errores_validacion:
                return Response({
                    "error": "Validación fallida. No se actualizó ningún registro.",
                    "total_errores": len(errores_validacion),
                    "detalles_errores": errores_validacion
                }, status=400)

            if not filas_nuevas and not filas_existentes:
                return Response({
                    "error": "El archivo CSV no contiene filas válidas para procesar",
                    "total_filas": 0,
                    "procesados": 0
                }, status=400)

            # =====================================================
            # PASO 2: ACTUALIZAR/CREAR EN UNA TRANSACCIÓN
            # =====================================================
            resultados = []
            try:
                with transaction.atomic():
                    # Procesar colaboradores NUEVOS (crear)
                    for fila_data in filas_nuevas:
                        try:
                            colaborador = Colaboradores.objects.create(
                                cccolaborador=fila_data['cedula'],
                                nombrecolaborador=fila_data['nombre'],
                                apellidocolaborador=fila_data['apellido'],
                                centroop=fila_data['centro_op'],
                                cargocolaborador=fila_data['cargo_obj'],
                                correocolaborador=fila_data['correo'],
                                telefocolaborador=fila_data['telefono'],
                                nivelcolaborador=fila_data['nivel_obj'],
                                regionalcolab=fila_data['regional_obj'],
                                estadocolaborador=1,
                            )

                            usuario = Usuarios(
                                usuario=fila_data['cedula'],
                                tipousuario=0,
                                idcolaboradoru=colaborador,
                                estadousuario=1,
                            )
                            usuario.set_password(fila_data['cedula'])
                            usuario.save()

                            resultados.append({
                                "fila": fila_data['num_fila'],
                                "cedula": fila_data['cedula'],
                                "nombre": colaborador.nombrecolaborador,
                                "apellido": colaborador.apellidocolaborador,
                                "usuario_id": usuario.id,
                                "colaborador_id": colaborador.idcolaborador,
                                "accion": "CREADO",
                                "success": True
                            })

                        except Exception as e:
                            raise Exception(f"Error creando fila {fila_data['num_fila']} (cedula {fila_data['cedula']}): {str(e)}")

                    # Procesar colaboradores EXISTENTES (actualizar)
                    for fila_data in filas_existentes:
                        try:
                            colaborador = fila_data['colaborador']
                            cambios = []

                            # Actualizar nombre y apellido si vienen en el CSV
                            if fila_data['nombre'] and fila_data['nombre'] != colaborador.nombrecolaborador:
                                colaborador.nombrecolaborador = fila_data['nombre']
                                cambios.append('nombre')
                            if fila_data['apellido'] and fila_data['apellido'] != colaborador.apellidocolaborador:
                                colaborador.apellidocolaborador = fila_data['apellido']
                                cambios.append('apellido')

                            # Actualizar correo y teléfono si vienen
                            if fila_data['correo'] and fila_data['correo'] != (colaborador.correocolaborador or ''):
                                colaborador.correocolaborador = fila_data['correo']
                                cambios.append('correo')
                            if fila_data['telefono'] and fila_data['telefono'] != (colaborador.telefocolaborador or ''):
                                colaborador.telefocolaborador = fila_data['telefono']
                                cambios.append('telefono')

                            # Actualizar centro de operación
                            if fila_data['centro_op'] and fila_data['centro_op'] != colaborador.centroop:
                                colaborador.centroop = fila_data['centro_op']
                                cambios.append('centro_operacion')

                            # Actualizar cargo
                            if fila_data['cargo_obj'] and fila_data['cargo_obj'] != colaborador.cargocolaborador:
                                colaborador.cargocolaborador = fila_data['cargo_obj']
                                cambios.append('cargo')

                            # Actualizar nivel
                            if fila_data['nivel_obj'] and fila_data['nivel_obj'] != colaborador.nivelcolaborador:
                                colaborador.nivelcolaborador = fila_data['nivel_obj']
                                cambios.append('nivel')

                            # Actualizar regional
                            if fila_data['regional_obj'] and fila_data['regional_obj'] != colaborador.regionalcolab:
                                colaborador.regionalcolab = fila_data['regional_obj']
                                cambios.append('regional')

                            # NO se actualiza estadocolaborador
                            colaborador.save()

                            resultados.append({
                                "fila": fila_data['num_fila'],
                                "cedula": fila_data['cedula'],
                                "nombre": colaborador.nombrecolaborador,
                                "apellido": colaborador.apellidocolaborador,
                                "colaborador_id": colaborador.idcolaborador,
                                "campos_actualizados": cambios,
                                "accion": "ACTUALIZADO",
                                "success": True
                            })

                        except Exception as e:
                            raise Exception(f"Error en fila {fila_data['num_fila']} (cédula {fila_data['cedula']}): {str(e)}")

                # Contar acciones
                creados = len([r for r in resultados if r.get('accion') == 'CREADO'])
                actualizados = len([r for r in resultados if r.get('accion') == 'ACTUALIZADO'])

                return Response({
                    "mensaje": f"Procesamiento completado: {creados} creados, {actualizados} actualizados",
                    "total_creados": creados,
                    "total_actualizados": actualizados,
                    "total_procesados": len(resultados),
                    "detalles": resultados
                }, status=200)

            except Exception as e:
                return Response({
                    "error": f"Error durante el procesamiento. Ningun usuario fue modificado o creado. Detalles: {str(e)}",
                    "total_intentados": len(filas_nuevas) + len(filas_existentes),
                    "creados": 0,
                    "actualizados": 0
                }, status=500)

        except Exception as e:
            return Response(
                {"error": f"Error al procesar el archivo CSV: {str(e)}"},
                status=500
            )

    def get(self, request):
        """Devuelve la plantilla CSV de ejemplo como descarga."""
        import os
        from django.http import FileResponse
        template_path = os.path.join(
            os.path.dirname(__file__), 'templates', 'Registrar usurios - ejemplo.csv'
        )
        if os.path.exists(template_path):
            return FileResponse(
                open(template_path, 'rb'),
                as_attachment=True,
                filename='plantilla_registro_masivo.csv'
            )
        return Response(
            {"error": "Plantilla no encontrada"},
            status=404
        )


class ReporteUsuariosView(APIView):
    """
    Vista para generar reportes de usuarios en formato Excel con tabla dinámica.
    
    GET: Genera un archivo Excel con todos los colaboradores (activos e inactivos)
    Excluye colaboradores con estado 3 (eliminados/suspendidos)
    La tabla incluye filtros automáticos en los encabezados.
    
    Columnas: cédula, empresa, unidad, proyecto, centro op, nombre, apellido,
    correo, celular, región, nivel, cargo y estado del usuario.
    """
    permission_classes = [IsAuthenticated, IsSuperAdmin, IsAdminUser]

    def get(self, request):
        """
        GET /usuarios/reporte/ - Genera y descarga un archivo Excel con tabla dinámica
        """
        try:
            # Obtener todos los colaboradores EXCEPTO los que están en estado 3
            colaboradores = (
                Colaboradores.objects
                .select_related(
                    'centroop__id_proyecto__id_unidad__id_empresa',
                    'cargocolaborador',
                    'nivelcolaborador',
                    'regionalcolab'
                )
                .prefetch_related(
                    Prefetch(
                        'usuarios_set',
                        queryset=Usuarios.objects.all(),
                        to_attr='usuarios_list'
                    )
                )
                .exclude(estadocolaborador=3)  # Excluir estado 3
                .order_by('idcolaborador')
            )

            # Crear workbook y worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Colaboradores"

            # Definir estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Definir columnas
            columnas = [
                'Cédula',
                'Empresa',
                'Unidad',
                'Descripción Unidad',
                'Proyecto',
                'Centro OP',
                'Nombre',
                'Apellido',
                'Correo',
                'Celular',
                'Región',
                'Nivel',
                'Cargo',
                'Estado Usuario'
            ]

            # Agregar headers
            for col_num, col_titulo in enumerate(columnas, start=1):
                celda = ws.cell(row=1, column=col_num)
                celda.value = col_titulo
                celda.fill = header_fill
                celda.font = header_font
                celda.alignment = header_alignment
                celda.border = border

            # Agregar datos
            for row_num, colaborador in enumerate(colaboradores, start=2):
                # Obtener relaciones
                centro = colaborador.centroop
                proyecto = centro.id_proyecto if centro else None
                unidad = proyecto.id_unidad if proyecto else None
                empresa = unidad.id_empresa if unidad else None

                # Obtener estado del usuario
                usuarios_list = getattr(colaborador, 'usuarios_list', [])
                estado_usuario = usuarios_list[0].estadousuario if usuarios_list else None

                # Datos a insertar
                datos = [
                    colaborador.cccolaborador or '',
                    empresa.nombre_empresa if empresa else '',
                    unidad.nombreunidad if unidad else '',
                    unidad.descripcionunidad if unidad else '',
                    proyecto.nombreproyecto if proyecto else '',
                    centro.nombrecentrop if centro else '',
                    colaborador.nombrecolaborador or '',
                    colaborador.apellidocolaborador or '',
                    colaborador.correocolaborador or '',
                    colaborador.telefocolaborador or '',
                    colaborador.regionalcolab.nombreregional if colaborador.regionalcolab else '',
                    colaborador.nivelcolaborador.nombrenivel if colaborador.nivelcolaborador else '',
                    colaborador.cargocolaborador.nombrecargo if colaborador.cargocolaborador else '',
                    'Activado' if estado_usuario == 1 else ('Desactivado' if estado_usuario == 0 else 'N/A')
                ]

                # Insertar datos en la fila
                for col_num, valor in enumerate(datos, start=1):
                    celda = ws.cell(row=row_num, column=col_num)
                    celda.value = valor
                    celda.alignment = center_alignment
                    celda.border = border

            # Ajustar ancho de columnas
            anchos = [15, 20, 20, 25, 20, 20, 20, 20, 25, 15, 20, 20, 20, 15]
            for col_num, ancho in enumerate(anchos, start=1):
                ws.column_dimensions[chr(64 + col_num)].width = ancho

            # Crear tabla dinámica con filtros automáticos
            # La tabla comienza en A1 y termina en la última columna y fila con datos
            tab = Table(displayName="TablaColaboradores", ref=f"A1:N{max(2, ws.max_row)}")
            
            # Aplicar estilo predefinido a la tabla
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # Guardar en BytesIO
            archivo_excel = BytesIO()
            wb.save(archivo_excel)
            archivo_excel.seek(0)

            # Retornar como descarga
            response = FileResponse(
                archivo_excel,
                as_attachment=True,
                filename='Reporte_Usuarios.xlsx',
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            return response

        except Exception as e:
            return Response(
                {"error": f"Error al generar el reporte: {str(e)}"},
                status=500
            )